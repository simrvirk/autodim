#!/usr/bin/env python3
"""
fai_dimension_numberer.py
=========================
FAI Dimension Numberer — engineering drawing callout annotation and
measurement record population tool.

Workflow:
  1. App loads measurement_template.xlsx from its own directory on startup.
  2. User opens a PDF engineering drawing.
  3. Regex engine detects dimensions; callout circles are overlaid on the drawing.
  4. User adjusts which callouts to include via the right-hand checklist.
  5. User exports an annotated PDF (moveable FreeText annotations) and/or
     a filled copy of the measurement template (formulas intact).
"""

# ---------------------------------------------------------------------------
# Sys-path setup — must run before any local imports so both dev and frozen
# PyInstaller exe can find dimension_detector.py in the same directory.
# ---------------------------------------------------------------------------
import sys
import os

_APP_DIR: str = (
    os.path.dirname(sys.executable)
    if getattr(sys, "frozen", False)
    else os.path.dirname(os.path.abspath(__file__))
)
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

# ---------------------------------------------------------------------------
# Standard library
# ---------------------------------------------------------------------------
import base64
import json
import math
import re
import subprocess
import tempfile
from typing import Any, Callable, Optional

# ---------------------------------------------------------------------------
# PyQt6
# ---------------------------------------------------------------------------
from PyQt6.QtCore import (
    QEvent,
    QPointF,
    QRectF,
    QSize,
    Qt,
    QThread,
    QTimer,
    pyqtSignal,
    pyqtSlot,
)
from PyQt6.QtGui import (
    QBrush,
    QColor,
    QFont,
    QFontMetricsF,
    QImage,
    QPainter,
    QPainterPath,
    QPen,
    QPalette,
    QPixmap,
)
from PyQt6.QtWidgets import (
    QApplication,
    QColorDialog,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGraphicsItem,
    QGraphicsScene,
    QGraphicsView,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSlider,
    QSpinBox,
    QSplitter,
    QStackedWidget,
    QStatusBar,
    QVBoxLayout,
    QWidget,
)

# ---------------------------------------------------------------------------
# Third-party — non-fatal import guards (errors shown in main())
# ---------------------------------------------------------------------------
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]
except ImportError:
    load_workbook = None  # type: ignore[assignment]

# ---------------------------------------------------------------------------
# Local module — dimension_detector.py (same directory)
# ---------------------------------------------------------------------------
_DETECTOR_IMPORT_ERROR: Optional[str] = None
detect_dimensions: Optional[Callable] = None
parse_spreadsheet_template: Optional[Callable] = None

try:
    from dimension_detector import (  # type: ignore[import]
        detect_dimensions,
        parse_spreadsheet_template,
    )
except ImportError as _ie:
    _DETECTOR_IMPORT_ERROR = str(_ie)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TEMPLATE_FILENAME = "MSI-VSA_Template_Dim_Check.xlsx"
CONFIG_FILENAME = "config.json"
SIDECAR_EXT = ".autodim"   # Session sidecar saved alongside every exported PDF
RENDER_ZOOM = 2.0          # Render PDF at 144 DPI (2× 72 pt/inch)
DEFAULT_CALLOUT_RADIUS = 14  # Scene pixels
DEFAULT_CALLOUT_COLOR = "#E74C3C"

# MSI-VSA measurement template fixed column layout
_COL_DIM_NUM    = 2   # Column B — Callout / Balloon number
_COL_NOMINAL    = 3   # Column C — Nominal value
_COL_PLUS_TOL   = 4   # Column D — Plus tolerance (positive)
_COL_MINUS_TOL  = 5   # Column E — Minus tolerance (positive)
_DATA_START_ROW = 16  # First data entry row in the MSI-VSA template

CATEGORY_COLORS: dict[str, str] = {
    "linear":   "#2980B9",
    "angular":  "#8E44AD",
    "GD&T":     "#D35400",
    "thread":   "#27AE60",
    "finish":   "#B7950B",
    "diameter": "#C0392B",
    "radius":   "#C2185B",
    "datum":    "#5D4037",
    "other":    "#7F8C8D",
}

# MSI-VSA measurement template — embedded so no external file is needed
_TEMPLATE_B64 = (
    "UEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAAYAAAAeGwvZHJhd2luZ3MvZHJhd2luZzEueG1s7Vdbc9o4FP4F+x88fg/CYC5h"
    "Ah3STDqZYVsmbLvPQpaxJrKkSoJAfv0eSRaYXnYbpju7D32xj845/nSuX8LNm33Nkx3VhkkxTbNON02oILJgYjNNP/5xfzVO"
    "E2OxKDCXgk7TAzXpm9lvN/tCT57NnU7ge2EmcJymlbVqgpAhFa2x6UhFBVhLqWts4ag3qND4GZBrjnrd7hAZpSkuTEWpvQuW"
    "tMHDF6DVmIn4/Q9FI8uSEXonybamwgYQTTm2UAtTMWUiGrkgGlJhbY8A+68Qaka0NLK0HSLrJpSIAQhZHhDo/oSRvRpkgK7R"
    "+Eug+ofSqbF+2qorwFVQkDXjzB58ZhGm2NQXlKVgeKNxHUH2/Y/i6dVpdZHhoNIRxXCaDV6N0osoMz/NMN5vKedzQSqpg6rU"
    "sg4SkXzW692gKEflh7KcdY9qd/IWLZ9nWVA7Mepa3s0pHE730L1N3LD0s7zfG6QJORxlFDyMCm+xW6llEyd5v1vqpKCGNHNv"
    "oAx4xzYMlrZjoP1USSZCIVxjqXU1WGvYa3Trngu5kQbNP324+n31cMXh1FFuG1kB96eJwDXs/qrCiib9GAnc6kJATRLtgIyX"
    "8GRf6tq9ofgJZDXIrvvDLlCMy2rYGzkZOfsx7W4+dnafdpAd/AlIaWPfUVknTpimmhKbOj3eLYwNrtHFqYW8Z5z7K7g4U6Cg"
    "QadonWT3t7I4OMc1vKGm2E/DWwt1hbDCaZoSC6O3fhBmml5nuesTbx+MIvcMolhgY5dYA5dlaaLbHs8aq2lqPm+hM2liW6YQ"
    "5HxrZcmahEIoPgdjV/bAqU9I+QfEyEQB9OUD5Dvu39DjhRe0DQrMN0Dv3MNDeLe0bKSlNckOe6dw25l1Xtq/8Wus6+3q5WjO"
    "8tjT9fZeCpvYg6IlJjA+c80wj7b3sG5Nv0Ju/mG9Jjy9gorCVfAR0jQvJ3Tk00ftlqH2chDOoCR32OKkXEjyZP5ktlq5vzMx"
    "ge9s/D+TQP9VJND9WSTQG4yy0X9IAvk5CeQ/iwQGwyatb5NALx8NB79I4BcJ/L9IIP82CWSDHsT2LzBBNs5H40GzExnsRP/E"
    "BYqRSAZLRs7ZwG1uN24uq/GGZn6lT8vrP4FFoiXVj/5/3x19pIa90HaRzqDXnCm3QL6HICcEfiBY8Fca6AXGbULrNYWb9UOR"
    "hVYbq6kllRNL+PIRNrYZoGhA58ht/rh03Vtrjc4KddFgoPiLZ/YXUEsHCJ2CEhOZAwAANA0AAFBLAwQUAAgICACkk7ZcAAAA"
    "AAAAAAAAAAAAIwAAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcxLnhtbC5yZWxzjc9BagMxDAXQE/QORvtYM12UUsaTTShk"
    "W5IDCFvjMR3LxnZKc/sYskmgiy4l8d9H0/43buqHSw1JDIx6AMVikwviDZxPn7t3ULWRONqSsIErV9jPL9MXb9R6pq4hV9UR"
    "qQbW1vIHYrUrR6o6ZZZ+WVKJ1PpYPGay3+QZX4fhDcujAfOTqY7OQDm6EdTpmvk/dlqWYPmQ7CWytD8qMMTe3UEqnpsBrTGy"
    "C3TfjzqLB5wnfPptvgFQSwcIuovj97YAAAAjAQAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAAYAAAAeGwvZHJhd2luZ3Mv"
    "ZHJhd2luZzIueG1sndBdbsIwDAfwE+wOVd5pWhgTQxRe0E4wDuAlbhuRj8oOo9x+0Uo2aXsBHm3LP/nvzW50tvhEYhN8I+qy"
    "EgV6FbTxXSMO72+zlSg4gtdgg8dGXJDFbvu0GTWtz7ynIu17XqeyEX2Mw1pKVj064DIM6NO0DeQgppI6qQnOSXZWzqvqRfJA"
    "CJp7xLifJuLqwQOaA+Pz/k3XhLY1CvdBnRz6OCGEFmL6Bfdm4KypB65RPVD8AcZ/gjOKAoc2liq46ynZSEL9PAk4/hr13chS"
    "vsrVX8jdFMcBHU/DLLlDesiHsSZevpNlRnfugbdoAx2By8i4OPjj3bEqyTa1KCtssV7ercyzIrdfUEsHCAdiaYMFAQAABwMA"
    "AFBLAwQUAAgICACkk7ZcAAAAAAAAAAAAAAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbLzd65Ikx5El6CfYd4BAuCND"
    "2W5U2sVvGJIjJFFBIpAA3D1x4/zDgMUmpAGCA4DdnLdft6gwNdOwo2Y2riv7A7ycjJPhGRZ+MivrY/ev/vs/v//unf948+NP"
    "3/7wt1+/a957evedN3/75oc/f/u3f/v1u59/dvnX+d13fvr567/9+evvfvjbm1+/+7/f/PTuf//N//Wr//zhx3//6a9v3vz8"
    "zvEJ/vbTr9/9688///39V69++uavb77/+qf3fvj7m78dH/nLDz9+//XPx3/98d9e/fT3H998/edb6fvvXtmnp/HV919/+7d3"
    "336G93/s+Rw//OUv337z5oMfvvnH92/+9vPbT/Ljm+++/vm4/J/++u3ff4qf7ft/Fp/u+2+/+fGHn374y8/vffPD9/fPdFzB"
    "N6/e/PObN7cLmtkFff9NzxV9//WP//6Pv//r8Sn/flzF//z2u29//t+366JP8x+/fvcfP/7t/fvn+Fe6jNB5/3j+9//j++/i"
    "g/9pfN91Fy/m8mphV/9PM5z7TObplTEPn8p/Xb4W/Zf19Tf0mb7v+zR0Ive3yG9+dfuU64+/+dXfv/63Ny9vfv787+uPr37z"
    "q1eU3/7DF9+++c+fsv/8Tnib/s8ffvj38F8+/POv3316l0r5Yy+3A11/fOebf/z08w/f//HNt//215+P2+Hdd/785i9f/+O7"
    "n3//w3dffvvnn/96ZP497yjff/hPevDw3u2zf/PDdz/d/vX+yWLv3Xe+//Zvb//963/e/v0/7x+x79nl3sQde+/Y1HmiJ8MV"
    "d6+4VPHvzWO14+8dT53lvWGqVoZ7ZcivrPHVjPfOSJ2pdWXTvTJlV9aozPF1Ti/a/N5kqh0TXzUz519P4zUwS2wteatxffYp"
    "HupT9kW13gjx3WPZW+F4Q1Zb8cuyLnuqxmth47vB+v+Tp4pvCJveEW55z/h6K74l7JgdVv0NbuN7wk79B2zju8LO2SvRurp4"
    "vu6p/5lcPClnimd69XYibsvzwdc/f/2bX/34w3++8+OtfF+S6Xj5HrfoeLbwoN8ej/rp9tgj/elI/+M3T7969R/h094f8bu3"
    "j7C3CwzB5TH4+G3gKPjkMfj0MVgfg+0x2B+Dl8fgs8fg88fgi8fgy8fgt/evzmdfv+Ff/29/Xz7EPjzkdXxIOI/j5aczsF1n"
    "YG/tga7qd/bxJX8MPraPL/lj8OljsD4G22OwPwYvj8Fnj8Hnj8EXj8Fv71+MT8nvi+S1FV4/1/X6uVt7zA7IPRzQ71xxEUXy"
    "2gkX4bsuwhcX4R9uJP94qo/Bx/7xVB+DTx+D9THYHoP9MXh5DD57DD5/DL54DL58DH57/+ryV7hIXnvhFR7ur/D4XuUFHm7l"
    "KXuBh4cX+O0jZnrC3z8GHzwGrx+Dy2Pwh8fgj4/Bh4/B9TH46DF4fgw+fgw+eQw+fQzWx2B7DPbH4OUx+Owx+Pwx+OIx+PIx"
    "+OptsFDwp8fgt/dXuTz3sevOGt9+i3pK+3hPTDrpe2Kzd8fI3x0f3B+S3rOv70l6h17uSZriP9yTMZ3/PcnfiBN/qg/vD8ne"
    "EsVTfVQkz8WTf1w8+SfgyWf+5J/eH5Je/7V4qq14qv1tEn58pE+88E/8Unyaz4pP83nxlX9RtL68P1U6va/uSVrCP40P+/I/"
    "7g/JFud+fOW7aup6V01vP2G6rt/dk/TV/P6e5KNuHn48+uD+mCm9re5JNir3JB3JH94mLr2n/3hP8p/EzMOPIh/eH5Nep2vx"
    "XB8VyXPx7B8Xz/4JevaHn3I+vT8m++ZTPNdWPNd+b7GfsB6+Pb8Un+ez4vN8Xjz7F0Xry/tjhuxNMglvkrnrTTK//YTpBvzd"
    "PUnH/ft7Mudf4MO3/g/uj0lfzuu3iU9HcLkn6bb4wz1Jx/3He+Ly53r4Lvjh/THpjX0tnuujInkunv3j4tk/Qc/+sLKf3h+T"
    "jmAtnmsrnmu/J+xGexjVl+LzfFZ8ns/vSTqdL4rWl/dkzt4ks/AmWbreJMvb9pJf/MMo/+7tY4Z0Gb8vkg+K5HWRXIrkD0Xy"
    "xyL5sEiuRfJRkTwXycdF8kmRfFoka5FsRbIXyUuRfFYknxfJF0XyZZF8VSR/uicme0sswlvi+DGk64/VT+BNsTz+yfqpeFeU"
    "0Qdl9LqMLmX0hzL6Yxl9WEbXMvqojJ7L6OMy+qSMPi2jtYy2MtrL6KWMPiujz8voizL6soy+KqM/xSi9Wf5HjGz2/olvgPIN"
    "1Pl7mbe/UhjysbXFr2buD/LZG+htlP2Zp0heF8mlSP5QJH8skg+L5FokHxXJc5F8XCSfFMmnRbIWyVYke5G8FMlnRfJ5kXxR"
    "JF8WyVdF8qd4PEP2VrlHY/5WkX57ZPp+fWTuvz/K/1BgzeNb5f6g7E/HZfRBGb0uo0uMsp9n4TU8/AD5x/JTfVhG1zL6qIye"
    "y+jjMvqkjD4to7WMtjLay6/6BX7VDz/cflZ+qs/L6Isy+rJ8wq/gEz78sPmnsvfbeIrgPdb3KzZz/x1b/udD+/j7l/ig7M/l"
    "ZfRBGb0uo0uMbPYeu0dsEsfH99j9QT57jxXRtYw+KqPnGGV/Io9R/pOqffhJ9ZP7g7I/835aRmsZbWW036PsT8sv8RrYm2B+"
    "fNfdH5T/scQ+/ADyeXzQkr0PQc89fN/5sjycr+7RlH+nvEf5j1VG+iWr6fstq3n7K8QpPfPvYuSyd10RfVBGr8voEiOfvet8"
    "ecs58/iu8+Wy+ce78Ao/1cNIflR+qufyU33cc1WflJ/q0/JTrT1XtZWfai8/1Ut8+dIN81mM0je8z2OU/SmtfNSX5VF8FaP8"
    "W+w9Ylsn/ZrZDH3vsfuvOPOtc+5x6+4Pyn9j4h6W+PfxQfmvJd3Dan4AH/Qwa6/hgx525xIflA+kexiGP8QH+bd/Jnnv6fGt"
    "fP/4cHz8L7fP89evf3zz53ff+fHNX24ffv/D20O+vb1kH17+6y/+8Avjfv1f/tc/fvj5v60//vDNm59+eueTf3z/P9/8+Db7"
    "l7f/dv8vR+G4hF+zjP2X46P/j/nlL3/16i/gb9o+jFc3Pl7d2+t5Bf7y5xo7E/6KP4ofn9+O5OPHn+PHl7cv6OPHP75/fHmC"
    "r9jx4fc/ja+Yvb1izRcgfvXFDX1/JoO+evv2q3/8rV3s2Ern4RVbG6/Y1njF9sYr9hKvKf/LXPrSsvdu8f0KPMg/fPIvej7T"
    "l/FBvvKZvkoHmz3o4Uz+FB805A8q/4JY+isQ0/d3IObtL7uX7DeRMcp+FRmj/Cf6InpdRpcY5T/Rx792yX99EDOTfZ8rsytl"
    "6dv0RyB7psxl39lidj+an38Ubqvx/f322G/D3xOE2+rDl9f7ftxc46vjn1/++rP989f3O+sdurXGf/3F8SrF++udx/srPvcg"
    "PffbZ3sFyp9qyqumvGnKu6b80nlYx+Pef7GTefu5BvG0KD7eaa/+WMbHm+3Vh2V8vI9ePQtH/tsvXu+//cPr8ETvh8f88lyC"
    "3zCf0dc/Vr/+43Hvf0Zf/3j7+o87E30TeCd9n3z7xR2v3asX4Yu7PeiTzz/+3ev9v/7i9qDLb59fwKPCkz2F//Dxb7+iL+q/"
    "fPfzfztuhn95+pc8fRXuj/jQX4T0F/HBv3j76IePvL2jXv3ig/A6/VJ6qT6nl2qqvlTH497/nF6q6fZSvf7/86V6TS/Vh5+k"
    "l+rf4kt1/Our/CO/jA/NXqp/Sy9V+Pd/5R8+XqrX1ZfqC3qp5upLdTzu/S/opZpvL9Vn9FK9I71W8L3w63C9xUOPzxa+mON7"
    "wP99/+Bvv3vz48/oM799IcrC5etvv/vHj2+Kz/nruedzSi/Rl/QSLdWX6Hjc+1/SS7TcXqLP2y8RfA/gl+jz/9OXqCwUL9Hn"
    "/1+8RF/Fl8jkfySPGfszefyL9OwHmPQTIf9ppe/v1s1U/rQylT+tTOVPK0X0uowuMcp/WpnuX1j2l6WUueynlTK7Upb/NqjM"
    "ninLfx80df60Mh0/rUxv34aeff+bju9/k/DTyhS2dRJ/Wpl6vnN74acVRXnVlDdNeaczEL/hyuWX9mG9/fEElT+jcvWZR1j+"
    "nMriN7+33+1Q+Qsqi98O3u4/Kn9JZXEo3y4jKn9Fr3Y+ITFjEyIxC9PnLMxcDsZcDsZcDkYRvS6jS4zywZjBYMQsH4wyu1KW"
    "D0aZPVOWD8bcORjzMRjz/fcaT2wx5mMxZmEx5rAYs7gYc8+t9/a0wWRo2quqvanaOx1E9d4V2i/tI6vNRiyfmo1YPjUbsXxq"
    "NmL51GzElzufjZix2ZDgjemTN2YpZ2MpZ2MpZ6OIXpfRJUb5bCxgNmKWz0aZXSnLZ6PMninLZ2PpnI3lmI3l/n42bDaWYzYW"
    "YTaWMBuLOBtL181nhNnQtFdVe1O1dzqI+mzg9kv7yGqzEcunZiOWT81GLJ+ajVg+NRvx5c5nI2ZsNiScZftwln0qZiNG2WzE"
    "KJuNMnpdRpcYZbNxj9hsUJbNBsiulGWzAbJnyrLZiFlrNuzT+/vtsd/e1Ek2G/bp1fEPng37dMzG8UFhNujJ6zefxbOhaq+q"
    "9qZq73QQ9dnA7Zf2kVVmg8pnZoPKZ2aDymdmg8pnZoNe7mw2KMtnw0okz/aRPGvK2TDlbJhyNorodRldYpTPhgGzEbN8Nsrs"
    "Slk+G2X2TFk+G6ZzNswxG/dfrhn2lzDWHLNhhNkwYTaMOBum6+YT/hZG1V5V7U3V3ukg6rMh/E1M+8hqsxHLp2Yjlk/NRiyf"
    "mo1YPjUb8eXOZyNmbDbE/3Fv5/+615azYcvZsOVsFNHrMrrEKJ8NC2YjZvlslNmVsnw2yuyZsnw2bOds2GM27P39zH4bau0x"
    "G1aYDRtmw4qzYbtuPuHXoar2qmpvqvZOB1GfDeFXou0jq81GLJ+ajVg+NRuxfGo2YvnUbMSXO5+NmLHZkMSt7RO31pWz4crZ"
    "cOVsFNHrMrrEKJ8NB2YjZvlslNmVsnw2yuyZsnw2XOdsuGM23P39zBCBdcdsOGE2XJgNJ86G67r58B3wqaq9qtqbqr3TQdRn"
    "A7df2kdWm41YPjUbsXxqNmL51GzE8qnZiC93PhsxY7MhAWnbB6StL2fDl7Phy9kootdldIlRPhsezEbM8tkosytl+WyU2TNl"
    "+Wz4ztnwx2z4+/t5ZLPhj9nwwmz4MBtenA3fdfPhN/Gnqvaqam+q9k4HUZ8N3H5pH1ltNmL51GzE8qnZiOVTsxHLp2Yjvtz5"
    "bMSMzYZk3m2febdDORtDORtDORtF9LqMLvfImCnbjZjluxGzfDfK7EpZvhtl9kxZvhtD524Mx25E6T6x3RiO3RiE3RjCbgzi"
    "bgxddx9+I36qaq+q9qZq73QQ9d3A7Zf2kdV2I5ZP7UYsn9qNWD61G7F8ajfiy53vRszYbkgu3fa5dFu69BjluxHZWT4cJUwv"
    "ows18+GIWT4cMcuHo8yulOXDUWbPlOXD0SnT7XgMx12mm5kNx3gMh0DTbaDpVqTp9OT12w+/mT5VtVdVe1O1dzqI+nDg9kv7"
    "yGrD0cTdteFocufacDQBcG04mjS2NhwAjVLGhkP6v75k+4ioLYlojPLhiNgsH47SiJbRhZr5cAAkSlk+HACJUpYPB0CilOXD"
    "0YlE7XQMxx2JmoUNx3QMh6BEbVCiVlSi9OT12w+/Hz5VtVdVe1O1dzqI+nDg9kv7yGrDoaGiVD41HBoqSuVTwwGoKGVsOCQq"
    "avuoqC2paIzy4YjcLB+O0oqW0YWa+XAALEpZPhwAi1KWDwfAopTlw9GJRe18DMcdi1qGRe18DIeARW3AolbEovTk1dvPClhU"
    "1V5V7U3V3ukgqnev0H5pH1ltODRYlMqnhkODRal8ajgAFqWMDYeERW0fFrUlFo1RPhwRnOXDUWrRMrpQMx8OwEUpy4cDcFHK"
    "8uEAXJSyfDg6uahdjuG4c1HLuKhdjuEQuKgNXNSKXJSevH77CVxU1V5V7U3V3ukg6sMhcNH2kdWGQ8NFqXxqODRclMqnhgNw"
    "UcrYcEhc1PVxUVdyUVdyURfJWTYcrvSiZXShZjYcDoBRB8AoyK4OgFGQPTsARl0nGHVP7+/uDkYtA6Pu6dXxDx4OF8CoE8Go"
    "62KXVgCjqvaqam+q9u66wKjQfmkfWWU4nAaMOg0YdRow6jRg1AEw6hAYdRIYdX1g1JVg1JVg1EV0lg9HKUbL6ELNfDgAGXWA"
    "jILs6gAZBdmzA2TUdZJRZ47huJNRy8ioM8dwCGTUBTLqRDLquuClFcioqr2q2puqvbsuMiq0X9pHVhsODRl1GjLqNGTUacio"
    "A2TUITLqJDLq+sioK8moK8moi+wsH47SjJbRhZr5cAA06gAaBdnVATQKsmcH0KjrRKPOHsNxR6OWoVFnj+EQ0KgLaNSJaNR1"
    "0UsroFFVe1W1N1V7d11oVGi/tI+sNhwaNOo0aNRp0KjToFEH0KhDaNSJ/4+QOv8/IZVo1JVo1EV4lg9HqUbL6ELNfDgAG3WA"
    "jYLs6gAbBdmzA2zUdbJR547huLNRy9ioc8dwCGzUBTbqRDbquvClFdioqr2q2puqvbsuNiq0X9pHVhsODRt1GjbqNGzUadio"
    "A2zUITbqJDbq+tioK9lojPLhiPQsH47SjZbRhZr5cAA4Slk+HACOUpYPB4CjlOXD0QlHnT+G4w5HLYOjzh/DIcBRF+CoE+Eo"
    "PXn99hPgqKq9qtqbqr3TQdSHQ4Cj7SOrDYcGjlL51HBo4CiVTw0HgKOUseGQ4Kjrg6OuhKMxyocj2rN8OEo5WkYXaubDAeQo"
    "ZflwADlKWT4cQI5Slg9Hpxx1wzEcdzlqmRx1wzEcghx1QY46UY7Sk9dvP0GOqtqrqr2p2jsdRH04BDnaPrLacGjkKJVPDYdG"
    "jlL51HAAOUoZGw5Jjro+OepKORqjfDiAHI0Py4ejlKPUzIcDyFHK8uEAcpSyfDiAHKUsH45OOerGYzjuctQyOerGYzgEOeqC"
    "HHWiHKUnr99+ghxVtVdVe1O1dzqI+nAIcrR9ZLXh0MhRKp8aDo0cpfKp4QBylDI2HJIcdX1y1JVyNEb5cAA5Gh+WD0cpR6mZ"
    "DweQo5TlwwHkKGX5cAA5Slk+HJ1y1E3HcNzlqGVy1E3HcAhy1AU56kQ5Sk9ev/0EOapqr6r2pmrvdBD14RDkaPvIasOhkaNU"
    "PjUcGjlK5VPDAeQoZWw4JDnq+uSoK+VojPLhAHI0PiwfjlKOUjMfDiBHKcuHA8hRyvLhAHKUsnw4OuWom4/huMtRx+Som4/h"
    "EOSoC3LUiXKUnrx6+zlBjqraq6q9qdo7HUT17hXaL+0jqw2HRo5S+dRwaOQolU8NB5CjlLHhkOSo65OjrpSjMcqHA8jR+LB8"
    "OEo5Ss18OIAcpSwfDiBHKcuHA8hRyvLh6JSjbjmG4y5HHZOjbjmGQ5CjLshRJ8pRevL67SfIUVV7VbU3VXung6gPhyBH20dW"
    "Gw6NHKXyqeHQyFEqnxoOIEcpY8MhyVHfJ0d9KUd9KUc9kKO+lKNldPFAjnogRz2QoyC7eiBHQfbsgRz1nXLUP72/+7scdUyO"
    "+qdXxz94OHyQo16Uo77LXzpBjqraq6q9qdq775KjQvulfWSV4fAaOeo1ctRr5KjXyFEP5KhHctRLctT3yVFfylFfylEP5Kgv"
    "5WgZXTyQox7IUQ/kKMiuHshRkD17IEd9pxz15hiOuxx1TI56cwyHIEd9kKNelKO+y186QY6q2quqvanau++So0L7pX1kteHQ"
    "yFGvkaNeI0e9Ro56IEc9kqNekqO+T476Uo76Uo56IEd9KUfL6OKBHPVAjnogR0F29UCOguzZAznqO+Wot8dw3OWoY3LU22M4"
    "BDnqgxz1ohz1Xf7SCXJU1V5V7U3V3n2XHBXaL+0jqw2HRo56jRz1GjnqNXLUAznqkRz1khz1fXLUl3LUl3LUAznqSzlaRhcP"
    "5KgHctQDOQqyqwdyFGTPHshR3ylHvTuG4y5HHZOj3h3DIchRH+SoF+Wo7/KXTpCjqvaqam+q9u675KjQfmkfWW04NHLUa+So"
    "18hRr5GjHshRj+Sol+So75OjvpSjMcqHA8jR+LB8OEo5Ss18OIAcpSwfDiBHKcuHA8hRyvLh6JSj3h/DcZejjslR74/hEOSo"
    "D3LUi3KUnrx++wlyVNVeVe1N1d7pIOrDIcjR9pHVhkMjR6l8ajg0cpTKp4YDyFHK2HBIctT3yVFfytEY5cMB5ChlS7YcZXah"
    "LJ8OYEcpy6cD2FHK8ukAdpSyfDo67agfjum421HH7KgfjukQ7KgPdtSLdpSevH4DCnZU1V5V7U3V3ukg6tMh2NH2kdWmQ2NH"
    "qXxqOjR2lMqnpgPYUcrYdEh21PfZUV/a0Rjl0wHsKGX5dJTZhbJ8OoAepSyfDqBHKcunA+hRyvLp6NSjfjym465HHdOjfjym"
    "Q9CjPuhRL+pRevL6DSjoUVV7VbU3VXung6hPh6BH20dWmw6NHqXyqenQ6FEqn5oOoEcpY9Mh6VHfp0d9qUdjlE8H0KOU5dNR"
    "ZhfK8ukAfpSyfDqAH6Usnw7gRynLp6PTj/rpmI67H3XMj/rpmA7Bj/rgR73oR+nJ6zeg4EdV7VXV3lTtnQ6iPh2CH20fWW06"
    "NH6UyqemQ+NHqXxqOoAfpYxNh+RHfZ8f9aUfjVE+HcCPUpZPR5ldKMunAwhSyvLpAIKUsnw6gCClLJ+OTkHq52M67oLUM0Hq"
    "52M6BEHqgyD1oiClJ6/egF4QpKr2qmpvqvZOB1G9f4X2S/vIatOhEaRUPjUdGkFK5VPTAQQpZWw6JEHq+wSpLwVpjPLpAIKU"
    "snw6yuxCWT4dwJBSlk8HMKSU5dMBDCll+XR0GlK/HNNxN6SeGVK/HNMhGFIfDKkXDSk9ef0GFAypqr2q2puqvdNB1KdDMKTt"
    "I6tNh8aQUvnUdGgMKZVPTQcwpJSx6ZAM6dBnSIfSkA6lIR2AIaUsmw6QXQagSAegSAegSEF2HYAiBdnzABTp0KlIh6f39+Gu"
    "SD1TpMPTq+MfPB1DUKSDqEiHLovpBUWqaq+q9qZq70OXIhXaL+0jq0zHoFGkg0aRDhpFOmgU6QAU6YAU6SAp0qFPkQ6lIh1K"
    "RToARUpZPh1ldhmAIx2AIx2AIwXZdQCOFGTPA3CkQ6cjHcwxHXdH6pkjHcwxHYIjHYIjHURHOnRpTC84UlV7VbU3VXsfuhyp"
    "0H5pH1ltOjSOdNA40kHjSAeNIx2AIx2QIx0kRzr0OdKhdKRD6UgH4Egpy6ejzC4DkKQDkKQDkKQguw5AkoLseQCSdOiUpIM9"
    "puMuST2TpIM9pkOQpEOQpIMoSYcuj+kFSapqr6r2pmrvQ5ckFdov7SOrTYdGkg4aSTpoJOmgkaQDkKQDkqSDJEmHPkk6lJJ0"
    "KCXpACQpZfl0lNllAJZ0AJZ0AJYUZNcBWFKQPQ/Akg6dlnRwx3TcLalnlnRwx3QIlnQIlnQQLenQJTK9YElV7VXV3lTtfeiy"
    "pEL7pX1ktenQWNJBY0kHjSUdNJZ0AJZ0QJZ0kCzp0GdJh9KSxiifDmBJKcuno8wulOXTATQpZfl0AE1KWT4dQJNSlk9HpyYd"
    "/DEdd03qmSYd/DEdgiYdgiYdRE1KT16/AQVNqmqvqvamau90EPXpEDRp+8hq06HRpFQ+NR0aTUrlU9MBNCllbDokTTr0adKh"
    "1KQxyqcDaFLK8ukAmpSyfDqAJqUsnw6gSSnLpwNoUsry6ejUpMNwTMddk3qmSYfhmA5Bkw5Bkw6iJqUnr9+AgiZVtVdVe1O1"
    "dzqI+nQImrR9ZLXp0GhSKp+aDo0mpfKp6QCalDI2HZImHfo06VBq0hjl0wE0KWX5dABNSlk+HUCTUpZPB9CklOXTATQpZfl0"
    "dGrSYTym465JPdOkw3hMh6BJh6BJB1GT0pPXb0BBk6raq6q9qdo7HUR9OgRN2j6y2nRoNCmVT02HRpNS+dR0AE1KGZsOSZMO"
    "fZp0KDVpjPLpAJqUsnw6gCalLJ8OoEkpy6cDaFLK8ukAmpSyfDo6NekwHdNx16SeadJhOqZD0KRD0KSDqEnpyes3oKBJVe1V"
    "1d5U7Z0Ooj4dgiZtH1ltOjSalMqnpkOjSal8ajqAJqWMTYekSYc+TTqUmjRG+XQATUpZPh1Ak1KWTwfQpJTl0wE0KWX5dABN"
    "Slk+HZ2adJiP6bhr0oFp0mE+pkPQpEPQpIOoSenJqzfgIGhSVXtVtTdVe6eDqN6/QvulfWS16dBoUiqfmg6NJqXyqekAmpQy"
    "Nh2SJh36NOlQatIY5dMBNCll+XQATUpZPh1Ak1KWTwfQpJTl0wE0KWX5dHRq0mE5puOuSQemSYflmA5Bkw5Bkw6iJqUnr9+A"
    "giZVtVdVe1O1dzqI+nQImrR9ZLXp0GhSKp+aDo0mpfKp6QCalDI2HZImHfs06Vhq0rHUpCPQpCPQpCC7jECTjkCTjkCTguw6"
    "Ak0KsucRaNKxU5OOT+/v412TDkyTjk+vjn/wdIxBk46iJh27TOYgaFJVe1W1N1V7H7s0qdB+aR9ZZTpGjSYdNZp01GjSUaNJ"
    "R6BJR6RJR0mTjn2adCw16Vhq0hFo0hFoUpBdRqBJR6BJR6BJQXYdgSYF2fMINOnYqUlHc0zHXZMOTJOO5pgOQZOOQZOOoiYd"
    "u0zmIGhSVXtVtTdVex+7NKnQfmkfWW06NJp01GjSUaNJR40mHYEmHZEmHSVNOvZp0rHUpGOpSUegSUegSUF2GYEmHYEmHYEm"
    "Bdl1BJoUZM8j0KRjpyYd7TEdd006ME062mM6BE06Bk06ipp07DKZg6BJVe1V1d5U7X3s0qRC+6V9ZLXp0GjSUaNJR40mHTWa"
    "dASadESadJQ06dinScdSk46lJh2BJh2BJgXZZQSadASadASaFGTXEWhSkD2PQJOOnZp0dMd03DXpwDTp6I7pEDTpGDTpKGrS"
    "sctkDoImVbVXVXtTtfexS5MK7Zf2kdWmQ6NJR40mHTWadNRo0hFo0hFp0lHSpGOfJh1LTRqjfDqAJqUsnw6gSSnLpwNoUsry"
    "6QCalLJ8OoAmpSyfjk5NOvpjOu6adGCadPTHdAiadAyadBQ1KT15/QYUNKmqvaram6q900HUp0PQpO0jq02HRpNS+dR0aDQp"
    "lU9NB9CklLHpkDTp2KdJx1KTxiifDqBJKcunA2hSyvLpAJqUsnw6gCalLJ8OoEkpy6ejU5OOwzEdd006ME06Dsd0CJp0DJp0"
    "FDUpPXn9BhQ0qaq9qtqbqr3TQdSnQ9Ck7SOrTYdGk1L51HRoNCmVT00H0KSUsemQNOnYp0nHUpPGKJ8OoEkpy6cDaFLK8ukA"
    "mpSyfDqAJqUsnw6gSSnLp6NTk47jMR13TTowTTqOx3QImnQMmnQUNSk9ef0GFDSpqr2q2puqvdNB1KdD0KTtI6tNh0aTUvnU"
    "dGg0KZVPTQfQpJSx6ZA06dinScdSk8Yonw6gSSnLpwNoUsry6QCalLJ8OoAmpSyfDqBJKcuno1OTjtMxHXdNOjBNOk7HdAia"
    "dAyadBQ1KT15/QYUNKmqvaram6q900HUp0PQpO0jq02HRpNS+dR0aDQplU9NB9CklLHpkDTp2KdJx1KTxiifDqBJKcunA2hS"
    "yvLpAJqUsnw6gCalLJ8OoEkpy6ejU5OO8zEdd006Mk06zsd0CJp0DJp0FDUpPXn1BhwFTapqr6r2pmrvdBD1+1fQpO0jq02H"
    "RpNS+dR0aDQplU9NB9CklLHpkDTp2KdJx1KTxiifDqBJKcunA2hSyvLpAJqUsnw6gCalLJ8OoEkpy6ejU5OOyzEdd006Mk06"
    "Lsd0CJp0DJp0FDUpPXn9BhQ0qaq9qtqbqr3TQdTvX0GTto+sNh0aTUrlU9Oh0aRUPjUdQJNSxqZD0qRTnyadSk06lZp0App0"
    "ApoUZJcJaNIJaNIJaFKQXSegSUH2PAFNOnVq0unp/X26a9KRadLp6dXxD56OKWjSSdSkU5fJHAVNqmqvqvamau9TlyYV2i/t"
    "I6tMx6TRpJNGk04aTTppNOkENOmENOkkadKpT5NOpSadSk06AU06AU0KsssENOkENOkENCnIrhPQpCB7noAmnTo16WSO6bhr"
    "0pFp0skc0yFo0ilo0knUpFOXyRwFTapqr6r2pmrvU5cmFdov7SOrTYdGk04aTTppNOmk0aQT0KQT0qSTpEmnPk06lZp0KjXp"
    "BDTpBDQpyC4T0KQT0KQT0KQgu05Ak4LseQKadOrUpJM9puOuSUemSSd7TIegSaegSSdRk05dJnMUNKmqvaram6q9T12aVGi/"
    "tI+sNh0aTTppNOmk0aSTRpNOQJNOSJNOkiad+jTpVGrSqdSkE9CkE9CkILtMQJNOQJNOQJOC7DoBTQqy5wlo0qlTk07umI67"
    "Jh2ZJp3cMR2CJp2CJp1ETTp1mcxR0KSq9qpqb6r2PnVpUqH90j6y2nRoNOmk0aSTRpNOGk06AU06IU06SZp06tOkU6lJY5RP"
    "B9CklOXTATQpZfl0AE1KWT4dQJNSlk8H0KSU5dPRqUknf0zHXZOOTJNO/pgOQZNOQZNOoialJ6/fgIImVbVXVXtTtXc6iPr9"
    "K2jS9pHVpkOjSal8ajo0mpTKp6YDaFLK2HRImnTq06RTqUljlE8H0KSU5dMBNCll+XQATUpZPh1Ak1KWTwfQpJTl09GpSafh"
    "mI67Jh2ZJp2GYzoETToFTTqJmpSevH4DCppU1V5V7U3V3ukg6vevoEnbR1abDo0mpfKp6dBoUiqfmg6gSSlj0yFp0qlPk06l"
    "Jo1RPh1Ak1KWTwfQpJTl0wE0KWX5dABNSlk+HUCTUpZPR6cmncZjOu6adGSadBqP6RA06RQ06SRqUnry+g0oaFJVe1W1N1V7"
    "p4Oo37+CJm0fWW06NJqUyqemQ6NJqXxqOoAmpYxNh6RJpz5NOpWaNEb5dABNSlk+HUCTUpZPB9CklOXTATQpZfl0AE1KWT4d"
    "nZp0mo7puGvSkWnSaTqmQ9CkU9Ckk6hJ6cnrN6CgSVXtVdXeVO2dDqJ+/wqatH1ktenQaFIqn5oOjSal8qnpAJqUMjYdkiad"
    "+jTpVGrSGOXTATQpZfl0AE1KWT4dQJNSlk8H0KSU5dMBNCll+XR0atJpPqbjrkknpkmn+ZgOQZNOQZNOoialJ6/egJOgSVXt"
    "VdXeVO2dDqJ6/wrtl/aR1aZDo0mpfGo6NJqUyqemA2hSyth0SJp06tOkU6lJY5RPB9CklOXTATQpZfl0AE1KWT4dQJNSlk8H"
    "0KSU5dPRqUmn5ZiOuyadmCadlmM6BE06BU06iZqUnrx+AwqaVNVeVe1N1d7pIOrTIWjS9pHVpkOjSal8ajo0mpTKp6YDaFLK"
    "2HRImnTu06RzqUnnUpPOQJPOQJOC7DIDTToDTToDTQqy6ww0KcieZ6BJ505NOj+9v893TToxTTo/vTr+wdMxB006i5p07jKZ"
    "k6BJVe1V1d5U7X3u0qRC+6V9ZJXpmDWadNZo0lmjSWeNJp2BJp2RJp0lTTr3adK51KRzqUlnoElnoElBdpmBJp2BJp2BJgXZ"
    "dQaaFGTPM9Ckc6cmnc0xHXdNOjFNOptjOgRNOgdNOouadO4ymZOgSVXtVdXeVO197tKkQvulfWS16dBo0lmjSWeNJp01mnQG"
    "mnRGmnSWNOncp0nnUpPOpSadgSadgSYF2WUGmnQGmnQGmhRk1xloUpA9z0CTzp2adLbHdNw16cQ06WyP6RA06Rw06Sxq0rnL"
    "ZE6CJlW1V1V7U7X3uUuTCu2X9pHVpkOjSWeNJp01mnTWaNIZaNIZadJZ0qRznyadS006l5p0Bpp0BpoUZJcZaNIZaNIZaFKQ"
    "XWegSUH2PANNOndq0tkd03HXpBPTpLM7pkPQpHPQpLOoSecukzkJmlTVXlXtTdXe5y5NKrRf2kdWmw6NJp01mnTWaNJZo0ln"
    "oElnpElnSZPOfZp0LjVpjPLpAJqUsnw6gCalLJ8OoEkpy6cDaFLK8ukAmpSyfDo6Nensj+m4a9KJadLZH9MhaNI5aNJZ1KT0"
    "5PUbUNCkqvaqam+q9k4HUZ8OQZO2j6w2HRpNSuVT06HRpFQ+NR1Ak1LGpkPSpHOfJp1LTRqjfDqAJqUsnw6gSSnLpwNoUsry"
    "6QCalLJ8OoAmpSyfjk5NOg/HdNw16cQ06Twc0yFo0jlo0lnUpPTk9RtQ0KSq9qpqb6r2TgdRnw5Bk7aPrDYdGk1K5VPTodGk"
    "VD41HUCTUsamQ9Kkc58mnUtNGqN8OoAmpSyfDqBJKcunA2hSyvLpAJqUsnw6gCalLJ+OTk06j8d03DXpxDTpPB7TIWjSOWjS"
    "WdSk9OT1G1DQpKr2qmpvqvZOB1GfDkGTto+sNh0aTUrlU9Oh0aRUPjUdQJNSxqZD0qRznyadS00ao3w6gCalLJ8OoEkpy6cD"
    "aFLK8ukAmpSyfDqAJqUsn45OTTpPx3TcNenENOk8HdMhaNI5aNJZ1KT05PUbUNCkqvaqam+q9k4HUZ8OQZO2j6w2HRpNSuVT"
    "06HRpFQ+NR1Ak1LGpkPSpHOfJp1LTRqjfDqAJqUsnw6gSSnLpwNoUsry6QCalLJ8OoAmpSyfjk5NOs/HdNw16cw06Twf0yFo"
    "0jlo0lnUpPTk1RtwFjSpqr2q2puqvdNBVO9fof3SPrLadGg0KZVPTYdGk1L51HQATUoZmw5Jk859mnQuNWmM8ukAmpSyfDqA"
    "JqUsnw6gSSnLpwNoUsry6QCalLJ8Ojo16bwc03HXpDPTpPNyTIegSeegSWdRk9KT129AQZOq2quqvanaOx1EfToETdo+stp0"
    "aDQplU9Nh0aTUvnUdABNShmbDkmTLn2adCk16VJq0gVo0gVoUpBdFqBJF6BJF6BJQXZdgCYF2fMCNOnSqUmXp/f35a5JZ6ZJ"
    "l6dXxz94OpagSRdRky5dJnMWNKmqvaram6q9L12aVGi/tI+sMh2LRpMuGk26aDTpotGkC9CkC9Kki6RJlz5NupSadCk16QI0"
    "6QI0KcguC9CkC9CkC9CkILsuQJOC7HkBmnTp1KSLOabjrklnpkkXc0yHoEmXoEkXUZMuXSZzFjSpqr2q2puqvS9dmlRov7SP"
    "rDYdGk26aDTpotGki0aTLkCTLkiTLpImXfo06VJq0qXUpAvQpAvQpCC7LECTLkCTLkCTguy6AE0KsucFaNKlU5Mu9piOuyad"
    "mSZd7DEdgiZdgiZdRE26dJnMWdCkqvaqam+q9r50aVKh/dI+stp0aDTpotGki0aTLhpNugBNuiBNukiadOnTpEupSZdSky5A"
    "ky5Ak4LssgBNugBNugBNCrLrAjQpyJ4XoEmXTk26uGM67pp0Zpp0ccd0CJp0CZp0ETXp0mUyZ0GTqtqrqr2p2vvSpUmF9kv7"
    "yGrTodGki0aTLhpNumg06QI06YI06SJp0qVPky6lJo1RPh1Ak1KWTwfQpJTl0wE0KWX5dABNSlk+HUCTUpZPR6cmXfwxHXdN"
    "OjNNuvhjOgRNugRNuoialJ68fgMKmlTVXlXtTdXe6SDq0yFo0vaR1aZDo0mpfGo6NJqUyqemA2hSyth0SJp06dOkS6lJY5RP"
    "B9CklOXTATQpZfl0AE1KWT4dQJNSlk8H0KSU5dPRqUmX4ZiOuyadmSZdhmM6BE26BE26iJqUnrx+AwqaVNVeVe1N1d7pIOrT"
    "IWjS9pHVpkOjSal8ajo0mpTKp6YDaFLK2HRImnTp06RLqUljlE8H0KSU5dMBNCll+XQATUpZPh1Ak1KWTwfQpJTl09GpSZfx"
    "mI67Jp2ZJl3GYzoETboETbqImpSevH4DCppU1V5V7U3V3ukg6tMhaNL2kdWmQ6NJqXxqOjSalMqnpgNoUsrYdEiadOnTpEup"
    "SWOUTwfQpJTl0wE0KWX5dABNSlk+HUCTUpZPB9CklOXT0alJl+mYjrsmnZkmXaZjOgRNugRNuoialJ68fgMKmlTVXlXtTdXe"
    "6SDq0yFo0vaR1aZDo0mpfGo6NJqUyqemA2hSyth0SJp06dOkS6lJY5RPB9CklOXTATQpZfl0AE1KWT4dQJNSlk8H0KSU5dPR"
    "qUmX+ZiOuyZdmCZd5mM6BE26BE26iJqUnrx6Ay6CJlW1V1V7U7V3Oojq/Su0X9pHVpsOjSal8qnp0GhSKp+aDqBJKWPTIWnS"
    "pU+TLqUmjVE+HUCTUpZPB9CklOXTATQpZfl0AE1KWT4dQJNSlk9HpyZdlmM67pp0YZp0WY7pEDTpEjTpImpSevL6DShoUlV7"
    "VbU3VXung6hPh6BJ20dWmw6NJqXyqenQaFIqn5oOoEkpY9MhaVLz1MdJb497GA/KsvWIGZuPFGb7gcJLCrMFSWE2ISnMNgSF"
    "1xRmK4LC5xRmO0Jha0iOB76/v330tzcwk03JEb8K/4LH5PjAsSbhw8KcpGuo35UCMdXVV11909X3dCj1TRGYacfxVUYltc+s"
    "SmqfmZXUPrMrqX1mWNKLni1LCvNpiSnalj5ventcsS2lOI3Zw7YAc4rCSwrZtgB2mkK2LQCeppBtC6CnKWTb0olPjweGbbnz"
    "08XxbTFhWwSAenzgti0iQU3XUL89BYOqq6+6+qar7+lQ6tsiONSO46tui0aipva5bdFY1NQ+ty1Ao6aQb4vkUc1TH0i9Pa7Y"
    "lpKkxuxhWwBKReElhWxbgEtNIdsWIFNTyLYF2NQUsm3p1KnHA8O23H3q4vm22LAtglA9PnDbFtGopmuo354CUtXVV11909X3"
    "dCj1bRGgasfxVbdFQ1VT+9y2aLBqap/bFsBVU8i3RQKr5qlPrN4eV2xLaVZj9rAtQK2i8JJCti0ArqaQbQugqylk2wLwagrZ"
    "tnTy1eOBYVvugHUZ+La4sC0CYT0+cNsWEbGma6jfnoJi1dVXXX3T1fd0KPVtESRrx/FVt0VjWVP73LZoNGtqn9sW4FlTyLdF"
    "Eq23l7xrW0rTShnbFqBaU8i2BbjWFLJtAbI1hWxbgG1NIdsWoFtTyLal07ceDwjbcheuy8i3xYdtEYzr8YHbtojKNV1D/fYU"
    "mKuuvurqm66+p0Opb4tAXTuOr7otGuya2ue2RcNdU/vctgDwmkK+LRJ5vR1417aU6JUyti2AvaaQbQuArylk2wLoawrZtgD8"
    "mkK2LYC/ppBtSyeAPR4YtuVOYJeJb8sQtkVAsMcHbtsiMth0DfXbU3Cwuvqqq2+6+p4Opb4tgoXtOL7qtmg0bGqf2xaNh03t"
    "c9sCRGwK+bZIJvb2onVtS6liKWPbAlxsCtm2ABmbQrYtwMamkG0L0LEpZNsCfGwK2bZ0CtnjgWFb7kZ2mfm2jGFbBCV7fOC2"
    "LaKTTddQvz0FKKurr7r6pqvv6VDq2yJg2Y7jq26Lhsum9rlt0YDZ1D63LYDMppBvi4Rmb19217aUbJYyti0AzqaQbQugsylk"
    "2wLwbArZtgA+m0K2LQDQppBtSyehPR4YtuWOaJeFb8sUtkVgtMcHbtsiQtp0DfXbU5C0uvqqq2+6+p4Opb4tgqbtOL7qtmg8"
    "bWqf2xaNqE3tc9sCTG0K+bZIqvZ24V3bUrpayti2AFmbQrYtwNamkG0L0LUpZNsCfG0K2bYAYZtCti2dxvZ4YNiWu7INf+3P"
    "xmUO4yJA2+MDt3ERqW26iOr9GZ5UWBdVf1X2N2V/TydTvcel/kvHIVYXRsNuU/vcwmjgbWqfWxhAb1PIF0bCt7en7lqYkt9S"
    "xhYGANwUsoUBBDeFbGEAwk0hWxjAcFPIFgZA3BSyhemkuMcDw8Is8c1u+MIsYWEEj3t84LYwoshNF9G4QwWTq+yvyv6m7O/p"
    "ZBoLI9jcjkOsLoxG56b2uYXR+NzUPrcwQOimkC+MaHRNp9E1wOgaYHQNMroGGV0QXlKYL4xBRtcgowvCawrzhQHhcwrzhTG9"
    "RtcEo2ue4pudI10TkK6RkK65IV0jI13T5VzDk+KF0fVXZX9T9vd0Mo2FkaRu+xBrC2NUUteopK5RSV2jkroGSV0Dpa4Rpa7p"
    "lLoGSF0DpK5BUtcgqQvCSwrZwiCpa5DUBeE1hWxhkNQ1SOqaXqlrgtQ1Jr7ZOdU1geoaieqaG9U1MtU1Xdo1PKmwMKr+quxv"
    "yv6eTqaxMJLXbR9idWFUXteovK5ReV2j8roGeV0Dva4Rva7p9LoGeF0DvK5BXtcgrwvCSwrZwiCva5DXBeE1hWxhkNc1yOua"
    "Xq9rgtc1Nr7ZOdg1AewaCeyaG9g1Mtg1XeY1PKmwMKr+quxvyv6eTqaxMJLabR9idWFUateo1K5RqV2jUrsGqV0D1a4R1a7p"
    "VLsGqF0D1K5BatcgtQvCSwrZwiC1a5DaBeE1hWxhkNo1SO2aXrVrgto1Lr7ZOds1ge0aie2aG9s1Mts1XfI1fFhYGFV/VfY3"
    "ZX9PJ9NYGMnutg+xujAqu2tUdteo7K5R2V2D7K6BdteIdtd02l0D7G7M2MIgu0shWxhkdylkC4PsLoVsYZDdpZAtDLK7FLKF"
    "6bW7Jthd4+ObneNdE/CukfCuueFdI+NduojGHSrpXV1/VfY3ZX9PJ9NYGEnwtg+xujAqwUvtcwujErzUPrcwSPBSyBdGFLym"
    "U/AaIHhjxhYGCV4K2cIgwUshWxgkeClkC4MEL4VsYZDgpZAtTK/gNUHwmiG+2TnhNYHwGonwmhvhNTLhpYto3KGS4dX1V2V/"
    "U/b3dDKNhZEcb/sQqwujcrzUPrcwKsdL7XMLgxwvhXxhRMdrOh2vAY43ZmxhkOOlkC0McrwUsoVBjpdCtjDI8VLIFgY5XgrZ"
    "wvQ6XhMcrxnjm51DXhMgr5Egr7lBXiNDXrqIxh0qSV5df1X2N2V/TyfTWBhJ87YPsbowKs1L7XMLo9K81D63MEjzUsgXRtS8"
    "plPzGqB5Y8YWBmleCtnCIM1LIVsYpHkpZAuDNC+FbGGQ5qWQLUyv5jVB85opvtk55zWB8xqJ85ob5zUy56WLaNyhkufV9Vdl"
    "f1P293QyjYWRTG/7EKsLozK91D63MCrTS+1zC4NML4V8YUTTazpNrwGmN2ZsYZDppZAtDDK9FLKFQaaXQrYwyPRSyBYGmV4K"
    "2cL0ml4TTK+Jptdw02uC6TWS6TU302tk00sXUb9DjWR6df1V2d+U/T2dTH1hhP5LxyFWF0Zleql9bmFUppfa5xYGmV4K+cKI"
    "ptd0ml4DTG/M2MIg00shWxhkeilkC4NML4VsYZDppZAtDDK9FLKF6TW9JpheE02v4abXBNNrJNNrbqbXyKaXLqJxh0qmV9df"
    "lf1N2d/TyTQWRjK97UOsLozK9FL73MKoTC+1zy0MMr0U8oURTa/tNL0WmF4LTK9Fptci0wvCSwrzhbHI9FpkekF4TWG+MCB8"
    "TmG+MLbX9Npgem00vYabXhtMr5VMr72ZXiubXttnYo1kenX9VdnflP09nUxjYSTT2z7E2sJYlem1KtNrVabXqkyvRabXQtNr"
    "RdNrO02vBabXAtNrkem1yPSC8JJCtjDI9FpkekF4TSFbGGR6LTK9ttf02mB6bTS9hpteG0yvlUyvvZleK5te22dijWR6df1V"
    "2d+U/T2dTGNhJNPbPsTqwqhMr1WZXqsyvVZlei0yvRaaXiuaXttpei0wvRaYXotMr0WmF4SXFLKFQabXItMLwmsK2cIg02uR"
    "6bW9ptcG02uj6TXc9Npgeq1keu3N9FrZ9No+E2sk06vrr8r+puzv6WQaCyOZ3vYhVhdGZXqtyvRalem1KtNrkem10PRa0fTa"
    "TtNrgem1wPRaZHotMr0gvKSQLQwyvRaZXhBeU8gWBplei0yv7TW9NpheG02v4abXBtNrJdNrb6bXyqbX9plYI5leXX9V9jdl"
    "f08n01gYyfS2D7G6MCrTa1Wm16pMr1WZXotMr4Wm14qm13aaXgtMb8zYwiDTSyFbGGR6KWQLg0wvhWxhkOmlkC0MMr0UsoXp"
    "Nb02mF4bTa/hptcG02sl02tvptfKppcuonGHSqZX11+V/U3Z39PJNBZGMr3tQ6wujMr0UvvcwqhML7XPLQwyvRTyhRFNr+00"
    "vRaY3pixhUGml0K2MMj0UsgWBpleCtnCINNLIVsYZHopZAvTa3ptML02ml7DTa8NptdKptfeTK+VTS9dROMOlUyvrr8q+5uy"
    "v6eTaSyMZHrbh1hdGJXppfa5hVGZXmqfWxhkeinkCyOaXttpei0wvTFjC4NML4VsYZDppZAtDDK9FLKFQaaXQrYwyPRSyBam"
    "1/TaYHptNL2Gm14bTK+VTK+9mV4rm166iMYdKpleXX9V9jdlf08n01gYyfS2D7G6MCrTS+1zC6MyvdQ+tzDI9FLIF0Y0vbbT"
    "9FpgemPGFgaZXgrZwiDTSyFbGGR6KWQLg0wvhWxhkOmlkC1Mr+m1wfTaaHoNN702mF4rmV57M71WNr10EY07VDK9uv6q7G/K"
    "/p5OprEwkultH2J1YVSml9rnFkZleql9bmGQ6aWQL4xoem2n6bXA9MaMLQwyvRSyhUGml0K2MMj0UsgWBpleCtnCINNLIVuY"
    "XtNrg+m10fRabnptML1WMr32ZnqtbHrpIup3qJVMr66/Kvubsr+nk6kvjNB/6TjE6sKoTC+1zy2MyvRS+9zCINNLIV8Y0fTa"
    "TtNrgemNGVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2ML2m1wbTa6Pptdz02mB6rWR67c30Wtn00kU07lDJ9Or6q7K/"
    "Kft7OpnGwkimt32I1YVRmV5qn1sYleml9rmFQaaXQr4woul1nabXAdPrgOl1yPQ6ZHpBeElhvjAOmV6HTC8IrynMFwaEzynM"
    "F8b1ml4XTK+Lptdy0+uC6XWS6XU30+tk0+v6TKyVTK+uvyr7m7K/p5NpLIxketuHWFsYpzK9TmV6ncr0OpXpdcj0Omh6nWh6"
    "XafpdcD0OmB6HTK9DpleEF5SyBYGmV6HTC8IrylkC4NMr0Om1/WaXhdMr4um13LT64LpdZLpdTfT62TT6/pMrJVMr66/Kvub"
    "sr+nk2ksjGR624dYXRiV6XUq0+tUptepTK9DptdB0+tE0+s6Ta8DptcB0+uQ6XXI9ILwkkK2MMj0OmR6QXhNIVsYZHodMr2u"
    "1/S6YHpdNL2Wm14XTK+TTK+7mV4nm17XZ2KtZHp1/VXZ35T9PZ1MY2Ek09s+xOrCqEyvU5lepzK9TmV6HTK9DppeJ5pe12l6"
    "HTC9Dpheh0yvQ6YXhJcUsoVBptch0wvCawrZwiDT65Dpdb2m1wXT66Lptdz0umB6nWR63c30Otn0uj4TayXTq+uvyv6m7O/p"
    "ZBoLI5ne9iFWF0Zlep3K9DqV6XUq0+uQ6XXQ9DrR9LpO0+uA6Y0ZWxhkeilkC4NML4VsYZDppZAtDDK9FLKFQaaXQrYwvabX"
    "BdProum13PS6YHqdZHrdzfQ62fTSRTTuUMn06vqrsr8p+3s6mcbCSKa3fYjVhVGZXmqfWxiV6aX2uYVBppdCvjCi6XWdptcB"
    "0xsztjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQLg0wvhWxhek2vC6bXRdNruel1wfQ6yfS6m+l1sumli2jcoZLp1fVXZX9T9vd0"
    "Mo2FkUxv+xCrC6MyvdQ+tzAq00vtcwuDTC+FfGFE0+s6Ta8DpjdmbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgWBpleCtnC9Jpe"
    "F0yvi6bXctPrgul1kul1N9PrZNNLF9G4QyXTq+uvyv6m7O/pZBoLI5ne9iFWF0Zleql9bmFUppfa5xYGmV4K+cKIptd1ml4H"
    "TG/M2MIg00shWxhkeilkC4NML4VsYZDppZAtDDK9FLKF6TW9LpheF02v5abXBdPrJNPrbqbXyaaXLqJxh0qmV9dflf1N2d/T"
    "yTQWRjK97UOsLozK9FL73MKoTC+1zy0MMr0U8oURTa/rNL0OmN6YsYVBppdCtjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQL02t6"
    "XTC9Lppex02vC6bXSabX3Uyvk00vXUT9DnWS6dX1V2V/U/b3dDL1hRH6Lx2HWF0Yleml9rmFUZleap9bGGR6KeQLI5pe12l6"
    "HTC9MWMLg0wvhWxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFqbX9Lpgel00vY6bXhdMr5NMr7uZXiebXrqIxh0qmV5df1X2N2V/"
    "TyfTWBjJ9LYPsbowKtNL7XMLozK91D63MMj0UsgXRjS9vtP0emB6PTC9Hplej0wvCC8pzBfGI9PrkekF4TWF+cKA8DmF+cL4"
    "XtPrg+n10fQ6bnp9ML1eMr3+Znq9bHp9n4l1kunV9Vdlf1P293QyjYWRTG/7EGsL41Wm16tMr1eZXq8yvR6ZXg9NrxdNr+80"
    "vR6YXg9Mr0em1yPTC8JLCtnCINPrkekF4TWFbGGQ6fXI9Ppe0+uD6fXR9Dpuen0wvV4yvf5mer1sen2fiXWS6dX1V2V/U/b3"
    "dDKNhZFMb/sQqwujMr1eZXq9yvR6len1yPR6aHq9aHp9p+n1wPR6YHo9Mr0emV4QXlLIFgaZXo9MLwivKWQLg0yvR6bX95pe"
    "H0yvj6bXcdPrg+n1kun1N9PrZdPr+0ysk0yvrr8q+5uyv6eTaSyMZHrbh1hdGJXp9SrT61Wm16tMr0em10PT60XT6ztNrwem"
    "1wPT65Hp9cj0gvCSQrYwyPR6ZHpBeE0hWxhkej0yvb7X9Ppgen00vY6bXh9Mr5dMr7+ZXi+bXt9nYp1kenX9VdnflP09nUxj"
    "YSTT2z7E6sKoTK9XmV6vMr1eZXo9Mr0eml4vml7faXo9ML0xYwuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgWptf0+mB6"
    "fTS9jpteH0yvl0yvv5leL5teuojGHSqZXl1/VfY3ZX9PJ9NYGMn0tg+xujAq00vtcwujMr3UPrcwyPRSyBdGNL2+0/R6YHpj"
    "xhYGmV4K2cIg00shWxhkeilkC4NML4VsYZDppZAtTK/p9cH0+mh6HTe9PpheL5lefzO9Xja9dBGNO1Qyvbr+quxvyv6eTqax"
    "MJLpbR9idWFUppfa5xZGZXqpfW5hkOmlkC+MaHp9p+n1wPTGjC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwiDTSyFbmF7T64Pp"
    "9dH0Om56fTC9XjK9/mZ6vWx66SIad6hkenX9VdnflP09nUxjYSTT2z7E6sKoTC+1zy2MyvRS+9zCINNLIV8Y0fT6TtPrgemN"
    "GVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2ML2m1wfT66Ppddz0+mB6vWR6/c30etn00kU07lDJ9Or6q7K/Kft7OpnG"
    "wkimt32I1YVRmV5qn1sYleml9rmFQaaXQr4woun1nabXA9MbM7YwyPRSyBYGmV4K2cIg00shWxhkeilkC4NML4VsYXpNrw+m"
    "10fT67np9cH0esn0+pvp9bLppYuo36FeMr26/qrsb8r+nk6mvjBC/6XjEKsLozK91D63MCrTS+1zC4NML4V8YUTT6ztNrwem"
    "N2ZsYZDppZAtDDK9FLKFQaaXQrYwyPRSyBYGmV4K2cL0ml4fTK+Pptdz0+uD6fWS6fU30+tl00sX0bhDJdOr66/K/qbs7+lk"
    "Ggsjmd72IVYXRmV6qX1uYVSml9rnFgaZXgr5woimd+g0vQMwvQMwvQMyvQMyvSC8pDBfmAGZ3gGZXhBeU5gvDAifU5gvzNBr"
    "eodgeodoej03vUMwvYNkeoeb6R1k0zv0mVgvmV5df1X2N2V/TyfTWBjJ9LYPsbYwg8r0DirTO6hM76AyvQMyvQM0vYNoeodO"
    "0zsA0zsA0zsg0zsg0wvCSwrZwiDTOyDTC8JrCtnCINM7INM79JreIZjeIZpez03vEEzvIJne4WZ6B9n0Dn0m1kumV9dflf1N"
    "2d/TyTQWRjK97UOsLozK9A4q0zuoTO+gMr0DMr0DNL2DaHqHTtM7ANM7ANM7INM7INMLwksK2cIg0zsg0wvCawrZwiDTOyDT"
    "O/Sa3iGY3iGaXs9N7xBM7yCZ3uFmegfZ9A59JtZLplfXX5X9Tdnf08k0FkYyve1DrC6MyvQOKtM7qEzvoDK9AzK9AzS9g2h6"
    "h07TOwDTOwDTOyDTOyDTC8JLCtnCINM7INMLwmsK2cIg0zsg0zv0mt4hmN4hml7PTe8QTO8gmd7hZnoH2fQOfSbWS6ZX11+V"
    "/U3Z39PJNBZGMr3tQ6wujMr0DirTO6hM76AyvQMyvQM0vYNoeodO0zsA0xsztjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQLg0wv"
    "hWxhek3vEEzvEE2v56Z3CKZ3kEzvcDO9g2x66SIad6hkenX9VdnflP09nUxjYSTT2z7E6sKoTC+1zy2MyvRS+9zCINNLIV8Y"
    "0fQOnaZ3AKY3ZmxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwvSa3iGY3iGaXs9N7xBM7yCZ3uFmegfZ9NJFNO5QyfTq"
    "+quyvyn7ezqZxsJIprd9iNWFUZleap9bGJXppfa5hUGml0K+MKLpHTpN7wBMb8zYwiDTSyFbGGR6KWQLg0wvhWxhkOmlkC0M"
    "Mr0UsoXpNb1DML1DNL2em94hmN5BMr3DzfQOsumli2jcoZLp1fVXZX9T9vd0Mo2FkUxv+xCrC6MyvdQ+tzAq00vtcwuDTC+F"
    "fGFE0zt0mt4BmN6YsYVBppdCtjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQL02t6h2B6h2h6PTe9QzC9g2R6h5vpHWTTSxfRuEMl"
    "06vrr8r+puzv6WQaCyOZ3vYhVhdGZXqpfW5hVKaX2ucWBpleCvnCiKZ36DS9AzC9MWMLg0wvhWxhkOmlkC0MMr0UsoVBppdC"
    "tjDI9FLIFqbX9A7B9A7R9A7c9A7B9A6S6R1upneQTS9dRP0OHSTTq+uvyv6m7O/pZOoLI/RfOg6xujAq00vtcwujMr3UPrcw"
    "yPRSyBdGNL1Dp+kdgOmNGVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2ML2mdwimd4imd+Cmdwimd5BM73AzvYNseuki"
    "GneoZHp1/VXZ35T9PZ1MY2Ek09s+xOrCqEwvtc8tjMr0UvvcwiDTSyFfGNH0jp2mdwSmdwSmd0Smd0SmF4SXFOYLMyLTOyLT"
    "C8JrCvOFAeFzCvOFGXtN7xhM7xhN78BN7xhM7yiZ3vFmekfZ9I59JnaQTK+uvyr7m7K/p5NpLIxketuHWFuYUWV6R5XpHVWm"
    "d1SZ3hGZ3hGa3lE0vWOn6R2B6R2B6R2R6R2R6QXhJYVsYZDpHZHpBeE1hWxhkOkdkekde03vGEzvGE3vwE3vGEzvKJne8WZ6"
    "R9n0jn0mdpBMr66/Kvubsr+nk2ksjGR624dYXRiV6R1VpndUmd5RZXpHZHpHaHpH0fSOnaZ3BKZ3BKZ3RKZ3RKYXhJcUsoVB"
    "pndEpheE1xSyhUGmd0Smd+w1vWMwvWM0vQM3vWMwvaNkeseb6R1l0zv2mdhBMr26/qrsb8r+nk6msTCS6W0fYnVhVKZ3VJne"
    "UWV6R5XpHZHpHaHpHUXTO3aa3hGY3hGY3hGZ3hGZXhBeUsgWBpneEZleEF5TyBYGmd4Rmd6x1/SOwfSO0fQO3PSOwfSOkukd"
    "b6Z3lE3v2GdiB8n06vqrsr8p+3s6mcbCSKa3fYjVhVGZ3lFlekeV6R1VpndEpneEpncUTe/YaXpHYHpjxhYGmV4K2cIg00sh"
    "WxhkeilkC4NML4VsYZDppZAtTK/pHYPpHaPpHbjpHYPpHSXTO95M7yibXrqIxh0qmV5df1X2N2V/TyfTWBjJ9LYPsbowKtNL"
    "7XMLozK91D63MMj0UsgXRjS9Y6fpHYHpjRlbGGR6KWQLg0wvhWxhkOmlkC0MMr0UsoVBppdCtjC9pncMpneMpnfgpncMpneU"
    "TO94M72jbHrpIhp3qGR6df1V2d+U/T2dTGNhJNPbPsTqwqhML7XPLYzK9FL73MIg00shXxjR9I6dpncEpjdmbGGQ6aWQLQwy"
    "vRSyhUGml0K2MMj0UsgWBpleCtnC9JreMZjeMZregZveMZjeUTK94830jrLppYto3KGS6dX1V2V/U/b3dDKNhZFMb/sQqwuj"
    "Mr3UPrcwKtNL7XMLg0wvhXxhRNM7dpreEZjemLGFQaaXQrYwyPRSyBYGmV4K2cIg00shWxhkeilkC9Nresdgesdoegduesdg"
    "ekfJ9I430zvKppcuonGHSqZX11+V/U3Z39PJNBZGMr3tQ6wujMr0UvvcwqhML7XPLQwyvRTyhRFN79hpekdgemPGFgaZXgrZ"
    "wiDTSyFbGGR6KWQLg0wvhWxhkOmlkC1Mr+kdg+kdo+kduekdg+kdJdM73kzvKJteuoj6HTpKplfXX5X9Tdnf08nUF0bov3Qc"
    "YnVhVKaX2ucWRmV6qX1uYZDppZAvjGh6x07TOwLTGzO2MMj0UsgWBpleCtnCINNLIVsYZHopZAuDTC+FbGF6Te8YTO8YTe/I"
    "Te8YTO8omd7xZnpH2fTSRTTuUMn06vqrsr8p+3s6mcbCSKa3fYjVhVGZXmqfWxiV6aX2uYVBppdCvjCi6Z06Te8ETO8ETO+E"
    "TO+ETC8ILynMF2ZCpndCpheE1xTmCwPC5xTmCzP1mt4pmN4pmt6Rm94pmN5JMr3TzfROsumd+kzsKJleXX9V9jdlf08n01gY"
    "yfS2D7G2MJPK9E4q0zupTO+kMr0TMr0TNL2TaHqnTtM7AdM7AdM7IdM7IdMLwksK2cIg0zsh0wvCawrZwiDTOyHTO/Wa3imY"
    "3ima3pGb3imY3kkyvdPN9E6y6Z36TOwomV5df1X2N2V/TyfTWBjJ9LYPsbowKtM7qUzvpDK9k8r0Tsj0TtD0TqLpnTpN7wRM"
    "7wRM74RM74RMLwgvKWQLg0zvhEwvCK8pZAuDTO+ETO/Ua3qnYHqnaHpHbnqnYHonyfRON9M7yaZ36jOxo2R6df1V2d+U/T2d"
    "TGNhJNPbPsTqwqhM76QyvZPK9E4q0zsh0ztB0zuJpnfqNL0TML0TML0TMr0TMr0gvKSQLQwyvRMyvSC8ppAtDDK9EzK9U6/p"
    "nYLpnaLpHbnpnYLpnSTTO91M7ySb3qnPxI6S6dX1V2V/U/b3dDKNhZFMb/sQqwujMr2TyvROKtM7qUzvhEzvBE3vJJreqdP0"
    "TsD0xowtDDK9FLKFQaaXQrYwyPRSyBYGmV4K2cIg00shW5he0zsF0ztF0zty0zsF0ztJpne6md5JNr10EY07VDK9uv6q7G/K"
    "/p5OprEwkultH2J1YVSml9rnFkZleql9bmGQ6aWQL4xoeqdO0zsB0xsztjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQLg0wvhWxh"
    "ek3vFEzvFE3vyE3vFEzvJJne6WZ6J9n00kU07lDJ9Or6q7K/Kft7OpnGwkimt32I1YVRmV5qn1sYleml9rmFQaaXQr4woumd"
    "Ok3vBExvzNjCINNLIVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhek1vVMwvVM0vSM3vVMwvZNkeqeb6Z1k00sX0bhDJdOr66/K"
    "/qbs7+lkGgsjmd72IVYXRmV6qX1uYVSml9rnFgaZXgr5woimd+o0vRMwvTFjC4NML4VsYZDppZAtDDK9FLKFQaaXQrYwyPRS"
    "yBam1/ROwfRO0fSO3PROwfROkumdbqZ3kk0vXUTjDpVMr66/Kvubsr+nk2ksjGR624dYXRiV6aX2uYVRmV5qn1sYZHop5Asj"
    "mt6p0/ROwPTGjC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwiDTSyFbmF7TOwXTO0XTO3HTOwXTO0mmd7qZ3kk2vXQR9Tt0kkyv"
    "rr8q+5uyv6eTqS+M0H/pOMTqwqhML7XPLYzK9FL73MIg00shXxjR9E6dpncCpjdmbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgW"
    "BpleCtnC9JreKZjeKZreiZveKZjeSTK90830TrLppYto3KGS6dX1V2V/U/b3dDKNhZFMb/sQqwujMr3UPrcwKtNL7XMLg0wv"
    "hXxhRNM7d5reGZjeGZjeGZneGZleEF5SmC/MjEzvjEwvCK8pzBcGhM8pzBdm7jW9czC9czS9Eze9czC9s2R655vpnWXTO/eZ"
    "2Ekyvbr+quxvyv6eTqaxMJLpbR9ibWFmlemdVaZ3VpneWWV6Z2R6Z2h6Z9H0zp2mdwamdwamd0amd0amF4SXFLKFQaZ3RqYX"
    "hNcUsoVBpndGpnfuNb1zML1zNL0TN71zML2zZHrnm+mdZdM795nYSTK9uv6q7G/K/p5OprEwkultH2J1YVSmd1aZ3lllemeV"
    "6Z2R6Z2h6Z1F0zt3mt4ZmN4ZmN4Zmd4ZmV4QXlLIFgaZ3hmZXhBeU8gWBpneGZneudf0zsH0ztH0Ttz0zsH0zpLpnW+md5ZN"
    "79xnYifJ9Or6q7K/Kft7OpnGwkimt32I1YVRmd5ZZXpnlemdVaZ3RqZ3hqZ3Fk3v3Gl6Z2B6Z2B6Z2R6Z2R6QXhJIVsYZHpn"
    "ZHpBeE0hWxhkemdkeude0zsH0ztH0ztx0zsH0ztLpne+md5ZNr1zn4mdJNOr66/K/qbs7+lkGgsjmd72IVYXRmV6Z5XpnVWm"
    "d1aZ3hmZ3hma3lk0vXOn6Z2B6Y0ZWxhkeilkC4NML4VsYZDppZAtDDK9FLKFQaaXQrYwvaZ3DqZ3jqZ34qZ3DqZ3lkzvfDO9"
    "s2x66SIad6hkenX9VdnflP09nUxjYSTT2z7E6sKoTC+1zy2MyvRS+9zCINNLIV8Y0fTOnaZ3BqY3ZmxhkOmlkC0MMr0UsoVB"
    "ppdCtjDI9FLIFgaZXgrZwvSa3jmY3jma3omb3jmY3lkyvfPN9M6y6aWLaNyhkunV9Vdlf1P293QyjYWRTG/7EKsLozK91D63"
    "MCrTS+1zC4NML4V8YUTTO3ea3hmY3pixhUGml0K2MMj0UsgWBpleCtnCINNLIVsYZHopZAvTa3rnYHrnaHonbnrnYHpnyfTO"
    "N9M7y6aXLqJxh0qmV9dflf1N2d/TyTQWRjK97UOsLozK9FL73MKoTC+1zy0MMr0U8oURTe/caXpnYHpjxhYGmV4K2cIg00sh"
    "WxhkeilkC4NML4VsYZDppZAtTK/pnYPpnaPpnbjpnYPpnSXTO99M7yybXrqIxh0qmV5df1X2N2V/TyfTWBjJ9LYPsbowKtNL"
    "7XMLozK91D63MMj0UsgXRjS9c6fpnYHpjRlbGGR6KWQLg0wvhWxhkOmlkC0MMr0UsoVBppdCtjC9pncOpneOpnfmpncOpneW"
    "TO98M72zbHrpIup36CyZXl1/VfY3ZX9PJ1NfGKH/0nGI1YVRmV5qn1sYleml9rmFQaaXQr4woumdO03vDExvzNjCINNLIVsY"
    "ZHopZAuDTC+FbGGQ6aWQLQwyvRSyhek1vXMwvXM0vTM3vXMwvbNkeueb6Z1l00sX0bhDJdOr66/K/qbs7+lkGgsjmd72IVYX"
    "RmV6qX1uYVSml9rnFgaZXgr5woimd+k0vQswvQswvQsyvQsyvSC8pDBfmAWZ3gWZXhBeU5gvDAifU5gvzNJrepdgepdoemdu"
    "epdgehfJ9C4307vIpnfpM7GzZHp1/VXZ35T9PZ1MY2Ek09s+xNrCLCrTu6hM76IyvYvK9C7I9C7Q9C6i6V06Te8CTO8CTO+C"
    "TO+CTC8ILylkC4NM74JMLwivKWQLg0zvgkzv0mt6l2B6l2h6Z256l2B6F8n0LjfTu8imd+kzsbNkenX9VdnflP09nUxjYSTT"
    "2z7E6sKoTO+iMr2LyvQuKtO7INO7QNO7iKZ36TS9CzC9CzC9CzK9CzK9ILykkC0MMr0LMr0gvKaQLQwyvQsyvUuv6V2C6V2i"
    "6Z256V2C6V0k07vcTO8im96lz8TOkunV9Vdlf1P293QyjYWRTG/7EKsLozK9i8r0LirTu6hM74JM7wJN7yKa3qXT9C7A9C7A"
    "9C7I9C7I9ILwkkK2MMj0Lsj0gvCaQrYwyPQuyPQuvaZ3CaZ3iaZ35qZ3CaZ3kUzvcjO9i2x6lz4TO0umV9dflf1N2d/TyTQW"
    "RjK97UOsLozK9C4q07uoTO+iMr0LMr0LNL2LaHqXTtO7ANMbM7YwyPRSyBYGmV4K2cIg00shWxhkeilkC4NML4VsYXpN7xJM"
    "7xJN78xN7xJM7yKZ3uVmehfZ9NJFNO5QyfTq+quyvyn7ezqZxsJIprd9iNWFUZleap9bGJXppfa5hUGml0K+MKLpXTpN7wJM"
    "b8zYwiDTSyFbGGR6KWQLg0wvhWxhkOmlkC0MMr0UsoXpNb1LML1LNL0zN71LML2LZHqXm+ldZNNLF9G4QyXTq+uvyv6m7O/p"
    "ZBoLI5ne9iFWF0Zleql9bmFUppfa5xYGmV4K+cKIpnfpNL0LML0xYwuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgWptf0"
    "LsH0LtH0ztz0LsH0LpLpXW6md5FNL11E4w6VTK+uvyr7m7K/p5NpLIxketuHWF0Yleml9rmFUZleap9bGGR6KeQLI5repdP0"
    "LsD0xowtDDK9FLKFQaaXQrYwyPRSyBYGmV4K2cIg00shW5he07sE07tE0ztz07sE07tIpne5md5FNr10EY07VDK9uv6q7G/K"
    "/p5OprEwkultH2J1YVSml9rnFkZleql9bmGQ6aWQL4xoepdO07sA0xsztjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQLg0wvhWxh"
    "ek3vEkzvEk3vwk3vEkzvIpne5WZ6F9n00kXU79BFMr26/qrsb8r+nk6mvjBC/6XjEKsLozK91D63MCrTS+1zC4NML4V8YUTT"
    "u3Sa3gWY3pixhUGml0K2MMj0UsgWBpleCtnCINNLIVsYZHopZAvTa3qXYHqXaHoXbnqXYHoXyfQuN9O7yKaXLqJxh0qmV9df"
    "lf1N2d/TyTQWRjK97UOsLozK9FL73MKoTC+1zy0MMr0U8oWRTK996jO9t8c9LAxl2cLEjC1MCrOFQeElhdnCpDBbmBRmC4PC"
    "awqzhUHhcwqzhaGwtTDHA9/f3z769mZnpvfIX4V/wQtzfOBYmPBhYWHSRTTuUMH0Kvursr8p+3s6mcbCCKa34xArC5PaZxYm"
    "tc8sTGqfWZjUPrMw6VXPFiaF+cLEFC1Mn+m9Pa5YmNL0xuxhYYDpReElhWxhgOlNIVsYYHpTyBYGmN4UsoXpNL3HA8PCRNO7"
    "OL4wJiyMYHqPD9wWRjS96SIad6hgepX9VdnflP09nUxjYQTT23GI1YXRmN7UPrcwGtOb2ucWBpjeFPKFkUyvfeozvbfHFQtT"
    "mt6YPSwMML0ovKSQLQwwvSlkCwNMbwrZwgDTm0K2MJ2m93hgWJhoehfPF8aGhRFM7/GB28KIpjddROMOFUyvsr8q+5uyv6eT"
    "aSyMYHo7DrG6MBrTm9rnFkZjelP73MIA05tCvjCS6bVPfab39rhiYUrTG7OHhQGmF4WXFLKFAaY3hWxhgOlNIVsYYHpTyBam"
    "0/QeDwwLE03vMvCFcWFhBNN7fOC2MKLpTRfRuEMF06vsr8r+puzv6WQaCyOY3o5DrC6MxvSm9rmF0Zje1D63MMD0ppAvjGR6"
    "7VOf6b09rliY0vTG7GFhgOlF4SWFbGGA6U0hWxhgelPIFgaY3hSyhek0vccDw8JE07uMfGF8WBjB9B4fuC2MaHrTRTTuUMH0"
    "Kvursr8p+3s6mcbCCKa34xCrC6Mxval9bmE0pje1zy0MML0p5AsjmV771Gd6b48rFqY0vTF7WBhgelF4SSFbGGB6U8gWBpje"
    "FLKFAaY3hWxhOk3v8cCwMNH0LhNfmCEsjGB6jw/cFkY0vekiGneoYHqV/VXZ35T9PZ1MY2EE09txiNWF0Zje1D63MBrTm9rn"
    "FgaY3hTyhZFMr33qM723xxULU5remD0sDDC9KLykkC0MML0pZAsDTG8K2cIA05tCtjCdpvd4YFiYaHqXmS/MGBZGML3HB24L"
    "I5redBGNO1Qwvcr+quxvyv6eTqaxMILp7TjE6sJoTG9qn1sYjelN7XMLA0xvCvnCSKbXPvWZ3tvjioUpTW/MHhYGmF4UXlLI"
    "FgaY3hSyhQGmN4VsYYDpTSFbmE7TezwwLEw0vcvCF2YKCyOY3uMDt4URTW+6iMYdKpheZX9V9jdlf08n01gYwfR2HGJ1YTSm"
    "N7XPLYzG9Kb2uYUBpjeFfGEk02uf+kzv7XHFwpSmN2YPCwNMLwovKWQLA0xvCtnCANObQrYwwPSmkC1Mp+k9HhgW5m56AxBg"
    "CzOHhRFM7/GB28KIpjddRPUODU8qLIyqvyr7m7K/p5Op3uNS/6XjEKsLozG9qX1uYTSmN7XPLQwwvSnkCyOZXvvUZ3pvjysW"
    "pjS9MXtYGGB6UXhJIVsYYHpTyBYGmN4UsoUBpjeFbGE6Te/xwLAwS3yzG74wS1gYwfQeH7gtjGh600U07lDB9Cr7q7K/Kft7"
    "OpnGwgimt+MQqwujMb2pfW5hNKY3tc8tDDC9KeQLI5pe02l6DTC9Bpheg0yvQaYXhJcU5gtjkOk1yPSC8JrCfGFA+JzCfGFM"
    "r+k1wfSap/hm56bXBNNrJNNrbqbXyKbXdJnY8KR4YXT9VdnflP09nUxjYSTT2z7E2sIYlek1KtNrVKbXqEyvQabXQNNrRNNr"
    "Ok2vAabXANNrkOk1yPSC8JJCtjDI9BpkekF4TSFbGGR6DTK9ptf0mmB6jYlvdm56TTC9RjK95mZ6jWx6TZeJDU8qLIyqvyr7"
    "m7K/p5NpLIxketuHWF0Ylek1KtNrVKbXqEyvQabXQNNrRNNrOk2vAabXANNrkOk1yPSC8JJCtjDI9BpkekF4TSFbGGR6DTK9"
    "ptf0mmB6jY1vdm56TTC9RjK95mZ6jWx6TZeJDU8qLIyqvyr7m7K/p5NpLIxketuHWF0Ylek1KtNrVKbXqEyvQabXQNNrRNNr"
    "Ok2vAabXANNrkOk1yPSC8JJCtjDI9BpkekF4TSFbGGR6DTK9ptf0mmB6jYtvdm56TTC9RjK95mZ6jWx6TZeJDU8qLIyqvyr7"
    "m7K/p5NpLIxketuHWF0Ylek1KtNrVKbXqEyvQabXQNNrRNNrOk2vAaY3ZmxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZ"
    "wvSaXhNMr/Hxzc5Nrwmm10im19xMr5FNL11E4w6VTK+uvyr7m7K/p5NpLIxketuHWF0Yleml9rmFUZleap9bGGR6KeQLI5pe"
    "02l6DTC9MWMLg0wvhWxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFqbX9Jpges0Q3+zc9Jpgeo1kes3N9BrZ9NJFNO5QyfTq+quy"
    "vyn7ezqZxsJIprd9iNWFUZleap9bGJXppfa5hUGml0K+MKLpNZ2m1wDTGzO2MMj0UsgWBpleCtnCINNLIVsYZHopZAuDTC+F"
    "bGF6Ta8JpteM8c3OTa8JptdIptfcTK+RTS9dROMOlUyvrr8q+5uyv6eTaSyMZHrbh1hdGJXppfa5hVGZXmqfWxhkeinkCyOa"
    "XtNpeg0wvTFjC4NML4VsYZDppZAtDDK9FLKFQaaXQrYwyPRSyBam1/SaYHrNFN/s3PSaYHqNZHrNzfQa2fTSRTTuUMn06vqr"
    "sr8p+3s6mcbCSKa3fYjVhVGZXmqfWxiV6aX2uYVBppdCvjCi6TWdptcA0xsztjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQLg0wv"
    "hWxhek2vCabXRNNruOk1wfQayfSam+k1sumli6jfoUYyvbr+quxvyv6eTqa+MEL/peMQqwujMr3UPrcwKtNL7XMLg0wvhXxh"
    "RNNrOk2vAaY3ZmxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwvSaXhNMr4mm13DTa4LpNZLpNTfTa2TTSxfRuEMl06vr"
    "r8r+puzv6WQaCyOZ3vYhVhdGZXqpfW5hVKaX2ucWBpleCvnCiKbXdppeC0yvBabXItNrkekF4SWF+cJYZHotMr0gvKYwXxgQ"
    "PqcwXxjba3ptML02ml7DTa8NptdKptfeTK+VTa/tM7FGMr26/qrsb8r+nk6msTCS6W0fYm1hrMr0WpXptSrTa1Wm1yLTa6Hp"
    "taLptZ2m1wLTa4Hptcj0WmR6QXhJIVsYZHotMr0gvKaQLQwyvRaZXttrem0wvTaaXsNNrw2m10qm195Mr5VNr+0zsUYyvbr+"
    "quxvyv6eTqaxMJLpbR9idWFUpteqTK9VmV6rMr0WmV4LTa8VTa/tNL0WmF4LTK9Fptci0wvCSwrZwiDTa5HpBeE1hWxhkOm1"
    "yPTaXtNrg+m10fQabnptML1WMr32ZnqtbHptn4k1kunV9Vdlf1P293QyjYWRTG/7EKsLozK9VmV6rcr0WpXptcj0Wmh6rWh6"
    "bafptcD0WmB6LTK9FpleEF5SyBYGmV6LTC8IrylkC4NMr0Wm1/aaXhtMr42m13DTa4PptZLptTfTa2XTa/tMrJFMr66/Kvub"
    "sr+nk2ksjGR624dYXRiV6bUq02tVpteqTK9FptdC02tF02s7Ta8FpjdmbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgWBpleCtnC"
    "9JpeG0yvjabXcNNrg+m1kum1N9NrZdNLF9G4QyXTq+uvyv6m7O/pZBoLI5ne9iFWF0Zleql9bmFUppfa5xYGmV4K+cKIptd2"
    "ml4LTG/M2MIg00shWxhkeilkC4NML4VsYZDppZAtDDK9FLKF6TW9NpheG02v4abXBtNrJdNrb6bXyqaXLqJxh0qmV9dflf1N"
    "2d/TyTQWRjK97UOsLozK9FL73MKoTC+1zy0MMr0U8oURTa/tNL0WmN6YsYVBppdCtjDI9FLIFgaZXgrZwiDTSyFbGGR6KWQL"
    "02t6bTC9Nppew02vDabXSqbX3kyvlU0vXUTjDpVMr66/Kvubsr+nk2ksjGR624dYXRiV6aX2uYVRmV5qn1sYZHop5Asjml7b"
    "aXotML0xYwuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2MMj0UsgWptf02mB6bTS9hpteG0yvlUyvvZleK5teuojGHSqZXl1/VfY3"
    "ZX9PJ9NYGMn0tg+xujAq00vtcwujMr3UPrcwyPRSyBdGNL220/RaYHpjxhYGmV4K2cIg00shWxhkeilkC4NML4VsYZDppZAt"
    "TK/ptcH02mh6LTe9NpheK5leezO9Vja9dBH1O9RKplfXX5X9Tdnf08nUF0bov3QcYnVhVKaX2ucWRmV6qX1uYZDppZAvjGh6"
    "bafptcD0xowtDDK9FLKFQaaXQrYwyPRSyBYGmV4K2cIg00shW5he02uD6bXR9Fpuem0wvVYyvfZmeq1seukiGneoZHp1/VXZ"
    "35T9PZ1MY2Ek09s+xOrCqEwvtc8tjMr0UvvcwiDTSyFfGNH0uk7T64DpdcD0OmR6HTK9ILykMF8Yh0yvQ6YXhNcU5gsDwucU"
    "5gvjek2vC6bXRdNruel1wfQ6yfS6m+l1sul1fSbWSqZX11+V/U3Z39PJNBZGMr3tQ6wtjFOZXqcyvU5lep3K9Dpkeh00vU40"
    "va7T9Dpgeh0wvQ6ZXodMLwgvKWQLg0yvQ6YXhNcUsoVBptch0+t6Ta8LptdF02u56XXB9DrJ9Lqb6XWy6XV9JtZKplfXX5X9"
    "Tdnf08k0FkYyve1DrC6MyvQ6lel1KtPrVKbXIdProOl1oul1nabXAdPrgOl1yPQ6ZHpBeEkhWxhkeh0yvSC8ppAtDDK9Dple"
    "12t6XTC9Lppey02vC6bXSabX3Uyvk02v6zOxVjK9uv6q7G/K/p5OprEwkultH2J1YVSm16lMr1OZXqcyvQ6ZXgdNrxNNr+s0"
    "vQ6YXgdMr0Om1yHTC8JLCtnCINPrkOkF4TWFbGGQ6XXI9Lpe0+uC6XXR9Fpuel0wvU4yve5mep1sel2fibWS6dX1V2V/U/b3"
    "dDKNhZFMb/sQqwujMr1OZXqdyvQ6lel1yPQ6aHqdaHpdp+l1wPTGjC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwiDTSyFbmF7T"
    "64LpddH0Wm56XTC9TjK97mZ6nWx66SIad6hkenX9VdnflP09nUxjYSTT2z7E6sKoTC+1zy2MyvRS+9zCINNLIV8Y0fS6TtPr"
    "gOmNGVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhUGml0K2ML2m1wXT66Lptdz0umB6nWR63c30Otn00kU07lDJ9Or6q7K/Kft7"
    "OpnGwkimt32I1YVRmV5qn1sYleml9rmFQaaXQr4woul1nabXAdMbM7YwyPRSyBYGmV4K2cIg00shWxhkeilkC4NML4VsYXpN"
    "rwum10XTa7npdcH0Osn0upvpdbLppYto3KGS6dX1V2V/U/b3dDKNhZFMb/sQqwujMr3UPrcwKtNL7XMLg0wvhXxhRNPrOk2v"
    "A6Y3ZmxhkOmlkC0MMr0UsoVBppdCtjDI9FLIFgaZXgrZwvSaXhdMr4um13LT64LpdZLpdTfT62TTSxfRuEMl06vrr8r+puzv"
    "6WQaCyOZ3vYhVhdGZXqpfW5hVKaX2ucWBpleCvnCiKbXdZpeB0xvzNjCINNLIVsYZHopZAuDTC+FbGGQ6aWQLQwyvRSyhek1"
    "vS6YXhdNr+Om1wXT6yTT626m18mmly6ifoc6yfTq+quyvyn7ezqZ+sII/ZeOQ6wujMr0UvvcwqhML7XPLQwyvRTyhRFNr+s0"
    "vQ6Y3pixhUGml0K2MMj0UsgWBpleCtnCINNLIVsYZHopZAvTa3pdML0uml7HTa8LptdJptfdTK+TTS9dROMOlUyvrr8q+5uy"
    "v6eTaSyMZHrbh1hdGJXppfa5hVGZXmqfWxhkeinkCyOaXt9pej0wvR6YXo9Mr0emF4SXFOYL45Hp9cj0gvCawnxhQPicwnxh"
    "fK/p9cH0+mh6HTe9PpheL5lefzO9Xja9vs/EOsn06vqrsr8p+3s6mcbCSKa3fYi1hfEq0+tVpterTK9XmV6PTK+HpteLptd3"
    "ml4PTK8Hptcj0+uR6QXhJYVsYZDp9cj0gvCaQrYwyPR6ZHp9r+n1wfT6aHodN70+mF4vmV5/M71eNr2+z8Q6yfTq+quyvyn7"
    "ezqZxsJIprd9iNWFUZlerzK9XmV6vcr0emR6PTS9XjS9vtP0emB6PTC9Hplej0wvCC8pZAuDTK9HpheE1xSyhUGm1yPT63tN"
    "rw+m10fT67jp9cH0esn0+pvp9bLp9X0m1kmmV9dflf1N2d/TyTQWRjK97UOsLozK9HqV6fUq0+tVptcj0+uh6fWi6fWdptcD"
    "0+uB6fXI9HpkekF4SSFbGGR6PTK9ILymkC0MMr0emV7fa3p9ML0+ml7HTa8PptdLptffTK+XTa/vM7FOMr26/qrsb8r+nk6m"
    "sTCS6W0fYnVhVKbXq0yvV5lerzK9HpleD02vF02v7zS9HpjemLGFQab3/23ubpskua0zDf8VhkLh8L5ZU1UJIHMsKYKWNAox"
    "mNt4y8Tq46w4pBkec6jhKGT/+y0UiexC30Dncadi7U+Srk6QxZNsiM+Zhz0bVjdMq9O7YXXDtDq9G1Y3TKvTu2F1w7Q6vRtW"
    "N4y00zvkTu9QOr2XutM75E7v0Ov0DrdO79Dv9G4fYuc7tNfpPXbeHjzvDp73j29m54bpdXr3X+KzN8yhTu92+mU3zKFO73b6"
    "ZTdMq9O7YX3DdDu9g7DTOzQ6vcWqG6bV6d2wumFand4Nqxum1endsLphWp3eDasbptXp3bC6YaSd3iF3eofS6b3Und4hd3qH"
    "Xqd3uHV6h36nd/sQO9+hvU7vsfP24Hl38Lx/fDM7N0yv07v/Ep+9YQ51erfTL7thDnV6t9Mvu2Fand4N6xum2+kdhJ3eodHp"
    "LVbdMK1O74bVDdPq9G5Y3TCtTu+G1Q3T6vRuWN0wrU7vhtUNI+30DrnTO5RO76Xu9A650zv0Or3DrdM79Du924fY+Q7tdXqP"
    "nbcHz7uD5/3jm9m5YXqd3v2X+OwNc6jTu51+2Q1zqNO7nX7ZDdPq9G5Y3zDdTu8g7PQOjU5vseqGaXV6N6xumFand8Pqhml1"
    "ejesbphWp3fD6oZpdXo3rG4Yaad3yJ3eoXR6L3Wnd8id3qHX6R1und6h3+ndPsTOd2iv03vsvD143h087x/fzM4N0+v07r/E"
    "Z2+YQ53e7fTLbphDnd7t9MtumFand8P6hul2egdhp3dodHqLVTdMq9O7YXXDtDq9G1Y3TKvTu2F1w7Q6vRtWN0yr07thdcNI"
    "O71D7vQOpdM71J3eIXd6h16nd7h1eod+p3f7EM9/hw69Tu+x8/bgeXfwvH98M8/fMJ3zQfASn71hDnV6t9Mvu2EOdXq30y+7"
    "YVqd3g3rG6bb6R2End6h0ektVt0wrU7vhtUN0+r0bljdMK1O74bVDdPq9G5Y3TCtTu+G1Q0j7fQOudM7lE7vUHd6h9zpHXqd"
    "3uHW6R36nd7tQ+x8h/Y6vcfO24Pn3cHz/vHN7NwwvU7v/kt89oY51OndTr/shjnU6d1Ov+yGaXV6N6xvmG6nVwk7varR6VWN"
    "Tq9qdXpVq9PbwDePeH/DqFanV7U6vQ384hHvb5gGfvmI9zeMknZ6Ve70qtLpHepOr8qdXtXr9Kpbp1f1O71K1okdep3eY+ft"
    "wfPu4Hn/+GZ2bphep3f/JT53w6hDnV51qNOrDnV61aFOr2p1elWz06u6nV4l7PSqRqdXNTq9qtXpVa1ObwPfPGJ1w7Q6varV"
    "6W3gF49Y3TCtTq9qdXqVtNOrcqdXlU7vUHd6Ve70ql6nV906varf6VWyTuzQ6/QeO28PnncHz/vHN7Nzw/Q6vfsv8dkb5lCn"
    "Vx3q9KpDnV51qNOrWp1e1ez0qm6nVwk7varR6VWNTq9qdXpVq9PbwDePWN0wrU6vanV6G/jFI1Y3TKvTq1qdXiXt9Krc6VWl"
    "0zvUnV6VO72q1+lVt06v6nd6lawTO/Q6vcfO24Pn3cHz/vHN7NwwvU7v/kt89oY51OlVhzq96lCnVx3q9KpWp1c1O72q2+lV"
    "wk6vanR6VaPTq1qdXtXq9DbwzSNWN0yr06tand4GfvGI1Q3T6vSqVqdXSTu9Knd6Ven0DnWnV+VOr+p1etWt06v6nV4l68QO"
    "vU7vsfP24Hl38Lx/fDM7N0yv07v/Ep+9YQ51etWhTq861OlVhzq9qtXpVc1Or+p2epWw06sand5i1Q3T6vRuWN0wrU7vhtUN"
    "0+r0bljdMK1O74bVDdPq9G5Y3TDSTq/KnV5VOr1D3elVudOrep1edev0qn6nd/sQO9+hvU7vsfP24Hl38Lx/fDM7N0yv07v/"
    "Ep+9YQ51erfTL7thDnV6t9Mvu2Fand4N6xum2+lVwk6vanR6i1U3TKvTu2F1w7Q6vRtWN0yr07thdcO0Or0bVjdMq9O7YXXD"
    "SDu9Knd6Ven0DnWnV+VOr+p1etWt06v6nd7tQ+x8h/Y6vcfO24Pn3cHz/vHN7NwwvU7v/kt89oY51OndTr/shjnU6d1Ov+yG"
    "aXV6N6xvmG6nVwk7varR6S1W3TCtTu+G1Q3T6vRuWN0wrU7vhtUN0+r0bljdMK1O74bVDSPt9Krc6VWl0zvUnV6VO72q1+lV"
    "t06v6nd6tw+x8x3a6/QeO28PnncHz/vHN7Nzw/Q6vfsv8dkb5lCndzv9shvmUKd3O/2yG6bV6d2wvmG6nV4l7PSqRqe3WHXD"
    "tDq9G1Y3TKvTu2F1w7Q6vRtWN0yr07thdcO0Or0bVjeMtNOrcqdXlU7vUHd6Ve70ql6nV906varf6d0+xM53aK/Te+y8PXje"
    "HTzvH9/Mzg3T6/Tuv8Rnb5hDnd7t9MtumEOd3u30y26YVqd3w/qG6XZ6lbDTqxqd3mLVDdPq9G5Y3TCtTu+G1Q3T6vRuWN0w"
    "rU7vhtUN0+r0bljdMNJOr8qdXlU6varu9Krc6VW9Tq+6dXpVv9O7fYjnv0NVr9N77Lw9eN4dPO8f38zzN0znfBC8xGdvmEOd"
    "3u30y26YQ53e7fTLbphWp3fD+obpdnqVsNOrGp3eYtUN0+r0bljdMK1O74bVDdPq9G5Y3TCtTu+G1Q3T6vRuWN0w0k6vyp1e"
    "VTq9qu70qtzpVb1Or7p1elW/07t9iJ3v0F6n99h5e/C8O3jeP76ZnRum1+ndf4nP3jCHOr3b6ZfdMIc6vdvpl90wrU7vhvUN"
    "0+30amGnVzc6vbrR6dWtTq9udXob+OYR728Y3er06lant4FfPOL9DdPALx/x/obR0k6vzp1eXTq9qu706tzp1b1Or751enW/"
    "06tlnVjV6/QeO28PnncHz/vHN7Nzw/Q6vfsv8bkbRh/q9OpDnV59qNOrD3V6davTq5udXt3t9Gphp1c3Or260enVrU6vbnV6"
    "G/jmEasbptXp1a1ObwO/eMTqhml1enWr06ulnV6dO726dHpV3enVudOre51efev06n6nV8s6sarX6T123h487w6e949vZueG"
    "6XV691/iszfMoU6vPtTp1Yc6vfpQp1e3Or262enV3U6vFnZ6daPTqxudXt3q9OpWp7eBbx6xumFanV7d6vQ28ItHrG6YVqdX"
    "tzq9Wtrp1bnTq0unV9WdXp07vbrX6dW3Tq/ud3q1rBOrep3eY+ftwfPu4Hn/+GZ2bphep3f/JT57wxzq9OpDnV59qNOrD3V6"
    "davTq5udXt3t9Gphp1c3Or260enVrU6vbnV6G/jmEasbptXp1a1ObwO/eMTqhml1enWr06ulnV6dO726dHpV3enVudOre51e"
    "fev06n6nV8s6sarX6T123h487w6e949vZueG6XV691/iszfMoU6vPtTp1Yc6vfpQp1e3Or262enV3U6vFnZ6daPTW6y6YVqd"
    "3g2rG6bV6d2wumFand4Nqxum1endsLphWp3eDasbRtrp1bnTq0unV9WdXp07vbrX6dW3Tq/ud3q3D7HzHdrr9B47bw+edwfP"
    "+8c3s3PD9Dq9+y/x2RvmUKd3O/2yG+ZQp3c7/bIbptXp3bC+YbqdXi3s9OpGp7dYdcO0Or0bVjdMq9O7YXXDtDq9G1Y3TKvT"
    "u2F1w7Q6vRtWN4y006tzp1eXTq+qO706d3p1r9Orb51e3e/0bh9i5zu01+k9dt4ePO8OnvePb2bnhul1evdf4rM3zKFO73b6"
    "ZTfMoU7vdvplN0yr07thfcN0O71a2OnVjU5vseqGaXV6N6xumFand8Pqhml1ejesbphWp3fD6oZpdXo3rG4YaadX506vLp1e"
    "VXd6de706l6nV986vbrf6d0+xM53aK/Te+y8PXjeHTzvH9/Mzg3T6/Tuv8Rnb5hDnd7t9MtumEOd3u30y26YVqd3w/qG6XZ6"
    "tbDTqxud3mLVDdPq9G5Y3TCtTu+G1Q3T6vRuWN0wrU7vhtUN0+r0bljdMNJOr86dXl06varu9Orc6dW9Tq++dXp1v9O7fYid"
    "79Bep/fYeXvwvDt43j++mZ0bptfp3X+Jz94whzq92+mX3TCHOr3b6ZfdMK1O74b1DdPt9Gphp1c3Or3Fqhum1endsLphWp3e"
    "DasbptXp3bC6YVqd3g2rG6bV6d2wumGknV6dO726dHp13enVudOre51efev06n6nd/sQz3+H6l6n99h5e/C8O3jeP76Z52+Y"
    "zvkgeInP3jCHOr3b6ZfdMIc6vdvpl90wrU7vhvUN0+30amGnVzc6vcWqG6bV6d2wumFand4Nqxum1endsLphWp3eDasbptXp"
    "3bC6YaSdXp07vbp0enXd6dW506t7nV596/Tqfqd3+xA736G9Tu+x8/bgeXfwvH98Mzs3TK/Tu/8Sn71hDnV6t9Mvu2EOdXq3"
    "0y+7YVqd3g3rG6bb6TXCTq9pdHpNo9NrWp1e0+r0NvDNI97fMKbV6TWtTm8Dv3jE+xumgV8+4v0NY6SdXpM7vaZ0enXd6TW5"
    "02t6nV5z6/SafqfXyDqxutfpPXbeHjzvDp73j29m54bpdXr3X+JzN4w51Ok1hzq95lCn1xzq9JpWp9c0O72m2+k1wk6vaXR6"
    "TaPTa1qdXtPq9DbwzSNWN0yr02tand4GfvGI1Q3T6vSaVqfXSDu9Jnd6Ten06rrTa3Kn1/Q6vebW6TX9Tq+RdWJ1r9N77Lw9"
    "eN4dPO8f38zODdPr9O6/xGdvmEOdXnOo02sOdXrNoU6vaXV6TbPTa7qdXiPs9JofO3vq/oL5kc73F8RTmgs9fsv/b9IDyZIc"
    "yZMCKZIW0kpKpDzMbn3RCOuL5sJhXjjMpzQXqob5lB5IluRInhRIkbSQVlIi5WF2m1pG2NQyA4c5cJhPaS5UDfMpPZAsyZE8"
    "KZAiaSGtpETKw+yWUoywlGIUh6k4zKc0F6qG+ZQeSJbkSJ4USJG0kFZSIuVhdn/93Qh//d1oDlNzmE9pLlQN8yk9kCzJkTwp"
    "kCJpIa2kRMrD7P5SoxH+UqMxHKbhMJ/SXKga5lN6IFmSI3lSIEXSQlpJiZSH2f1VFSP8VRUzcpgjh/mU5kLVMJ/SA8mSHMmT"
    "AimSFtJKSqQ8zO4C2QgXyGbiMCcO8ynNhaphPqUHkiU5kicFUiQtpJWUSHmY3V3ZKNyVja8wzJ/ofpigudD9MEEPJEtyJE8K"
    "pEhaSCspka7DHLtrgVG4FhhPHOaJw3xKc6FqmE/pgWRJjuRJgRRJC2klJVIeZjcBjcIENDIBjUxAoLlQNUwmIJAlOZInBVIk"
    "LaSVlEh5mN0ENAoT0MgENDIBgeZC1TCZgECW5EieFEiRtJBWUiLlYXYT0ChMQCMT0MgEBJoLVcNkAgJZkiN5UiBF0kJaSYmU"
    "h9lNQKMwAY1MQCMTEGguVA2TCQhkSY7kSYEUSQtpJSVSHmY3AY3CBDQyAY1MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBDQK"
    "E9DIBDQyAYHmQtUwmYBAluRInhRIkbSQVlIi5WF2E9AoTEAjE9DIBASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUCjMAGNTEAj"
    "ExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwFNwgQ0MQFNTECgudD9MEEPJEtyJE8KpEhaSCspka7DnLoJaBImoIkJaGICAs2F"
    "qmEyAYEsyZE8KZAiaSGtpETKw+wmoEmYgCYmoIkJCDQXqobJBASyJEfypECKpIW0khIpD7ObgCZhApqYgCYmINBcqBomExDI"
    "khzJkwIpkhbSSkqkPMxuApqECWhiApqYgEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CWgSJqCJCWhiAgLNhaphMgGBLMmRPCmQ"
    "ImkhraREysPsJqBJmIAmJqCJCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4AmYQKamIAmJiDQXKgaJhMQyJIcyZMCKZIW0kpK"
    "pDzMbgKahAloYgKamIBAc6FqmExAIEtyJE8KpEhaSCspkfIwuwloEiagiQloYgICzYWqYTIBgSzJkTwpkCJpIa2kRMrD7CWg"
    "yytZAro9Vw+z0N0wSXOhu2GSHkiW5EieFEiRtJBWUiJ9/rtirWHKEtDtuafDRAIizYWqYSIBkSzJkTwpkCJpIa2kRMrD7CWg"
    "yytZAro993SYSECkuVA1TCQgkiU5kicFUiQtpJWUSHmYvQR0eSVLQLfnng4TCYg0F6qGiQREsiRH8qRAiqSFtJISKQ+zl4Au"
    "r2QJ6Pbc02EiAZHmQtUwkYBIluRInhRIkbSQVlIi5WH2EtDllSwB3Z57OkwkINJcqBomEhDJkhzJkwIpkhbSSkqkPMxeArq8"
    "kiWg23NPh4kERJoLVcNEAiJZkiN5UiBF0kJaSYmUh9lLQJdXsgR0e+7pMJGASHOhaphIQCRLciRPCqRIWkgrKZHyMHsJ6PJK"
    "loBuzz0dJhIQaS5UDRMJiGRJjuRJgRRJC2klJVIeZi8BXV7JEtDtuafDRAIizYWqYSIBkSzJkTwpkCJpIa2kRMrD7CagkzAB"
    "nZiATkxAoLnQ/TBBDyRLciRPCqRIWkgrKZGuwzx1E9BJmIBOTEAnJiDQXKgaJhMQyJIcyZMCKZIW0kpKpDzMbgI6CRPQiQno"
    "xAQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1AJ2ECOjEBnZiAQHOhaphMQCBLciRPCqRIWkgrKZHyMLsJ6CRMQCcmoBMTEGgu"
    "VA2TCQhkSY7kSYEUSQtpJSVSHmY3AZ2ECejEBHRiAgLNhaphMgGBLMmRPCmQImkhraREysPsJqCTMAGdmIBOTECguVA1TCYg"
    "kCU5kicFUiQtpJWUSHmY3QR0EiagExPQiQkINBeqhskEBLIkR/KkQIqkhbSSEikPs5uATsIEdGICOjEBgeZC1TCZgECW5Eie"
    "FEiRtJBWUiLlYXYT0EmYgE5MQCcmINBcqBomExDIkhzJkwIpkhbSSkqkPMxuAjoLE9CZCejMBASaC90PE/RAsiRH8qRAiqSF"
    "tJIS6TrMczcBnYUJ6MwEdGYCAs2FqmEyAYEsyZE8KZAiaSGtpETKw+wmoLMwAZ2ZgM5MQKC5UDVMJiCQJTmSJwVSJC2klZRI"
    "eZjdBHQWJqAzE9CZCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4DOwgR0ZgI6MwGB5kLVMJmAQJbkSJ4USJG0kFZSIuVhdhPQ"
    "WZiAzkxAZyYg0FyoGiYTEMiSHMmTAimSFtJKSqQ8zG4COgsT0JkJ6MwEBJoLVcNkAgJZkiN5UiBF0kJaSYmUh9lNQGdhAjoz"
    "AZ2ZgEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CegsTEBnJqAzExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwGdhQnozAR0ZgIC"
    "zYWqYTIBgSzJkTwpkCJpIa2kRMrD7CagizABXZiALkxAoLnQ/TBBDyRLciRPCqRIWkgrKZGuw7x0E9BFmIAuTEAXJiDQXKga"
    "JhMQyJIcyZMCKZIW0kpKpDzMbgK6CBPQhQnowgQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1AF2ECujABXZiAQHOhaphMQCBL"
    "ciRPCqRIWkgrKZHyMLsJ6CJMQBcmoAsTEGguVA2TCQhkSY7kSYEUSQtpJSVSHmY3AV2ECejCBHRhAgLNhaphMgGBLMmRPCmQ"
    "ImkhraREysPsJqCLMAFdmIAuTECguVA1TCYgkCU5kicFUiQtpJWUSHmY3QR0ESagCxPQhQkINBeqhskEBLIkR/KkQIqkhbSS"
    "EikPs5uALsIEdGECujABgeZC1TCZgECW5EieFEiRtJBWUiLlYXYT0EWYgC5MQBcmINBcqBomExDIkhzJkwIpkhbSSkqkPMxu"
    "AhqECWhgAhqYgEBzofthgh5IluRInhRIkbSQVlIiXYc5dBPQIExAAxPQwAQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1AgzAB"
    "DUxAAxMQaC5UDZMJCGRJjuRJgRRJC2klJVIeZjcBDcIENDABDUxAoLlQNUwmIJAlOZInBVIkLaSVlEh5mN0ENAgT0MAENDAB"
    "geZC1TCZgECW5EieFEiRtJBWUiLlYXYT0CBMQAMT0MAEBJoLVcNkAgJZkiN5UiBF0kJaSYmUh9lNQIMwAQ1MQAMTEGguVA2T"
    "CQhkSY7kSYEUSQtpJSVSHmY3AQ3CBDQwAQ1MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBDQIE9DABDQwAYHmQtUwmYBAluRI"
    "nhRIkbSQVlIi5WF2E9AgTEADE9DABASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUBKmIAUE5BiAgLNhe6HCXogWZIjeVIgRdJC"
    "WkmJdB2m6iYgJUxAiglIMQGB5kLVMJmAQJbkSJ4USJG0kFZSIuVhdhOQEiYgxQSkmIBAc6FqmExAIEtyJE8KpEhaSCspkfIw"
    "uwlICROQYgJSTECguVA1TCYgkCU5kicFUiQtpJWUSHmY3QSkhAlIMQEpJiDQXKgaJhMQyJIcyZMCKZIW0kpKpDzMbgJSwgSk"
    "mIAUExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwEpYQJSTECKCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4CUMAEpJiDFBASa"
    "C1XDZAICWZIjeVIgRdJCWkmJlIfZTUBKmIAUE5BiAgLNhaphMgGBLMmRPCmQImkhraREysPsJiAlTECKCUgxAYHmQtUwmYBA"
    "luRInhRIkbSQVlIi5WF2E5AWJiDNBKSZgEBzofthgh5IluRInhRIkbSQVlIiXYepuwlICxOQZgLSTECguVA1TCYgkCU5kicF"
    "UiQtpJWUSHmY3QSkhQlIMwFpJiDQXKgaJhMQyJIcyZMCKZIW0kpKpDzMbgLSwgSkmYA0ExBoLlQNkwkIZEmO5EmBFEkLaSUl"
    "Uh5mNwFpYQLSTECaCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4C0MAFpJiDNBASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUBa"
    "mIA0E5BmAgLNhaphMgGBLMmRPCmQImkhraREysPsJiAtTECaCUgzAYHmQtUwmYBAluRInhRIkbSQVlIi5WF2E5AWJiDNBKSZ"
    "gEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CUgLE5BmAtJMQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBGSECcgwARkmINBc6H6Y"
    "oAeSJTmSJwVSJC2klZRI12GabgIywgRkmIAMExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwEZYQIyTECGCQg0F6qGyQQEsiRH"
    "8qRAiqSFtJISKQ+zm4CMMAEZJiDDBASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUBGmIAME5BhAgLNhaphMgGBLMmRPCmQImkh"
    "raREysPsJiAjTECGCcgwAYHmQtUwmYBAluRInhRIkbSQVlIi5WF2E5ARJiDDBGSYgEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7"
    "CcgIE5BhAjJMQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBGSECcgwARkmINBcqBomExDIkhzJkwIpkhbSSkqkPMxuAjLCBGSY"
    "gAwTEGguVA2TCQhkSY7kSYEUSQtpJSVSHmY3AY3CBDQyAY1MQKC50P0wQQ8kS3IkTwqkSFpIKymRrsMcuwloFCagkQloZAIC"
    "zYWqYTIBgSzJkTwpkCJpIa2kRMrD7CagUZiARiagkQkINBeqhskEBLIkR/KkQIqkhbSSEikPs5uARmECGpmARiYg0FyoGiYT"
    "EMiSHMmTAimSFtJKSqQ8zG4CGoUJaGQCGpmAQHOhaphMQCBLciRPCqRIWkgrKZHyMLsJaBQmoJEJaGQCAs2FqmEyAYEsyZE8"
    "KZAiaSGtpETKw+wmoFGYgEYmoJEJCDQXqobJBASyJEfypECKpIW0khIpD7ObgEZhAhqZgEYmINBcqBomExDIkhzJkwIpkhbS"
    "SkqkPMxuAhqFCWhkAhqZgEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CWgUJqCRCWhkAgLNhaphMgGBLMmRPCmQImkhraREysPs"
    "JqBJmIAmJqCJCQg0F7ofJuiBZEmO5EmBFEkLaSUl0nWYUzcBTcIENDEBTUxAoLlQNUwmIJAlOZInBVIkLaSVlEh5mN0ENAkT"
    "0MQENDEBgeZC1TCZgECW5EieFEiRtJBWUiLlYXYT0CRMQBMT0MQEBJoLVcNkAgJZkiN5UiBF0kJaSYmUh9lNQJMwAU1MQBMT"
    "EGguVA2TCQhkSY7kSYEUSQtpJSVSHmY3AU3CBDQxAU1MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBDQJE9DEBDQxAYHmQtUw"
    "mYBAluRInhRIkbSQVlIi5WF2E9AkTEATE9DEBASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUCTMAFNTEATExBoLlQNkwkIZEmO"
    "5EmBFEkLaSUlUh5mNwFNwgQ0MQFNTECguVA1TCYgkCU5kicFUiQtpJWUSHmYvQQ0vJIloNtz9TAL3Q2TNBe6GybpgWRJjuRJ"
    "gRRJC2klJdLnvyvWGqYsAd2eezpMJCDSXKgaJhIQyZIcyZMCKZIW0kpKpDzMXgIaXskS0O25p8NEAiLNhaphIgGRLMmRPCmQ"
    "ImkhraREysPsJaDhlSwB3Z57OkwkINJcqBomEhDJkhzJkwIpkhbSSkqkPMxeAhpeyRLQ7bmnw0QCIs2FqmEiAZEsyZE8KZAi"
    "aSGtpETKw+wloOGVLAHdnns6TCQg0lyoGiYSEMmSHMmTAimSFtJKSqQ8zF4CGl7JEtDtuafDRAIizYWqYSIBkSzJkTwpkCJp"
    "Ia2kRMrD7CWg4ZUsAd2eezpMJCDSXKgaJhIQyZIcyZMCKZIW0kpKpDzMXgIaXskS0O25p8NEAiLNhaphIgGRLMmRPCmQImkh"
    "raREysPsJaDhlSwB3Z57OkwkINJcqBomEhDJkhzJkwIpkhbSSkqkPMxuAjoJE9CJCejEBASaC90PE/RAsiRH8qRAiqSFtJIS"
    "6TrMUzcBnYQJ6MQEdGICAs2FqmEyAYEsyZE8KZAiaSGtpETKw+wmoJMwAZ2YgE5MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjd"
    "BHQSJqATE9CJCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4BOwgR0YgI6MQGB5kLVMJmAQJbkSJ4USJG0kFZSIuVhdhPQSZiA"
    "TkxAJyYg0FyoGiYTEMiSHMmTAimSFtJKSqQ8zG4COgkT0IkJ6MQEBJoLVcNkAgJZkiN5UiBF0kJaSYmUh9lNQCdhAjoxAZ2Y"
    "gEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CegkTEAnJqATExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwGdhAnoxAR0YgICzYWq"
    "YTIBgSzJkTwpkCJpIa2kRMrD7CagszABnZmAzkxAoLnQ/TBBDyRLciRPCqRIWkgrKZGuwzx3E9BZmIDOTEBnJiDQXKgaJhMQ"
    "yJIcyZMCKZIW0kpKpDzMbgI6CxPQmQnozAQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1AZ2ECOjMBnZmAQHOhaphMQCBLciRP"
    "CqRIWkgrKZHyMLsJ6CxMQGcmoDMTEGguVA2TCQhkSY7kSYEUSQtpJSVSHmY3AZ2FCejMBHRmAgLNhaphMgGBLMmRPCmQImkh"
    "raREysPsJqCzMAGdmYDOTECguVA1TCYgkCU5kicFUiQtpJWUSHmY3QR0FiagMxPQmQkINBeqhskEBLIkR/KkQIqkhbSSEikP"
    "s5uAzsIEdGYCOjMBgeZC1TCZgECW5EieFEiRtJBWUiLlYXYT0FmYgM5MQGcmINBcqBomExDIkhzJkwIpkhbSSkqkPMxuAroI"
    "E9CFCejCBASaC90PE/RAsiRH8qRAiqSFtJIS6TrMSzcBXYQJ6MIEdGECAs2FqmEyAYEsyZE8KZAiaSGtpETKw+wmoIswAV2Y"
    "gC5MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBHQRJqALE9CFCQg0F6qGyQQEsiRH8qRAiqSFtJISKQ+zm4AuwgR0YQK6MAGB"
    "5kLVMJmAQJbkSJ4USJG0kFZSIuVhdhPQRZiALkxAFyYg0FyoGiYTEMiSHMmTAimSFtJKSqQ8zG4CuggT0IUJ6MIEBJoLVcNk"
    "AgJZkiN5UiBF0kJaSYmUh9lNQBdhArowAV2YgEBzoWqYTEAgS3IkTwqkSFpIKymR8jC7CegiTEAXJqALExBoLlQNkwkIZEmO"
    "5EmBFEkLaSUlUh5mNwFdhAnowgR0YQICzYWqYTIBgSzJkTwpkCJpIa2kRMrD7CagQZiABiaggQkINBe6HybogWRJjuRJgRRJ"
    "C2klJdJ1mEM3AQ3CBDQwAQ1MQKC5UDVMJiCQJTmSJwVSJC2klZRIeZjdBDQIE9DABDQwAYHmQtUwmYBAluRInhRIkbSQVlIi"
    "5WF2E9AgTEADE9DABASaC1XDZAICWZIjeVIgRdJCWkmJlIfZTUCDMAENTEADExBoLlQNkwkIZEmO5EmBFEkLaSUlUh5mNwEN"
    "wgQ0MAENTECguVA1TCYgkCU5kicFUiQtpJWUSHmY3QQ0CBPQwAQ0MAGB5kLVMJmAQJbkSJ4USJG0kFZSIuVhdhPQIExAAxPQ"
    "wAQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1AgzABDUxAAxMQaC5UDZMJCGRJjuRJgRRJC2klJVIeZjcBDcIENDABDUxAoLlQ"
    "NUwmIJAlOZInBVIkLaSVlEh5mN0EpIQJSDEBKSYg0FzofpigB5IlOZInBVIkLaSVlEjXYapuAlLCBKSYgBQTEGguVA2TCQhk"
    "SY7kSYEUSQtpJSVSHmY3ASlhAlJMQIoJCDQXqobJBASyJEfypECKpIW0khIpD7ObgJQwASkmIMUEBJoLVcNkAgJZkiN5UiBF"
    "0kJaSYmUh9lNQEqYgBQTkGICAs2FqmEyAYEsyZE8KZAiaSGtpETKw+wmICVMQIoJSDEBgeZC1TCZgECW5EieFEiRtJBWUiLl"
    "YXYTkBImIMUEpJiAQHOhaphMQCBLciRPCqRIWkgrKZHyMLsJSAkTkGICUkxAoLlQNUwmIJAlOZInBVIkLaSVlEh5mN0EpIQJ"
    "SDEBKSYg0FyoGiYTEMiSHMmTAimSFtJKSqQ8zG4CUsIEpJiAFBMQaC5UDZMJCGRJjuRJgRRJC2klJVIeZjcBaWEC0kxAmgkI"
    "NBe6HybogWRJjuRJgRRJC2klJdJ1mLqbgLQwAWkmIM0EBJoLVcNkAgJZkiN5UiBF0kJaSYmUh9lNQFqYgDQTkGYCAs2FqmEy"
    "AYEsyZE8KZAiaSGtpETKw+wmIC1MQJoJSDMBgeZC1TCZgECW5EieFEiRtJBWUiLlYXYTkBYmIM0EpJmAQHOhaphMQCBLciRP"
    "CqRIWkgrKZHyMLsJSAsTkGYC0kxAoLlQNUwmIJAlOZInBVIkLaSVlEh5mN0EpIUJSDMBaSYg0FyoGiYTEMiSHMmTAimSFtJK"
    "SqQ8zG4C0sIEpJmANBMQaC5UDZMJCGRJjuRJgRRJC2klJVIeZjcBaWEC0kxAmgkINBeqhskEBLIkR/KkQIqkhbSSEikPs5uA"
    "tDABaSYgzQQEmgtVw2QCAlmSI3lSIEXSQlpJiZSH2U1ARpiADBOQYQICzYXuhwl6IFmSI3lSIEXSQlpJiXQdpukmICNMQIYJ"
    "yDABgeZC1TCZgECW5EieFEiRtJBWUiLlYXYTkHk+Af3i7snn//H+/snn/9n1/snn/8Hs/snn/6nj/snn/y/1/snn///i/snn"
    "L8O7J3d+7+P7J5//2/j+SfE72vktWe+fFL+jnd9M8/5J8Tva+W0Q758Uv6Od38Du7smd353t/knxO9r5fbXunxS/o53fEen+"
    "SfE72vm9bO6fFL+jnd+F5P5J6TtSO79/xP2T0nekdn7y//2T0nekdn5m+/2T0nekdn7a9v2T0nekdn5O8v2T4ne08xNu758U"
    "v6Odn016/6T4He38VMn7J8XvaOfnAd4/KX5HOz/J7f5J8Tva+Rlc90+K39HOT0+6f1L8jnZ+7s39k+J3tPMTS+6fFL+jnZ81"
    "cf+k+B3t/JSA+yfF72jn3+++f1L8jnb+zdz7J8XvaOffqbx/UvyOdv5tuPsnxe9o599jun9S/I52/g2U+yfF72jn3x24f1L8"
    "jnZa3/dPit/RTl/3/knxO9ppWt4/KX5HOx25+yfF72in3XT/pPgd7fRS7p8Uv6OdRsH9k+J3tPNrwfdPit/Rzq/i3T8pfkc7"
    "v/5y/6T4He1szu+fFL+jnZ3n/ZPid7Szrbp/UvyOxHsGJd4zKPGeQYn3DEq8Z1DiPYMS7xmUeM+gxHsGJd4zKPGeQYn3DEq8"
    "Z1DiPYMS7xmUeM+gxHsGJd4zKPGeQYn3DEq8Z1DiPYMS7xmUeM+gxHsGJd4zKPGeQYn3DFq8Z9DiPYMW7xm0eM+gxXsGLd4z"
    "aPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0"
    "eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8"
    "Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4z"
    "aPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0"
    "eM+gxXsGLd4zaPGeQYv3DFq8Z9DiPYMW7xm0eM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8"
    "ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94z"
    "GPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmM"
    "eM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8"
    "ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94zGPGewYj3DEa8ZzDiPYMR7xmMeM9gxHsGI94z"
    "GPGewYj3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4"
    "zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8w"
    "ivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3"
    "DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyj"
    "eM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMo3jPMIr3DKN4zzCK9wyjeM8wivcMk3jP"
    "MIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ"
    "9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcM"
    "k3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4"
    "zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8w"
    "ifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wyTeM8wifcMk3jPMIn3DJN4zzCJ9wynV/uLhl/88M/v3n36"
    "7dtPb3/9y3999/Gbd7959/79D5/96cNfvrs+cv2H4jv+7OO7r3/1sz/o11/qfPSpm9dfmob/dnz9Zmw9P77+suUP5rVr/XG8"
    "eR1b/jC+dq0/jh9fx5Yv4+vU8s8vr/94abl6nVTrr0u/ftOaw4N+7Vru9evY8kW/Ti3/rXn9pvXXu5jXqeWfT6//OLX89Or1"
    "H0+vGl/5p9Pp+pVT4yu/P51f+9O58ZVw/UpqfuX/XL/yx+ZXfn8aXv/hNDS+8sX1K182vzJfv/LQ/Iq9fsU3v/L59RO8aX6C"
    "z0+X61daZ35//cqXp9Z7n69f8c2vhOtXwqn190S8fmVp/nnW61fSj1/5xeP32a9/+acP33317advP3z39v2bDx//9e2nT99+"
    "981nP/z5p0PXvzdOOv9U7q/9X96/++zTv3//7lc/+9P16B9++NlnX/3b13/46lc/u36Pf//x2w8fv/307/mb+rPr9/j3f/g6"
    "fvzLu9v//PD9u49vP3243gfv/vyXt++vf7Cvr3+iv7x/++u/+/NfPnz6x8/fv/v46cf/+stflC/98hc//hnzf2l9wr/JBz/d"
    "f/Dzyz74Z3/jT/5P10/+5mxOTz/6u3/7/uO7H364nto+/vn+41+efvzHT/vzz0/6Vz9+zPntN9+9+1sP+/fXj/yl7CNf7j/y"
    "0P/Iv98+8X/eZx3uP6t69rP+3Tef/vFXP//NSf+Pv//5b0/6v/33V/8w/md8ZHX/kfXzH/n9Tx/5f/39z3/3t/7I68m8Xhsf"
    "+bmLw/wXuThEn7y6Ocb/GjdHun709B8c+vRfYujCT14NPf/yyP+nqX91/YfT9e37b6//ef3aD0/hpw/7/tsfPv3ss7fv33/4"
    "6z+9f/vdv/z4f4X//OGvv/v48cPH+fqd+vabHz/nT3/V11v59ec//lX/9FlOP33Q33z/L//zN9e/zG//9Pb9k497un7M+k/f"
    "+ThfvfvTt/96HcLdVL75+O7tp3cf4z+//e7h4+9uI/qPfN7fnF+Z17978oFf/cOrZz/bLzC86yv87tPD97f/df2Hke+vf5b5"
    "7cdvvv3uh8/+74dP13hw/Vv0Fha+/vDh+mnz/8oh4t3br7b/8f7d159uT/3ss48/Ronbf7/+3fDT2Z/+uOHdp798/9n1b5h3"
    "3326/el/9bPvP3z89PHtt5/yI199fPvX/Hfhx9ffXv/u+viHr34MJH/98PFfbqHk1/8PUEsHCGUUhY9MoAAA5OcFAFBLAwQU"
    "AAgICACkk7ZcAAAAAAAAAAAAAAAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzjc9LCsIwEAbgE3iH"
    "MHuT1oWINO1GhG6lHmBIpg9sHiTx0dubjaLgwuXMz3zDXzUPM7MbhTg5K6HkBTCyyunJDhLO3XG9AxYTWo2zsyRhoQhNvapO"
    "NGPKN3GcfGQZsVHCmJLfCxHVSAYjd55sTnoXDKY8hkF4VBccSGyKYivCpwH1l8laLSG0ugTWLZ7+sV3fT4oOTl0N2fTjhdAB"
    "77lYJjEMlCRw/tq9w5JnFkRdia+K9RNQSwcIrajrTbMAAAAqAQAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAAYAAAAeGwv"
    "d29ya3NoZWV0cy9zaGVldDIueG1snd3NblvHGYDhK+g9CNzH4vx/Y1jKokHQLIoGTdOuGYmSCYukQNKxc/elZVtN6kUfd2OL"
    "0pzhQC+1eXDOzKtv328fLn5dH46b/e5qkV4sFxfr3c3+drO7v1r8/I/vv4nFxfG02t2uHva79dXit/Vx8e31n1692x/eHF+v"
    "16eL8wS749Xi9en0+PLy8njzer1dHV/sH9e780/u9oft6nR+ebi/PD4e1qvbp4u2D5d5ueyX29Vmt/g4w8uDzLG/u9vcrL/b"
    "37zdrnenj5Mc1g+r03n5x9ebx+Pn2bbvv5huu7k57I/7u9OLm/3200znFdxcrt/frJ8WFH9Y0PZGVrRdHd68ffzmPOXjeRW/"
    "bB42p9+e1vU8za9Xi7eH3ctPc3zzvIwP17w8v//LX7cPnwe/T9XW/cUvc17OP6z+fWr/30xpeZnSf01VV1/+LnxZq5vnmbY2"
    "zXORTx+R61dPU/54uH71uLpf/7Q+/fz44+Hy+tXl8/efvvjnZv3u+LuvLz58TH/Z7998ePHD7dViuXi+6Pdjv38K+uPh4ubt"
    "8bTf/mW9uX99Ov85LC5u13ertw+nP+8f/rW5Pb0+f6++qOX5+3/fv3se3F48zX6zfzg+/ftpss/XLS62m93H/1fnD+f58/Hu"
    "40/ixUhPy/p45dOCvludVtevDvt3F+c/i3y+5tM7jLb4Yo2X/xmYdWDRgVUHNh3YdeDQgaEDJw4sSx2oZYqWKVqmaJmiZYqW"
    "KVqmaJmiZaqWqVqmapmqZaqWqVqmapmqZaqWqVqmaZmmZZqWaVqmaZmmZZqWaVqmaZmmZbqW6Vqma5muZbqW6Vqma5muZbqW"
    "6VpmaJmhZYaWGVpmaJmhZYaWGVpmaJmhZULLhJYJLRNaJrRMaJnQMqFlQsuElplaZmqZqWWmlplaZmqZqWWmlplaZmqZtNQ0"
    "aalt0lLjpKXWSUvNk5baJy01UFpqobTURGnJjRI3StwocaPEjRI3StwocaPEjRI3StwocyOWgcQ0kNgGEuNAYh1IzAOJfSAx"
    "ECQWgsREkNgIEiNBYiVIzASJnSAxFCSWgsRUkNgKEmNBYi1IzAWJvSAxGCQWg8RkkNgMEqNBYjVIzAaJ3SAxHCSWg8R0kNgO"
    "EuNBYj1IzAeJ/SAxICQWhMSEkNgQEiNCYkVIzAiJHSExJCSWhMSUkNgSEmNCYk1IzAmJPSExKCQWhcSkkNgUEqNCYlVIzAqJ"
    "XSExLCSWhcS0kNgWEuNCYl1IzAuJfSExMCQWhsTEkNgYEiNDYmVIzAyJnSGzM2R2hszOkNkZMjtDZmfI7AyZnSGzM2R2hszO"
    "kNkZMjtDZmfI7AyZnSGzM2R2hszOkNkZMjtD/oo7ELiR34PgNyH4XQh+G4Lfh+A3IrAzZHaGzM6Q2RkyO0NmZ8jsDJmdIbMz"
    "ZHaGzM6Q2RkyO0NmZ8jsDJmdIbMzZHaGzM6Q2RkyO0NmZ8jsDJmdIbMzZHaGzM6Q2RkyO0NmZ8jsDJmdIbMzZHaGzM6Q2Rky"
    "O0NmZ8jsDJmdIbMzZHaGzM6Q2RkyO0NmZ8jsDJmdIbMzZHaGzM6Q2RkyO0NmZ8jsDJmdIbMzZHaGzM6Q2RkyO0NmZ8jsDJmd"
    "IbMzZHaGzM6Q2RkyO0NmZ8jsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q"
    "2BkKO0NhZyjsDIWdobAzFHaGws5Q2BmKP/Hgjzz4Mw9f8dADN/LHHvy5B3/wwZ988Ecf2BkKO0NhZyjsDIWdobAzFHaGws5Q"
    "2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjs"
    "DIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobAzFHaGws5Q2BkKO0NhZyjsDIWdobIzVHaG"
    "ys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0Nl"
    "Z6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDNX3WPBNFnyXBd9m4Sv2WeBGvtOCb7Xgey34ZgvsDJWdobIzVHaGys5Q2RkqO0Nl"
    "Z6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIz"
    "VHaGys5Q2RkqO0NlZ6jsDJWdobIzVHaGys5Q2RkqO0NlZ6jsDJWdobIzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q2Bka"
    "O0NjZ2jsDI2dobEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q2BkaO0NjZ2jsDI2d"
    "obEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaG5rs6+raOvq+jb+zoOzt+xdaO3Mg3d/TdHX17R3aGxs7Q2BkaO0NjZ2jsDI2d"
    "obEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q2BkaO0NjZ2jsDI2dobEzNHaGxs7Q"
    "2BkaO0NjZ2jsDI2dobEzNHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+js"
    "DJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaG"
    "zs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0P3cyT8IAk/ScKPkvCzJPwwia84TYIb+XkSfqAEO0NnZ+jsDJ2dobMzdHaG"
    "zs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs7Q2Rk6O0NnZ+jsDJ2dobMzdHaGzs4w2BkGO8Ng"
    "ZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYbAz"
    "DHaGwc4w2BkGO8NgZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYbAzDHaGwc4w2BkG"
    "O8NgZxjsDIOdYbAzDHaGwc4w2BkGO8NgZxjsDIOdYfjJlX50pZ9d6YdX+umVfnwlO8P4igMsuZEfYcnOMNgZBjvDYGcY7AyD"
    "nWGwMwx2hsHOMNgZBjvDYGcY7AyDnWGwMwx2hsHOMNgZBjvDYGcY7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzB"
    "zhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q"
    "7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwM"
    "wc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHOEOwMwc4Q7AzBzhDsDMHO"
    "EOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNM"
    "dobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7"
    "w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51h"
    "sjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzjDZGSY7w2RnmOwMk51hsjNMdobJzpCW/xsaLo+v1+vTd6vT6vrV42GzO/3t"
    "8bTZ747nHz2u7td/XR3uN7vjxS/70/nKq8XyaZ67/f60Pnx49WH+9er2+cXD+u70NGpxcfj4Lk9fn/aPn679NO9P69Pbx4v9"
    "YbPenVYf3vBq8bDa3R5vVo/rD2NuD6t3m939xeHl5vZqcfjh9uNi3+0Pb54WfP1vUEsHCMXCpah7DAAA0KQAAFBLAwQUAAgI"
    "CACkk7ZcAAAAAAAAAAAAAAAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQyLnhtbC5yZWxzjc9LCsIwEAbgE3iHMHuT"
    "tgsRadqNCN1KPcCQTB/YJiGJj97ebBQLLlzO/Mw3/GX9nCd2Jx9GayTkPANGRlk9ml7CpT1t98BCRKNxsoYkLBSgrjblmSaM"
    "6SYMowssISZIGGJ0ByGCGmjGwK0jk5LO+hljGn0vHKor9iSKLNsJ/21AtTJZoyX4RufA2sXRP7btulHR0arbTCb+eCG0x0cq"
    "lkj0PUUJnL93n7DgiQVRlWJVsXoBUEsHCIUB9RW0AAAAKgEAAFBLAwQUAAgICACkk7ZcAAAAAAAAAAAAAAAAEwAAAHhsL3Ro"
    "ZW1lL3RoZW1lMS54bWzNV1tv2jAU/gX7D5bf19wIAQRULRTtYdOksWnPJnESr44T2aZd//0cJxDn1lYrlQoP2MffOf7OxT5m"
    "ef03o+ABc0FytoLOlQ0BZmEeEZas4K+fu88zCIRELEI0Z3gFn7CA1+tPS7SQKc4wUOpMLNAKplIWC8sSoRIjcZUXmKm1OOcZ"
    "kmrKEyvi6FGZzajl2vbUyhBhsNbnr9HP45iEeJuHxwwzWRnhmCKpqIuUFAIChjLFcZ9iLAVcn0jeUVxqiFIQUr4PNfMeNrp3"
    "yh/Bk8OGcvCA6Ara+gOt9dI6A6js43b6U+NqQHTvvmTPrez1cR17GoDCUHnR39u/nd9u/RprgKph3/bdNth6Tgtv2Pd6+Bu/"
    "/LbwXoOfDHDfND4aoGro9/CTSeBuJi283+CnPXxg32wnQQuvQSkl7L4fcX/qbU7eniFxTr+8DG9QllE5lT6TY3WUoT853ymA"
    "Tq4qTwbkU4FjFCrcBlFy4KTcAC0wGlsJxfCK1TGfEfauezXmLdNpHYKsHYHv+njqCMSE0r18ovir0MRETkm0U0I90UrngBep"
    "GtbbtXAJR3oMeC5/E5nuU1SobRy9QyJq04kARS5U3uCobR2aY/Ytjyqp45zOoFJAspHb/lmuAikr6TRoivlsXs8SYRLwtdHX"
    "kzA2a5PwBkgE3utIOPalWMwHWMyc51hYRlbUoQGo7CD+pGIERIgojso8Vfqn7F4802PBbLvtDrg3n1ws0y0SRrm1SRhlmKII"
    "d8UXzvV8Ppxqd5BGMHuPXFv9u4Gy9gw8qjPn+cpMiIoVjNWlpoZZoewJlkCAaKIeKqGsA/0/N0vBhdwikVYwvVT5nxGJOaAk"
    "U7VupoGyhpvjBvbHJTe3P17krG6ScRzjUI5Imqlaq4wMrr4RXE7yoyK9T6NHcKBH/gOpQPmBUwYwIkKeoxkRbhR3E8XOdVUf"
    "xYHXnn7M0CJFdUcxL/MKrsdnOoYfmmnXK2sohIdkd4mu+7JS59IcaSDB6C32fk3eYOUNs/IH77r5zH6+S7y9IRjUZsPUvGFq"
    "Y73jgg8CY7vpSNzc0Wy+sRt0q9Yy3pV61vkDd5Ks/wFQSwcIuDMBxSoDAAC5DgAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAA"
    "AAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWyNVlFu4zgMPcHegfBivzaN22KwGBRJBoaTNsHEcRC7/VdtNhYqS15JTidzo7nG"
    "nGwpu9l2rXQxf/YjaT7ykUwmX77VAg6oDVdyGlyNLwNAWaiSy/00uM9vLz4HYCyTJRNK4jQ4ogm+zH6bGGOBQqWZBpW1zU0Y"
    "mqLCmpmxalCS5Unpmll61fvQNBpZaSpEW4vw+vLyr7BmXAZQqFbaafCJkrSS/91i/Ap8CmYTw2cTO0vSPN2l6wiydH2fr9JN"
    "Bg9ZBPNVsthk9BqtIV4u4q+wW2zTXT4J7WwSutA+fKMsZO1jza3Fcmi8jVYepDRsmbYGJEVqFMgMlmAVbLUq28JSn+ACIqLe"
    "HEE9ga24Ib9Gaeu8HhEMyu4xUVZpJRhkSrQuzgA1BVjTaHVgwgUzeQTD6kYgNC7pkE0ENdPPlP8tG0Kp2QvJc8rWNo3g5PLC"
    "bfUfNqZSnV/Ja5Smy28rZoFpBNKqePYa4hS9MQ0rSGiSzKA+YDD7+QOuL6+vRmcKGkMkBGi+r6xL2wWUYzhpNuqeRuBrOKLS"
    "y66azB4F/078E1irverYWc1KdJUboIZp3HNjUTsZ3hme3vjkJxiWSrjRNSNYr+Muh/te6yRsZYkaBC+oGdgTV0RAv/9ol5xI"
    "kUINantqOXckTIOk/gFBvUjalzHAUC03N7Bp60fUN0Nb1svkG3LNaRbyY4OeaSW7nMoP6jLN0RSaN04I/6tKiQ+p9AM3Z/bD"
    "lG7Iz9p3eOBulEJ46G+G55FQmCvpo7RrVbCzlGMlLSssLOg0+NHLO5jCEt2kwZ61e6SpWsWEJbzQqkZKOoK5A+aunzETvHHQ"
    "lhDaXNLYNXIEqXNJqWmF81I1rR3rDLFLELu1cDtz16eIk8ShSmmaKioMEjoHrXYeCSsqLnE0JJp2IW8Z/icUlrkriulSojE0"
    "xWi6MnJHJa948dzhr1y2Dt2K9l9yD6uMkAduWiYoaNlHuUvbe3jM8uVi5zV2/npNMtKdP/FeHO8QvZuLHZpW+KfqnUckmTga"
    "7rncIS0OE790dEjG09X6Nf++vUhBPrc5TW1XF9zSrMBG1STJkMeMviCH2B+QK0GkZeG1MyLcekPMDpzOxtUH+LWHa94Nisd4"
    "lcDv/m/ZWd5/EsUhdnEGo80bD7H7hrbEd3zxwbMjRQt0tuHf34CQ/inM/gFQSwcI/9fwFXIDAABnCAAAUEsDBBQACAgIAKST"
    "tlwAAAAAAAAAAAAAAAANAAAAeGwvc3R5bGVzLnhtbO1dW5ObNhT+Bf0PDNPH7nJn1xnbmcSpO33odqbZzrTT7QO2sa0JIA/g"
    "xM6vr4R8wbHAXCQkNvWDDUJ8+qRzdM6RENbw7S4MlM9+nAAYjVTjXlcVP5rDBYhWI/XP5+ndo6okqRctvABG/kjd+4n6dvzD"
    "MEn3gf9x7fupghCiZKSu03TzRtOS+doPveQebvwIXVnCOPRSdBqvtGQT+94iwTeFgWbququFHohUgvBmZ9je/AonBPMYJnCZ"
    "3s9hqMHlEsz9a6SBNtC8+REpvIah0Am9+NN2c4dgN14KZiAA6T5jpY6H0TachmmizOE2SkeqdUpSyM+vC9RYrq0qBHACF6ht"
    "Pmi/aX+jj6pRczuXuf/58c7WB/+G6POiLF5+elH26FNwr3t5r36v6zrOqh2IjodLGJ35Go5KUsbD5Kvy2QtQEpYtumMOAxgr"
    "KWoVhGNkxXmhT/JMvADMYoATs3Y7JIcggnFWGoEk3zOUsNVyBbi3C3gXAy+4QjohmK0R6lWSjqHnMeLVbKROD586VB5b18Vu"
    "hqDlBdS+aS9Q9NaVaoywLeRxEBHqEXo9EVVplGcQ+ony5H9R/oChF1XEylEi/bQ6YLHUTqpYBzT7weYBBMHJPGTWASWMh8jy"
    "pX4cTdGJcjh+3m9QM0TI2hOYLN+N3AFYrdNfYm9f/ZYEBmCBeawm+dp9cD9MfrYxzKzogpbDbFnauVvPii6wKO2gWnq+oHwa"
    "tYzsB0luBuMF8s9H2bkP6jFNOx7g3zSFoZJ5ZWSw/QXYhmqhNmoke3ZwvJ/8Bv4yrYyCM4+HMZY9Ok3hpvKdKO+3nNM1igQa"
    "MtakYNGk3SqWm+WVonI3MEjVBPEUVOwrFmR/qnbB9GbB8oihorFgbKg1ftaXKXSnJvXGDZIY1O5YCin01YqwLxVra0jFiICB"
    "CWwXNN8wgt1H5B2qGpvK8TeEzXkKKvYVC7I/VeveIHZqLPLWSqOrfun1ps3DHlRsI7YZSDAxTN0WLGK8JbdI28tGUCML7AAF"
    "BocVVJdKSiu/e0vSm8E2A9XkMxUpZHpPgnn4Iv3VetmN2PQKgSPqPnXr9r1U2NyRVDMRHLpgbWyph2LtRpmvb+zWTHKcJqc4"
    "DvuFM+7/rJ648ZakDxp7ZIX6UD3tldvNPrMV84yrN8+OpO3DdSt3OEjQLX4QfMQgfy3Pa4IR1G6pkDx4VbGuKniJ2fEQRunh"
    "8LzwGJ14m02wfxeAVRT6BIYkTSE5w0zyxZHC8+WaerOSd8uKFMZD73hRwevCQbT6HReV3ZysYxB9eoZTkGbnn/04BXO8aJK0"
    "n6p8ib3Ns79Lj4vxdstKdA0WdNcwBl9ROuaDFa0OAZM1gTlK8OM6FKxSCjkVqYBlM8RyRMvG7Y6AwaYF3kPSZVgoC53SQyml"
    "ovJZCLwIyzxjmS361RV1hMmmIS0qwUfess3cF52RLR0jh6oQA96Mijuky1H7HzhiP1JFW27i22F/h2pT1tsrUMreOWtMipw9"
    "bcOZH0+z18kYUJVEngOOPcNoEv20c2hGeTRcyNTg5NLaRAN0NytSWQz5HKthy+bHDLo9YtSleHpJ44GjKysA/x51p7TXV+CU"
    "vRMtgTcz5ItqjUee/WMgn5cwdelkYNKdqVBKpmxewrQ4KqpJNyKMwOlBLBsvUQAui+5YLHWH2XRHFe2WZQRkyhfVFgjYqD1T"
    "3dCu5+bWS1uOZ+hn0qMz7q6txEDy9OQmz2EvfWTK/LFDyQhEUJxSd7JdICGLLiSeCnBjaoKpBliCQrDSSXm6l2LT5BbPcMoS"
    "NM4pbU2eswwWT1djCZoIKG1Nns7G4ulsbEHjv7LWtMVan4rxlC0oEi3lxNOM2XQzxmbgZvMcFdo87ZHN8+GgzdOw2DwNi8Mz"
    "QnJ4LnVw6B2bjSo69EcgjGY/BBnO4kDSkTD0qTJHhLuGBLMdTo+ouvJJ2pEwVHQErcko5UQf8Avl5EoYmLoSDotF2fy2FrZz"
    "2UkYrrs8w/UqvqNzQyOW09Wq9PMadKNAQvQnYdyX+ZZMheZ8mp2bHRSzfLbiWNmgj5EuSZudxTHkYU5F7i59psywJAzC6Np6"
    "2czZv9rL2ND0oW5PyBcoST/Iu/T5kZ6wN+heVBz71pZQdrIXpk8Oy1ew7k1KqgX6asuoAkXeT0qylSxB44C6qCnYoEtnxerZ"
    "YNnMWj32vY6bKunl/+x5WkNLRrIFLxL0YtDV1J7Uo3p8if70/nz2Nv3Fq/un1HNBeP+fkfqEUQJVmW1BkIKIXLt4Kx9hLnbn"
    "F/Lx/mYLXFFcmXG24dDFXwzk9r4hOch+NHX31plMBgPq3jrnC9R9b0jb46sZyXKmpz9DkJyprr/XHRZMp9PBgMr0eKEWU63u"
    "PlDNgQp3YJocRHi1A9PxAuMaMd546hsi+Bv1Ny37C49sf8Txf1BLBwi5cj/3DgcAAGRxAABQSwMEFAAICAgApJO2XAAAAAAA"
    "AAAAAAAAABUAAAB4bC9wZXJzb25zL3BlcnNvbi54bWwdjDEOwjAMAF/AHyLv1JSpqpp2Y2KEB0SJSyI1dlVbqPyewnq6u2Ha"
    "6+LetGkR9tA2F3DEUVLhl4fn43buwKkFTmERJg8fUpjG07C3ncV+PULhe1Fzx4e1/2MP2WztETVmqkGbWuImKrM1USrKPJdI"
    "qOtGIWkmsrrg9dJ2aPmHKB1WJTYFHL9QSwcINGgDnIcAAAChAAAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAAPAAAAeGwv"
    "d29ya2Jvb2sueG1snZPLjoIwFIafYN6BdK8Fo8YhghsziRszi5kHKOUgjb2QtjD49nNAIDpuyKx6//q1/bs/tEoGDVgnjE5I"
    "tAxJAJqbXOhLQr6/PhY7EjjPdM6k0ZCQGzhySN/2P8ZeM2OuAa7XLiGl91VMqeMlKOaWpgKNI4Wxinls2gt1lQWWuxLAK0lX"
    "YbiliglN7oTYzmGYohAcjobXCrS/QyxI5tHelaJyI021LzgluDXOFH7JjRpIaMAptBx6od2TkOJzjBSz17paILJCi0xI4W+9"
    "14RpElJbHQ+MxaTRrYlx/7hRcpzcRut53i+X+U7fn+zbaPM/UhTSKPqDWrPXu5ivxfhEUvMw04sMEUmnuH1amu57vhvKLp0e"
    "g9kIJzIJJNBMYfMoMCFdqDES/bxTjvEmgY0FVuwp3xA6kxCca5XhD3kArR5A2w5ER6ccCqEhPyPCYT9nkvfOdDxB+gtQSwcI"
    "dfgb/VgBAABzAwAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO9k01O"
    "wzAQhU/AHazZEycBCkJ1ukFI3UI5gOVMftTYY9nmJ7fHENqmqERdRKysN9a892k8Xq4+dMfe0PmWjIAsSYGhUVS2phbwsnm8"
    "vAPmgzSl7MiggB49rIqL5RN2MsQe37TWs2hivIAmBHvPuVcNaukTsmjiTUVOyxClq7mVaitr5HmaLrgbe0Bx5MnWpQC3LjNg"
    "m97iOd5UVa3CB1KvGk04EcFD7MVoKF2NQcC3HIpZEs2An2bI52Twoe/iDPcQg56Kv5o1vpEOy+fg4gOPKcblKZjrP2B0qxx5"
    "qkKiSP9wxPzslmfpLwQbt43MIXvQu/pU+M2ck3gnt/UNYjiQ7Etfc4rH5FYs/hkm38Hwo79XfAJQSwcIPw+FGhQBAADDAwAA"
    "UEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAALAAAAX3JlbHMvLnJlbHONz0EOgjAQBdATeIdm9lJwYYyhsDEmbA0eoLZDIUCn"
    "aavC7e1SjQuXk/nzfqasl3liD/RhICugyHJgaBXpwRoB1/a8PQALUVotJ7IoYMUAdbUpLzjJmG5CP7jAEmKDgD5Gd+Q8qB5n"
    "GTJyaNOmIz/LmEZvuJNqlAb5Ls/33L8bUH2YrNECfKMLYO3q8B+bum5QeCJ1n9HGHxVfiSRLbzAKWCb+JD/eiMYsocCrkn88"
    "WL0AUEsHCKRvoSCyAAAAKAEAAFBLAwQUAAgICACkk7ZcAAAAAAAAAAAAAAAAEwAAAHhsL21lZGlhL2ltYWdlMS5wbmfdeYc3"
    "XN8Xb0IIEr1FzyCMHqIE0cbodfQuiDJKRI8yBINBEIwQJRidKNG7SJQvid57SbTRojPau9/vb72/4a31Zq17Z+bcvT/7nL3P"
    "budG62qrkJIwkty5c4dUTRWqd+fO3XbgCiEiBEbEJb+cA193vfRUIHcqBpm3gD/3HBW0FO7c+Zrw4MqGAPhP/EbV1OvOHdDE"
    "v9fde5m1WXfuvFhQgyoYvLXazSJuZFZq6ohiuWeqlKjkQV/Z+M6qO7ANHLvqaYhO+/BJbZpSl5JCSaGx8Ss4BvUWBiWARb+Z"
    "8XQYVMOziZKfPhNzVZuGEHG+AoFu92PFbqIkO7AlexXY3avFDX+2J0UuW1sV2HJsGSOIkyCP64ve3O5snbN168trM6+/R3Xr"
    "34acB6LJx3tIBc9svZa7zfJOzswQo3t1B/pTlWKRkxVucj6G+lIgJAUVamHQsszIvUND8EXUOOPu/RUlXvKMcnqXk0l3Fzuz"
    "/TLE1q24SN0TJAXEdjTtPJFWxz2uU+lvmL43027HT6WAqNn6okqHUxQFhGgLI9zoVT4iZSz46OdyRP7O5JxxX7JNnrjzd6xD"
    "AjEI+Z4/3rRJvOLEq2l6IJQ+pnTO6Tav/nK5ZSBBgQiUnzRuvq1eJGcyU9ScLz7bMN5Dhvh2pQ6IVdfxWmb3rzbHk0bgnLd3"
    "L43e6wKz0RPpqZUzdqm3Cs2S3q288Y2b3B+oVANwPmyW/XK4DJhOHPsjPic0wajoLtwhAkJ+mGeQrnUrHxHcqA2vNnWPS7UO"
    "Kgi5+QvgVxeMMOJLOzIuOlw6roOBdRBr59U1eLdrRPn0iAZuhzt+yoMQgbqk4yaDmbNMdTPKj37B2JEUSelf8c0rH70MlbIw"
    "48/YtU7LUyYCcVK2LqcM9SxXmyEOowYoOZEUo9l2Rh8Hj7SnS0YPi++29+pYjjEbAHgf1pedpqNHD5tXL7Pnf/2rPdvxPyfO"
    "A6cvpj1GnacfXaVJziIpRL7DyVhJpdXNZMSqyQANmvjCy2EutfTfyJKMjOPwpXxXQz7VvKOAKBPFTYKq9PM5enHfXnmJU0BU"
    "u0sa2JAsIBZ/cbMHLtLZdfWSzwDl2G5KuVvwoth7ta8wkhrAAPu4Oen0CmwWZhzGjfygPuy6QC4tpZ4I7CjbQWYIEQ/DnI5l"
    "nXiFGJvva/Ssa23qBGO4u4QjJQiJRo6Em6ty9xSEO3kyhVsLO1KDkDTMWXkRUiLqv83c2dwHeQF2sB2DS3JA/Uz7eKYlYLOn"
    "6fhRnfGxo7Xi9x0+1byngJBA8k78pm1gWH2A+N7AqU9VMm9j2b7YvVsEZ+Ph26v9iIaR4M5MH0Dd1AOnNNMKKd+2yOXIgafP"
    "T/YcVzpSvwcEVkheAoq3Yyyn20lTx1RjWc/dDX3L27jLp/2OGw6nA6YzIwB2yoFThkZF3qiswduu0KiVpbIduD6r5dd9lz3L"
    "Q+O2gVPFdVdO5IdvouZF3+OLoZ9Dfu/QZt3WCAlGhlRWH1kF+He333q+IbwKLPnS4P7M8QkIeQ8jWyu3807dQWJz37Ce6fhs"
    "1NTd2n5QG0J08Gs62j2SRYd416FB/3XknMEI4uZN1Mx2SFg9T0HcpOQ6P7B/dYJWvk99JytC3iy3Xmo23Oq89n/FOl8Rsyhz"
    "pFQzdfsqYWmBsIGv4Df3rtOgEYRIra+YuMwYfneGnu7miMlYwqmFdzb2hs6pZXS2cVNg/csT5Aca0dkHL7n19SjnrpYZG91S"
    "5veipj61IcSc5vSkbxJc0p+lUkCoAAPflU5NwjUlF2bs3UxkBl7zNyJSXJdOa2fmjI9G7CrD51KNlP1ASM5C4bciOybqO201"
    "6xl7wc+dLnnhih2Xubsabl6rpOv8XEgKFUZFtlAW9b4429/4QhHw0mWupcZfO2tvrPzXaNMKr+rJxR2fgZDHdt95Gzl4Gval"
    "D53HJuzi096waRoThHiP4DsICWewvKAt5Guz43Jz1naGEEXf8XlM+2aZ7iqujAldQXL6m23M4Yd3ygcKCDv8YSaRdALJ5nfn"
    "QCP3EsyEfbaFPJzPJu86R7KFGJQfBUueCmdRIZhfdtHXPGMdOJX7RDiWOjN0i5fwatc+jrdSZ/rRdGfmGuCQSnltD30Zaf8i"
    "o8kz9mTSIzQ+fj+2n0zhGGphK6xEMr4LpYBYdTeeGlb9KEbF/glbJ1G3drAcMWTSxfVPM6fZD/pC/kE5BgVqTMMMugxP3+xf"
    "pUztyBCv6ytLyG7q1Ri/znnBa5/n805ymxjEka7sQrzDybtqrN004739+gXul8bTbBEpvxtbgDBoJVWiM4iXw9qfCpLP8Zza"
    "l4L2sMQ/VrNhTzmFy1soGRH9g2OI/92FQT8Qz7xOqSHKzHmFFfjSkTGfkCnysvZeKQd9pyxrhpLbu26afI+R78HYXe6XRbqg"
    "pZA/YoXSCb2mxkPIi5Q+r4os+CkXqzmLB+2vSk8qiPKjgcnTqmfFaWOOL9ybj94a2GQ1X8RGVkpvfl87+xV6YTi4vtyg6QIA"
    "Eumq6CqMvTvkS4bLXxiOXmtbD9QSQVQLShrII7Lkuy3cb/sKmXR67oY0P1g3J+y42Wj5ZcVpZ3mxq9ueUiP7fKaWBCBWYJEn"
    "uN6uq9GU+PyeQSfVD/n7h47A08DE98mzz162p9SKO9OOiRCBUvJnnH/0WCsg3h0KhMvKXxiNOlw+GZMCHtwTu4rBhu/Gt5Gt"
    "W76S858X1b/0zmnlc3G7UC+MoIDwKnn95Zre5NtcCWmZ8l7aSgvqffb0ll0aCBBau2PCpw0N36/44DUnaDjgJOqhjEFKlrWv"
    "Pm21j4UfClW3Dy9RrFmb+sHLbcxIQPlUorOKVV+NDbJdD8p/2jAoeR+FF6WbXd5NgZlu77ZouYCQBnjjw3zYQe75l81ucleY"
    "66f8RajPIseokjq2+otOsarM/uZd78FtDgAKCCXKVjC2AOf2rfrrKm9T/zAjF0uE5/6XC4Bq1tXa1YwUlD++qcdfnGg8TvHl"
    "6xkFf5mqMwtCvyCFN/MJEp35NdzO5ndFsbts175EBZmddhvBR2PODty2uabb9WEz4HNrhpNf8aM+TSoS8Tw5CY2o8jAeV/N7"
    "Ydfs0vHLupYGIK6LBzktnIryV45SQdT7WcPNvylb2ihWH761cKC04FmZDRRsbfVGarnoMqW2L9WyQ1RLnMi0Q+OaMRW3dTyF"
    "ZWQ/OhCVjdvjbicLfFoCP/6mI7g6PflcvJatxmKIQGs2F+pVCsbTybPkGa+lHp79CqL8aH6v7jXoYiSC/Lsn3269RTXqg6cw"
    "hC5913cl6elSY7PCGl3nsIGNzeJ0vJuElgs7Eh3mER4bJfYnJoOgbN+0FU5atpc40J5XzScEyudnruF17aL1vdpqyNiT5Kcu"
    "HIhc29hCbjtmzEas8RXRhKhH+TF0JH9I0nVH4IuvklyeZ0YUBpKR88rp1KzecKuY0NDXzeDa4IbID5BwxjfdA1QGwZ8diHk7"
    "DgFl3HFupT1V7LVwaV5zHa670ZU5/PtNbsHeOYX030BlkqCuQcTz/ORUturAaHqbNe1fsrfBnVQmCS8ypUOiSwzFH85L4eU1"
    "jt8PfmVGD1QdorNC33rMbY7uTHCQA7e8UaD4+DSpRMTDWbxiNB1H1jzu3TtT/eNA4hvu1RZ3wlHLXg4fU3xtzgQvKD/nfftL"
    "jWknvp3syoDvAJnnFPWFikfTyYh/ofiMZ3++yksEr1vbCWysnwgUgRF+Sio+S5JBr2RSpL39gb5YNqV+sX2gVgSimutUbxUW"
    "V/99vQ+IxA8neR7U7bxQ8CmtHTjV+48VT2yaJHDael944Y1spp9iMMxisC34Y4mS7Pvbhd0qq5SPQIbuqfcLi6v7vq71eXtc"
    "ILH6iHCibFT8fk/6pPJ/Cbq76qvRNHvVaXuyzQKkuccqWXPrIgxrBM/Y5Smso4SQKNTIkItPxuBKqN1FF9zbmRE3+FXBoOvT"
    "pT2xgo2f9+ObCScrq4GSRMuFA0mDN67Lt0PNQFgQv1rQrWFzDZHa6O4uPA76dPALz6TjXwsQBVo1Cbvm04oP5V615eAiasE0"
    "Nuoq6+LcHX21UhDVivdV55n+5Gv7a8+4taJOHq6QSSq5fD2YRXB3BhhN9nS5pxupEPWlKG/vVyUajf/c/2TXvGD7ZuO30LeA"
    "yDez4TOEf/Z6u8yVTYaxDmVuZiwgTmPsbhNAapClrCuqK6NoIkon0HfD5FghPherVcaB/EPinEPrk4XgzhBsMGm9/sv3EoHA"
    "FSoKShTQxoqdMdpUShODCO6IDcbgytQS0vbqTTTT6GRjflpXfIKr89H1Wl5RfHG3f4L882vzFGWWaNO5WlDYIsBa3C4rwDs0"
    "IDT4vfXTdOOftsITSohnrnDjA9dM2rp/vtTpNBoHl+1LvlLn4SoeTKmbYbU3YwNx0mPxe3/0mCqcSf4mGxugS+bgmr+/vfuj"
    "8npLW19dmMEP51a1mFI7I9VZKwtRFRQ1S/0GEDsOCSjXfgTrtiFMvPc+ibFNtiSJc1k5/ofHyOHIq8ggf3JqA9wqviA/MCvE"
    "VYe99tvxKQwUEJdmDsH/EsLYIIkzSw78evFXeOo7zlgPohqfo2yqUClk0kXFJMQfxDApFzckZGLrKSSeOn9F8yXkCbBaImc0"
    "bS0/kmi+9Yv83abLUKy1GjFIMozRvjvIrjNM78zuyF0r4jKxi5S8cfyPGZlVyLD9GzlebhAnt7KWFbl4f8xF1E82iQVnyUZ4"
    "z9EbuaS3A9dDnyx+duoIBlPsqawxh3TiCmmpICL2808yOpVNhX8SzSZwCprptb6bv5a/hEmK2VKBOO/3aAc9c0XTek7IxLB8"
    "7epxDiTrvI7+zRuIPsJU9j7c1SmYd9sU+A+nQPg8eBrM90diqH3Ou2httZYn4TtAqiMY8M8fI0Hlaxxdo1yK59XwfU0+TuSr"
    "fqLGvym0WWHvTrnjdYbMX3nysTmF8D4RCkL+uScUFEG4q8P88+FstvpM5jFojIQY9Env4MESd7DUgHfeGhRfI+/a02Mq2V28"
    "CZ3sFqJ7ABY8OeX+D1e9WygIwHNtHvH0vE3uw5dYOh8+4kr1FAncY2+0Tkmkl4V4YlbpZePC9HyHg5tMfCHQKh5061/KJGra"
    "9r/w42TdzflGRyuT1qDTo9iPcrKOK4emh96HX67Ovt8pawnZUZnqvI713OaoFCm5zr083YmNNd4p2zvSvLC+JoUMUluE3PK8"
    "ueUF3W+g+A/Q8vZdR854rLVV/dU8vOZjZc3+KxOjSvfLU9qfRUN/j0o9lj8W2TDAPnvtaiZLsTSTZwT9EECm3Yg7Wl93xSz1"
    "06cuyNz/aEjsv8IBmFVnnkS8n6TR2xQ3lZkw4A+/mFzqYFzfx5p6Cqs+6Jfor80U2aJMqywzD7ktXUuQuUzqE5DU+MMec1iJ"
    "/N8C4+arJ2WTNac+ZZKz+pD+ko522ZOpXWVPRw+Q+5KbtXXe1dcsPZ8/6Yrlw7TlBCWf370gHdObAQs+NFVW1X6f1XjXFf3k"
    "TEKfVWzBIWiwPYDr11iUSpZfB88TVm7NIKL1qBd/Ml8lOp/KHJGlZZMPJS2J8nBrBz1aNw3o1Zwdr4VAVAsNvH0i6a0CZrwz"
    "Ds6H2HTCLy6OWaKs4eqx9/sffPpDu59B1sS8vjaWFd/RiaYjZxFFHvK1WWSF/udUf7hsfuBXMykU23fbD7ThOg7NxRR0sVM7"
    "z4jXab9zXpbntKoLgThVdPP4v17eHmiPWTYnTgV9jSOPaZDfcdRJJzva8t3eb9KxG74ePMRMuc1fpNjptNCu6zv/z62iEZsl"
    "BIUt3vhriSn7V9qtU5/BCa+tTTZvYvweJJS19voAzVF38/nCdVfs02w9DaI+ynGwhitz/QvRq4by4GAeZU0tdLu5sWgkj3p2"
    "gfvHJVkycMK/ZInK28LTTzU2T+U8yNURm5Jn314c/v3TyVBoNNK6903DJsDnRGaBbl7mNOz+S9yL8ebGw0S73ypVPF/iJpnH"
    "cESgJl22VgrxQSqpkz3Jwoz363ufugN9fJd2Lgp9fR0GLYnGSiWG0hzrq732v6zunlHHHzGZN/QkvNQ32d69+x//8KMrKdfM"
    "J+O02cm1iGZvuHdjZXHo8584TTuT5to0hWxjryA9Q12c56xs/39Bb9xcerXHVNAyXI9+5fLjQN1Zlqsz/H6B/mWuf5630ptj"
    "99dide+5tO9lQMzdQwb/Y3mf3UgMiDg6/sn6bCGPJY4tepSub//2t6y++a/8uWyGj+t/OlV+WQtzkzWqBPf93KDnF66rO4qd"
    "fJhW7z05FJIHRyA/zMvH1bxqrur5vIa6fFT3vfP7E40oNzafEZPqWubQzw3nH+fAOuh6byoIyR3nzCeecW24EO70gKGap26U"
    "C8U6yOTF3by31wqvf5LQKpAmLSRMsLBP01hBVO86PUQHqP8IDWn/6R3rq9vMHriipZOeTr5xSb13OeULW+rl+ujKmf8T/XDB"
    "Gl0vKj6rk/G/7HfpV5X4sSt4/vyvRVrFPxUvFtutEwYubscYvZ0j40QEt34fjkeNpvQ4XLoXTgCZp2AE5z6tqmErp5su8vaP"
    "3M7PRz71X2H50VgPC6/9GxcejHfBV8mnk//l3x0JhBhAu+wr1zroHZVGEBtJTsiTLN7+9nDgZ4MMzuzI+oeCf0p36p80pYPJ"
    "HxXvttRXBMvFNP9NpWTO+U/+Xsi67G/xKaIr+Pf81lkcdKwGJs5kGq7Cts6ll6du3nL/KuPH3vkI9qbhUzcGYicBJl9YOu8f"
    "lz7QUo9pXw7v5Tw5uvCvuLnO/y1Psm3sBKoYayVjyhNaGQv5Mqyv1rQcEnA1S5/Wgme3G8eIJdJlvtNIUsVNrC83/a9gabsW"
    "nDZ5sFN7P1ti4TFhAU+Z0Fn075JNcbCbadzqYeD6N9GI+uAD8f+I8Z2/EIhjbd9g/6jYpFeXln04s9uBN3TW2p+sZmSLBIoz"
    "x70OH6UxA2wFf/grlLT+3WkstmPvRv9ls8gcMNNfp4EpcaM0kw6Bi3JMDKB8k/cuKQ5u7c8EG+4XKB/wly1bdfRyvEi51pBN"
    "GnOgKdIMuFfIkw/UXIUpgOzuc7kKYia3u4jgUATjD6ChJCgsCRMfrjd8Mw/UxzLMg/MiWi7nV92VwuXvKfSe6+YJ5MBydssw"
    "PVs7SqkikDmjbvDwaQSaQu8lttzR5ofAh8Bsm5A1vx+Ka7E/GXpTuqyOlwCwg/re9U8yLINWnoI5ohHVkv/sByJp0ivrZUmY"
    "3ENbQ06DxnZJHaN+v34HL/0jOfVLNLAl6EIOKMznDPqGboKaZ+ZW5i/OHzzVymV+5kE/1gwUnQRii3gMBLPtzhVtQ69qGn6m"
    "VrZYE3N7cj8lqUXkciLRd8bfPlg9lctyaG5ZcbpSRxfGLJ50PeCDhpRhxtjq7/Odn902kRXyvEd+oHoa8Neu6i79N439gElv"
    "vx8XwyNT30UNvX8qClnfxP+oRAIs0qwh4ZL9P3iOw1oGJh21ytiRNJkpNhyC7wBdwH3e47lZE7n0fnGdpQHl/xXlHowilVTx"
    "YRFnOhv8TD42jD3/lnyqoPXfoUSfoZsiIFI4YNWoSnhjE54q69zctBKE7WKyYTRj6bgqyxgYYR0cMif/IdokVXgiO5ZOBCq4"
    "K7bOHIfnplPmSON0aQG0cB8CthuqFDaWpS61RRbm3li93bJPizQaoqefVZworzuj0EyOHTi9928zVWA7jjx5mTfYKcOBJ8rm"
    "dXP9lLT8s1z7JshrPEr/1i30o/3s7jypFh/QMsX3fLOrsnF0pXEKNLPMk59zdN82HQsAIF4dP2HrjZM2B8bGsIwEs7Je2+U6"
    "/3WHAsUejpvSHV4FW5ORz7z2IMAkpj4NCYiqhDzfx72+y3VCDeGVt9uUrgp1/PNTI/XTa7fnGfQeN3Sk2MKl1l/M1yv8gWu0"
    "aVzPMsJV0owAPJmBSb6qXMeXNHwJvXAnHyWNZz7ho68w9ZQQ1Qineu2IOJmq7V25bm/HR+hIoyXw5GwglSb1vMOlUGEd0LAr"
    "kukgJKDQteyneJOYvGmvymz7wbAVi09TvCzIDySjTtmkTEHy1la4mK3ByLxwK11GM8RScCPi29hqamqnqEI5uaZ4OZJCT1I3"
    "j/wPNpDsU8XDH/uBHMYyp7AcUYUnSIM05e1H03/xz6P4Pj9deERYkMvRC9bOugJrUnetf2Mo9AGm8MD5Ch1vQAzieOdBLUcs"
    "fUddJX3Ad+zf84o1eL1VJMvdAo5u5cBPcFLkB0qFthPWaQZdKT3XelknM3zQq65hpyUy6XvqZa/tllnHSIhAHDEH+lPILLwC"
    "Q79xpM9/ZzHPGXfzXvbKw+1ZpAYzrwEUjTwZej8KtNGvAJfc/50NwaiNQrPuFRjWLJpNZZ4AJEONk6ZVeTm1IvWCwo7SIOQj"
    "LL7GmygFeN9apgwERTsO1zXzJQZ1qbfV0e88BwuUDeQqDsxoEYHoeur2w1mICuxrZgJGMyeQFFTdhyk23BD4zHwwtnQ9nwv5"
    "QfZ90xv96URdM7pe7aAyyQwAJvEr/v1OsohRzBzzi9HMboAr2cGFascELFAaNynqdUoFUYVgd4ts7ivCR1zrNz7VfKSA6I26"
    "CkWwkBYoJSaPE4OsGMRsfKZndM1oOt6suAzMqBOBPk/vYhq+kUWNti4g/i78d9I5JMoV/ND3BXrd3m5ZY32HE/lB/j0LM6n0"
    "Y/Wa+kWp75kpgOS0nm9fqzgwonwZ+HIijhwg5CsoYvntNAjGONuzPLPODzC9X+lOcHCruab3OgV2mbYoYfBDabA6Ys7pSC2B"
    "Qo8ivZzDRkWpROiEmQNo90ec4kJZ6Arg5Q1+Eo4MICTNe1z33Lf42NH8ybnsp470wAh6QSjYsnwmgMDXFt1cujFcseH++r8z"
    "67txk5DpXNjskLKLmlvqJOO6JjsQCGDuMiTS4upu4oSOGTVRFBD1mMEWCAtzwQ9XS9tBRoBN8uR0442KSkkihwQxKMXuimB7"
    "inFXqWosj2Nu/U/cL1MIETBaYIZ4KbJztflPjezboxivMGAn6tScKExTjib+3tvVyagJpYAoUwD10jTPKHbWaUhMJRsYAA9M"
    "vp1WGMX2RdgBdqu6Hjd0ieuUtnFU++med/2YO58aQoSZWVSzIJHWMdPeLrccGAHUblswI8vii6l9IU544xzPBUJScGMdK81d"
    "6j8js3gFKgZC6dd43C29aIhByC6HX0ru71j4BT7nCQYnSzojAUM/Cxg1Gm06vLQeca1N7aNeDPzvpcdq1sLcnEdfHl5xb47k"
    "wc/14n9flwxOh5YDEyqyKdL8UUI93Ks938dMACz307P171ouv2clfmin42ok3xqXB6+0/KrUJgJx+nmHeechSHZq4Z4Zu8+c"
    "pcnL3YLtAZaI/JnFiknmVZdhwFAPnE/UpNOtkykgqvBpd5xzHe6bbJVX0Q83MmYzJjkfraNtSkDvYozk3vPLwj+kLhMr2lkn"
    "me4BMLRFi7Nlk8xBZFYPS1293li57XxrXzd+jKQQYR+hMB+YbKpzuGR4+xVc06xzEnrk0kLWtrgjAGhoNNyKiGmX7efYiNZf"
    "W71XGeXS+Xu9E9hdifQ24CmE6unrfoe6nmFWTw9JDUFUAnnL8iu3GjH9z+s3ei61ZMVihEuvndWIQMh7DuNIQ6OuUtd6OQEd"
    "M08b1ywDpinWDw9GZf6Grdm2Xef8EeO80RN0DoxpGWdMbfVTMgXQiUD//95ifuFW61eLuhBd+L9ujZe4kmxze6EJA45QGaOs"
    "3yceHhqJ6IRB4WtMca+X0NGSeeysLDGI+MLvwOIAdEDr4Y+XTStE28Gf9OYLL6uglVA7GGO3smq5WrJayN5H3/buGtKqZsGJ"
    "+T0x9GTV+yAK85kjf3VTLY/lHflGtOx5GEp7UBS3/tYjHtV/uGLrEbNZRvmZHeN5SJOl2CbLU2jvdUBGOdGZmhCTTaf4NQIV"
    "nP0iC8TKvd4ZEmMmK0UnO840fQlFiDzs9mLHWV5KKvBT98ZsSsSXtcscRS5eYyBJ6XSIAuSh1u8FDLkVd/tjjAiklL6TpUs0"
    "J+jnslxX/pW2BRcmSLGEZb0c436MF97PFxMQjyVVdqqgSG43VCv9rBhCiwjv9zygtjWdlodefujahrq3cgLAAWFp6AkltKOZ"
    "H+1YVQrqcOCHX6yavH9Kp2XX8y/89Upo48lF/gpO3J01a+gJT8+UrN9wsE09qv/n7JgSW6sx9KLFhEv3emyTQEYBtg4fdzdU"
    "I4Klxqv5h6MOf84VdYlBjSELuq9smNu5MS8OT+1sGY0oYe1KJ/YjU9062ecwBv9LOmHSWz5M0OXWtpce3ZUyMDlloY2JEBT4"
    "qJbLJhR1qF4mO8K1wk+5ZsxpqRJpxYM5VjKW1qU/jra4z53zImb1XRqaBVhA4kSnInx4yh+8xO6PTqNBZ8dK3al42KoMPdEp"
    "W+wqWlemtP882VhJV0kJk5V1jfZ422ZyrF07Rx59bcKLqVUytqY9VVQxAn4ZqLBwrYgD4rwl6JJzJIjBHey4QGFB+OflJdDH"
    "u21wpROd7eyu8Zz3iM1RG59h9zAwuXWFQLOIGdSfZsGnG/xvS5PjpGjtj2SeoFdnKWGiR+swjCcOeQy3+LPUoQ55q285elbK"
    "2ft7u0X+hrCXtgjp5+qIg/Wtm3OS1WFm2c1UODwg+Eg9InSzJnTbvPNR1w9t2pixqb7py6iXRbasXmxsZ8qwoca/FQHacszg"
    "n4FCfc7gvMF8GE3EhbYwCYI0bPRiMuJICS1erria2iXBcBJ9mGqLonQ+i0FVt3Ppu+W4IYol5LLsCsZUjt9aPsa4jZXSu9Mf"
    "w/EKhGIrcmrhE77lvXobvJWO85pXFZuXJx3rGyTq+6s5Jt7V55poEnSuzhBdnjw86NLwK5tKLNiNyqc9QdkjOrwddygmREzp"
    "PKzr5EY9hbg2Sh9as2U+pk2p9A+V6h0mQE8qkZp8+DjRJdrRsE3Npp4BLXeZnyqTjm0fLLmlukx0OA4YGdk0LF5sn/ibEE0k"
    "6ZNjDjWyL8NRSsacxYItLU4EuXf/LJiy90vHSzU10NdaMlylTMdWLt/1EHrb5tSqgna3uIQhtfV+m2OuLMeL1oLPAw7Vn33X"
    "k9SCXvTUseHuxYw2mbxaFMRIRT31EOXOMkaKHAuku4jrQ+cin5JYe+WxUiMSfauU97c/ZkEB9B37alR9bNkNl+xfynH5jiWl"
    "A0HKAq8Revip15dHET3SSlurxTouLtGyfz0Ux9wb23qOlLYs+HnVXg+L1vPi/OgtiyRJ8e3pvod3vR6kHvEFjTe5wM4fyZw2"
    "bjIZaDO3jNzpx43IlIBDjE/EPgnJ8mNEVnGwEW449+bppUNe52NdKXrEgQ/l2r6twQeCrEPkaNeEhNQ1ccpv5443tuC4WLVC"
    "rVdK8/vLjIvaCiFeYpTgPArncwzZEuWIeO64ijuTbNhLP7LP7qJcj3mHJoZi2uUvWrDJ+ZRnLVpXljnkxV+eK4fW+3nIjdAz"
    "8txLGRlS6QK2bmO76cO7vD8NLVdAJOie7udj1dWcfWVzg6QjopTXC0qp+yWVEOsHOe8hJZvvyI8bPmB/N5CycSRY8UU+3n3z"
    "0nrksAME6D5I/rh7j52XIx5HIkkDt1v4V5Cw9vMW8OiLa2jE8SGl2d6GLXiU7GV1LkfEyKkgHa3BFPjkbnF4RVVovQdMDtHF"
    "PmAoHo3b3HcZN2B222RHO10+54r7thKmraxi34zd/l7EZ5ZzZXmF2nzAlgdGIPAhTvMGBoVl4t5vyGcXu2s3v7eoZR+Pu/zr"
    "jY2Rf/wpU7ZKSS2bSdz2bSmbLKE+dTPS5MJa1vTSXN6fvns8OXk60nPNrk6tMiVpojbc4gMgtQsGwV8bBVk6eefz98vFXxdO"
    "5D/RciycVfVa6RsLTv6yxI3x9nBFgYh9VjSkwgynavgY3WlglnkW+jsEbhvO/LqE4oo9ww/h4+N4tQjm5ArZFfVZbtks7118"
    "dENZ+3DfDfj2eeSNOEaxDNZUMSs/R2eZqjYv5U5bEGfnNGpxjlFmxuxajiuhYpXnJYQZRe8kcqWpdMBGbiGpEyoM0+PEy/D9"
    "4KTicNvhjhZnxby+nN3f00LsWmajkcWHnazs/XPd0Rgh3RsPzRc7koTwz7+8rdi/926ff5nwLxbMqfUY/4bLVVhrLr0rTMHo"
    "RwdoZaRRHXpSy3BS9cFhvMqrKqzRa5Mm7k7oDZ367sPRX89azWpzx/h005/jmAsTVC8DL1+Wog59Rzaww8AyCk1J/qE1qCgL"
    "CFV0+uBXi4DF4WqHH6iz9Sp2Ded4WlqqQOd5MLuXktb8VGqE+dABN03qKvNe0PNQqrkgMwNvbzpcINEJmaUQPxyy5OSqySS7"
    "mhxJCBgs2P3wbv+cYfaSHOzG46r+UdIuFJBSnwGudaOc2C/yiG8I/q1w0QXLE4j4w0mZsl1KWlseGxWrhDchrD5H/UYCnFWG"
    "iGbqVnSc9JMPIbdx9Gc5ljDOG67LBNYR+UoR2BPYQCW2nFbLk4gEbc8QYc4AL1S/D23uEudIvdLoE0JB9t0zF0OgOE5ymEgw"
    "V0IuLufsWj5s6/BAXcbohyMbc7wwsAD6S7L82Etc6aX2ng2l2++xwSh1hOG7mz0M4at4zCZ9u6jHgxO7GvIbcyBTbtonqoZ/"
    "Dil+/zyy9W/Bp+QC7Bj76itKN2vOkzrLDS4xMtnXDhAromBejIjlRGI8ig3lFEgfC5k9kkEj3L/rsme0heMF82N2g2au86uL"
    "UPUeBwm2NLIf3xJHRFZC1ppK6fnZNClgN4I98YXD9w2UmYcoFCmdk44sAAlQH1ZjC11Zjx3622c8TNHFL4zzCN/MA93Q9fHJ"
    "FruAqSA/xdQ/VSEeKiQYaka3jR+VqIV8U2WVqdkY20DKgti8yM2DgU52IDBkeW08QHeUtSvYd8FUOKNe5qL6vQzvu8/Iz9HP"
    "BLubBij4dwn3kZwK5lSvnkuNEALOuAfm9k9a+46nVkt1Rskap0lCiVjjthVykja2lN9a3cQtztKycWj55GpZAx79UBi3E0CZ"
    "QttbTMYjiMOJuaFELB9NR/aYKsFj9gKhxp8pYawex5LVs739+1JOg5GL7/NEBWetshyiKwvmIM5P9noVkT3W0Dkuw9Sl/ScD"
    "IvwVsVY4FOqAQVgAh1tmp1xj66VWst7EEhQ/pFdsNbp6h+qPySNsjI+Zw8YkjAxcr9Na8WK8r3fimMpGYM0TiYMDrD67ymhj"
    "qhCB5kM45RqlCin1Bj7bzxp4CFV3Irv6hi2l/Alv+hzqh4khlEkl+hE3xi3ZqWV7wt8YsJ0Pe/4DqeW7BQku+0kaHsMmKPBQ"
    "o9SQnccfTcqU35iKOTo8AyMLn+OC+vD8TCFM0O9Z95eE6r9t9cQpYMEeE/5RHvL+XScKwlvjDIj+FkUgicE6jk4Dccdfv3kT"
    "WYKdFwIjqGdOtu65uerv/+MKhADGugDB2cWJ5/kM1fPl7fXR1B9MN1Jyg5I2GXk4yIqZjRVXfhu+UdlQJQ9rR/y9MMjPW8Wo"
    "Mn7l3lMqomejP2n0GU6h1ZVVewHxXkvGWD2FyarN6+3cV09tgCnISr2ngs2yk3Q8W+LDxLkpJ56GUdJ+Ex/wVMNJ8ZHBP/6o"
    "KsEP8/6qaqmsEjUdTxnC5dyqz7bd55PxdutaNqfa4qEH5Z5yBxvGM7mUg9tHNI3wdsKuAIuybiM6J7Jd3B60lRNlJ1afG6rd"
    "Fiq01DSemPL3ULjIh1mHYJk4yjU20rAHbdCHws/zluSllKbuKcGPjC6Lg89NaJ9oRIWKx4jEcEPRPRPFxtTwQmFOBqOsk0PN"
    "9nE7irAGYKfT61ArYpaqdJvsnzpbNr9O0l7V0piFizonnddlxIArRTal0nZU2NWzG66tKKMaLx6UV9jhhLwjbGsTB6i7wKNB"
    "y/tndkUMYuneMXf7y/7BUtuDj7doUrZsR90FPN42oPoFq3ECfbTo9mxZ5UfbZ75oP+ZkqrNnmpTvvSroT0p4YppRo9CmmE18"
    "hFvcoR9sqeo6Vg7lxJHBFHstv4UrYRu5DB1tSidcqhyBFJNuktqq6jYJ4jIVp8kRPjS2ZnHuv4OPZs2gF8dH/MICcGPIiAQ3"
    "fFoJekE3wCmRBc5ubZOvzaV/iuZSX+JbyROxGDe9f4vnRlHwAGyJFlSwH0Py6I+JqAHknSDhB7ozgr5hnxzV0odfgybcmzzW"
    "a30aFaAXJ2P3wUvPscwrtZtaLP0tRNizZc6VmdEgKSK4wV9ko8vvLcx6u5qXXQueJQne9UjhdlbGonDOsSrC1Re0lcSCPVNi"
    "C9LlH8UeS4H7ptTXr1WPSAUqJ3Ln2Et1WJ3tJvYl1RT8mzfx0e7MlwjWaMXEj3ajds1laTRs4ZEQJzsJSjc2e6LRe8HWwvZq"
    "5Iy4PK2s1p3vAezqSy6ZD/TSRzDtfZfhtWRig1K2iul+2/4Ytlo2Klhb9ajPTXFcfmgGIsJOQwre/vvhCLcUm+hu6Sl8fL/Q"
    "mm4s7HBwTpKdfU8xKQbz+uKxVs4bljG8W4oCF+8Hqe8KmrWNmXVZKyK/byCvg7F81cMam9BQ7B9nyuEHOQkj0RXYbggH5nw5"
    "kJ03L/upmRKed1Wk4M37/jeGQovtI/tNY0oGVkRi39p+0hDQxzCOX1SyQ4ROnQIzPIoSuglStlZfeXQjZ5733C3Yf+RH8Dqu"
    "YFqtLE6yInekeDcqBj75HoJ1yWJRyva4nzyKuFo1Ez3OqB6K3mA50T5mwbgPu2Mx7t7j/M7Wgs6zH66i3sai4jUYxZ4HoRr4"
    "pFTvsdS0TjT4ENW6odtAh0Hh/d4p9yAbfzdF1Bmzh+2hxe01N1nFlUk+uaZsHhzRTLhBa2nxLA5sM6qw3XmSKWtAieZL1p9Y"
    "WUts8qcKizdZqHibShxS1mq8KdwUIcH9/6x9f/9GnKJDUJuuKPYslpyB7entJFUHqEKALs6yO7mrDOgTivb8Mm7bVZm5qUT0"
    "3ta8RT3n9G0gPvE9YDx45eGYeO7odTBnW6Sze3O9/+y2sNfB2oM08dU7q83XdKmon9S41saOtPvcKC+ogRB9NqUcR+RySPfG"
    "O/No/+pEFPaEHkj9mYDb/eLPEaSP8VFCG5F2b3e1dHWzS21cNhm0KgbFPcLzJvMeP30UEwg1KHQXp3GnHDsuMatsVn1xpEV5"
    "vQyXUX04IM7OK12Oo9mmNohMl3CMLrZqQYEU7LHJb+UWhTC1/i3N0wmVKHB5+6e/A93svHzlLjHt02pQe2+gA75XzD9B34sC"
    "39YGwzh5ZoygchU6bDSKVVmo/qNdZ1szIyrYok6Mf57fRPFrkKQm1D5Ya6d3gl09o0Ewx7tUyIi25hxBqMRm8Rhz1TLB/Wxp"
    "GtfmBVTByXL9c0XPoEkUiOvQ4HGKxcc4iDULPw1MMnd8Sgl+nok6vI5Lkyi2iu364OQY2e/T1KeK6xfMP4sBasRS8h4UeF8X"
    "W4atzQlKbLeuwRKQpeS5U4/mLKnC2kHjTMxUyY5Kx3a+0gpXjU52tB61xWy2MJzVEkehvRzl2sI23oHrATUmbiiGutsYilap"
    "hFsdWIAwL46UE4RKpz4Bce5TKRClFoNNdqKsxmFm14WfCREZOTp3wPtW+zQJVD7vHtiywOR0Nvkc0e1Tc135m2U2aG69y0fZ"
    "0+0NT5QSQoOjFP19g0yvMs73MaQsv0ZphOhhrG3tWpzxv6hGEoqxnbFST2HHl9ppp2QHZpRrc7sMq2a2Fj2x4I7X2XTJYYfm"
    "XOsDEeyl5v8Q1n1C9Weynypmn95yREjxMqS0taDq/Xzsv7J1UMBmzGPVRuKp1mbKZUHbN91K0kcdseCh12dUyTlLYrBgcUTY"
    "oc6zb04VqEODOp0JDyK15xZQf99WbU6yTVWZY/xSNZ3l14tgTKs015oN5drUgbGt2TFxvbWkmrIR3DHukEQ4jF9QypZ8QoBT"
    "qu1eyoW/sKhQdPEOv/xaacDiExvqr0tzJKncI8Gt+lB/v4x7CDOo/8xV6Cb+HIEEnA2zJA0LLk1Q0LjlAgQkyTopKshlYo/I"
    "bo0oU3K5VoQp1yY8ECPc9UV4Yn0NlZGow9GpLv9w7AnqUdILwjFv+TUs0wd+efsytU2/kaPrUrMimxxPqWCV9w/6Vc54nL9l"
    "kbtms3i/1LXAWE1hRUHW0DYcJ6KqbfUKvUk9R6QsJigRDWYTGziLAGOwldh74883WxBoncVpwX/kaBTD+oNgD+PcQhLoFI/M"
    "KFOuA4UFZ2eULY9lwZHfxGkV4y5ng+0o1/KWFmmOJSZ6dYPRlen3FtzRqbJPLv90KBnEq7MP71ttqjyYWMmpBpDLSNFRjAF/"
    "u2XZeY1lIO/x+xUHmOa0lFRl2TEiQtSKz7Sg9h1WDwfi2HnNre26CdbysN262fcUqcTA2u6PLCXlof6bNE/WdCmvLa0NixRU"
    "HPOCkrX5IP4+gouy0QELmuzT5lC5dLKaszgyq+GVBVVj71He0kpsIDKV1KC+NDfC4qtYrUp3q8IJiczyug47r8SevjeVbpMz"
    "3XThUteaHrAQ7NnJ8CY1Pa953FF+9GjZQk5HGXyQdu5xfWWMx0UW6hCXtsl0bYTC6SPcNNir4qR4eYsrSoP3f5JkvU3A5jR7"
    "YlCjHYa/kiVQxVYJsUDcKM0d8X9ZbHHNmkVQ7OqW7DE8kg1mmzjbiAHrXjhnhlqD22bjtFUgc/lN8PE3MeAhzmmcAemlKgme"
    "lLzzLHEG6LwEhS116BuCGtyfoEUb+TDcsCaFKxkIXX+kIJmTv7P7HuKVj+qflP+We1GN6o+XfnxvLdf5zin/14HQMQpLwU0L"
    "BXRw0rTW+DfKa4f0e+jU9HtzZ226SyZqeMN39XyWlUbkTwySByPU+rmVDMj/lTCQ4W87eTL6Eq9Yk5SflZGlJKZ/EM9clByB"
    "9e0+1dIcYJxLHrRWonfVL13kPLVlAgUXPKKfMyjLd0/UtSpdfMSZhYG8CCfTVWnzGknBj3m3yxSpbNuXXrw8JThtVZkluE5a"
    "mspIZc79kaDYIms6i6CeO/x5zrEzqw59jZJB6n4sELJ4GZmpxYSeC3bQAKVrEcv0+jffqyZVT2U7e0z4WBmrJsvKOJ0cpwNp"
    "aXJzSZoYXCioGf1AvaUgjdscULogH0f+3ZSSyHVLwLpuxU+NdWekvo+nCEWDycXIXXWCpBY6CdBGdbmUhsOnHXS/3x3qJh+i"
    "UYfbdS06Woy0sZz51tCUYKwu/KTFFnwKhA1n1mnCwsZ2xj4O3pgHjI+ktwPKuTEiE8PMQhjPmrc+wV9Q/Yv63t106EAVYXLX"
    "EXcL5WdYlCbaKCGRdEEHrRyHr15ZeHdto/S0CGeOs7p5f/wyljbNKO6sWCdKzfaVr93lXm1G2v36D+q4/i4CtHjZlAssJ6AE"
    "NLKyWvyLYgweoLBWavryQehGzNBLinZ5H5glO3imve7U3GvI94CNr/klk27yKbWOQWO1KBbv7j7t2GlJ15scVLBT9k1q2r/m"
    "niL654K9NOHnzczqXex+Nre/T37xC7Hjx3dk/HSb4NyQlrD++Gc5mtYbuS8gHR/rKcVGRyPyYe9Hyfa+LnThzpGYJXHYYvt4"
    "bDshhJv1ucp7kkD0ZKOXm6i8cL8K9806QWkCNmx3UDwWDIJx+k/8wTcoZGSc8zNjexB/+oPOjUMbzWhuMSaE8n9TisLmuQQ2"
    "vxttroh5uV1pQXj+b8o7Jvxp/yRW1Dn/ySrx9d+QzEd1BjC9E7EEvnrDvFv+9dDdvmkfODsbxwXk5YPz8q/UijaB/a1WD+N2"
    "DdGyTcieUwGxwRhicFu6N6u5mWs2hwV79vTk6HiIX6evjcdF0LvRtseRL/lgG3fmCNfyhbWz1W0xeTf19D4K/UdK5rX49HIc"
    "zTn0NGjZPilb2eE6fBiiXRCYYQ9DHz2JgZAYZ/2wtiwd1959XuMBPDaNx7nAPp2mGzW0fCsPhfus3TfWDRBzjvA1T6u77v4y"
    "LBkjlWBo9wF7Jj0X9zWWkTDx1a9ZTLD3RF/pZs5xKXfLA3B5ukPiYP5mYIgB/QRB1oe40XjKI8IJDXzyqHvAjr3xLeNbz/Oc"
    "AHJGj+l7SWeW+heNUKjPSIMY4BPOzK6ddzBhu9sGeGiu/SEyFu+umlOvjejiz23rFxdRj7w56aWeLwBzLzG9J7TX7es4y6TX"
    "Fxm8BN36Da0kDxXw/ki4R0rMq08a5/boSqJ7ePPlPYPT2+iKaeXvpf7vsLnNrb515WR5d/lWcyQWgQAVuTv3EvY1Tn63+FKR"
    "JGyXoxWLp17kgFgJbEABxGGj2LK41q9K85sS6BekOHRsvzfCPpPQoKJUJyy3SR5q39spfl93IL2EvTRqG9NdP1KiZgRu6CaZ"
    "KOq0cl1KrSGx+kA4R2RidU6CdMsSmyJjKzCsYe2JBeJnfNdYyg4MF9vdOjKlABRJI3fvFf8S+/cA6j+97FbFeK5cvkMBlIvQ"
    "6dfNSTnhKrpWzkLc0U4t7EG+m1HB+Co+GxFDYWQTypFdkysti6a5tViYrdGmC9gg6rmOSwgFjJWN/qjHt4iZZO5N8FTuPlzB"
    "3vHrlxQB5W12YTJYpVRNgFt7Karf+IRnmBYtO9DAey0A+yyVXp3BDa3N2gPajZBSALlIlpvqDu7r+Acn8DODwm1d+zlwEf/H"
    "L7BTKtgFVkXJ+PBtqSfBhAZphJvom3+My1huVKD23KQVFFFZSga+88MG9WLUEzqxpi+lIz0+S9Ve57zw7nqrJyoI7q+ReH5n"
    "VjkuvD1L3r8YdPEZ1f/LVtR7q0pL10gxyXezXpYAiJaZK4NqTqhiixNPYTJwwnAel6WCj76PwSd3bTqYqanWVpXMgsMBYkhc"
    "jJsYyOyxLTeXSgb1zbmbMYPdSv75u99gn0vrHbtPexQ6OnoULtnQCYxUBp/V7/hbnHujmXn3GUdSfigBefpQsXZXrc6O2ZcK"
    "xtp8Y0I5bKxRHCw1O8bdnCbAqxJ3GcPDU+wq9gDIn1Su1D2Gu9OKlGImbu7CZLlWX2+h0Avf7yNwb95jv6w85qgXeF/mCqFb"
    "zPJ4w9NRm/f1Fb+Q2zPVl/cQrI0Id5VluSgZGPU2enAKUcFkTaHEHp2nnKtewra7BroMpS5Pw2unI1tdVB4xms8wMvAQ4wZI"
    "1qZGJHBQ1xbdErXJJOjFZpIsPDpGuszrsQ+pznFD43T7TClImdtl2dfgY/oZuhgSHlwVLyu9whZT/HDCPjrJChehdURrpMFd"
    "vC1BYqDcy5kTNI55agY90T8xKez2TS21Ajl0v6KcAMXWDbvejfVJ19TTDWZBxP8dSyHRZWD53dq5lSPxBSxkHSs8xzUxdOe2"
    "kOti5+UD7wocq82FBGVK3mu+JCJcN0rDE4blcq7IwEqNMIUebvuqagZM5blQze7hSXOxwB1axs48HIFGozfeMUZtWhN64TP+"
    "D4Np0sYdDsCVVtqTPD+iDnfEB2bDmsUf99fMyUWrDbv/Actl51y5ccVdzwl1mcwS6dfflWc5yZDc+lEEPf59UpqsYkkovskn"
    "F/duj0pPpUEtfLtabZah8T3ZuJvFFKhVSTPGLcXVshw3+XAXyNv7tTCZlglJ7ev1GXjrN/dYMKKhYuVLUBg2cUME6mzmuuQY"
    "ONGdmbdqxIuT0ncP3qcGavPmCv8o+YuLYPVqJTNI8xcG1Hk+kECNwu23SxkY/QvYZouw1JmynDmtzXEto1eodSvKaxQMO2yv"
    "VTSLIYRuSfnksjVXCfTLqctcSPs0DBxEH7rYMrLSw9pNtw955+oHcwYG9xmLwmE+z30nN/2gBvQsRJSMc/odi/kFT7A+5mi1"
    "hKxSCTc4HqnrY0wQojKGQuw8f0qU07RT+T63T7zCeIN+htRATFcXndiPhtmtbbW9ZGolg4/uv7/E0RH0S5NbDrdozcVMyo6T"
    "eVEDKRv/Yz2NYp4HN2ymfd1yC2WS1fOu/2+5rGRht/hdOeg4zZx36WvLqeuaXAk2WJ4rohR/JEYEXIF28r57/AOnq1P/3CW2"
    "WynKhbsJsLX0P/mVMgEjU8HF77/EjVWdSJe4UajqnUij4lV/Us81pPFXLWies0bTe3a7W7u70czF+RF5W8IMjBqiomWryg82"
    "FdvgQR2qL3JYrgzZcbnJxWU+7pwO3VFQNnp5BKLML9ejiOf4NxFlSmrdXLTHSIGp6xizxUQnH6NVCVEObsulEGgCoswgytk3"
    "QZZ39ikVmeUMPjKSAPCyIxmW14Xx/WoYZUYSlmI7aJ4GavOVR6oxPlYPdj9qs+zXFy62ueEHH9shNXE0itgGlmvCbxdCI5su"
    "zCxE3uGo4N7XJ33K9OO/XGjfHDp797EhEMV8T3J+T5RV2ObgL4uyqCXUhzdKdEjJ6V8aLEmsSOqWXcuZsTJ8afu9knUH+Kgp"
    "aUMrIC/D/g9QSwcINGjXlmQ7AABIPQAAUEsDBBQACAgIAKSTtlwAAAAAAAAAAAAAAAATAAAAW0NvbnRlbnRfVHlwZXNdLnht"
    "bMVUS24CMQw9Qe8wyraaBFhUVcXAoi3LtlLpAULi+Yj8FAcYbt/Mh6qg6YIW1M04nhf7vdiyp/Naq2QLHitrMjKmI5KAEVZW"
    "psjIx3KR3pMEAzeSK2sgI3tAMp/dTJd7B5jEYIMZKUNwD4yhKEFzpNaBiUhuveYhur5gjos1L4BNRqM7JqwJYEIamhxkNn2C"
    "nG9USB67/03qjHDnVCV4iLpYTEaS5zqCnczGZz/EVbrhcaY4imj8nyK+M22NPJGf2jyvBEgrNjqGULvKNxhvg1zEJEckVoaQ"
    "/5amrxD1oNo7WFYOb09fHlFsGF5jy3wl4S8vQeeBSywBglZ0Z/26PXecb9yHF65jUlYr9gUia82Y9i34Zx2T6+nAknuQ78HH"
    "ScAhLUcXLqlDer6LOYc4ewgPh4v24QzeM+uuMYVagKIuLhprhhg6BHt7xbaGvYLhfrbIJZlDXIcwRNUC3feKk9Raqnk1WPBm"
    "lFbWrg/8rN3os09QSwcIWOQmwnIBAAARBgAAUEsBAhQAFAAICAgApJO2XJ2CEhOZAwAANA0AABgAAAAAAAAAAAAAAAAAAAAA"
    "AHhsL2RyYXdpbmdzL2RyYXdpbmcxLnhtbFBLAQIUABQACAgIAKSTtly6i+P3tgAAACMBAAAjAAAAAAAAAAAAAAAAAN8DAAB4"
    "bC9kcmF3aW5ncy9fcmVscy9kcmF3aW5nMS54bWwucmVsc1BLAQIUABQACAgIAKSTtlwHYmmDBQEAAAcDAAAYAAAAAAAAAAAA"
    "AAAAAOYEAAB4bC9kcmF3aW5ncy9kcmF3aW5nMi54bWxQSwECFAAUAAgICACkk7ZcZRSFj0ygAADk5wUAGAAAAAAAAAAAAAAA"
    "AAAxBgAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQAFAAICAgApJO2XK2o602zAAAAKgEAACMAAAAAAAAAAAAAAAAA"
    "w6YAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzUEsBAhQAFAAICAgApJO2XMXCpah7DAAA0KQAABgAAAAA"
    "AAAAAAAAAAAAx6cAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQIUABQACAgIAKSTtlyFAfUVtAAAACoBAAAjAAAAAAAA"
    "AAAAAAAAAIi0AAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0Mi54bWwucmVsc1BLAQIUABQACAgIAKSTtly4MwHFKgMAALkO"
    "AAATAAAAAAAAAAAAAAAAAI21AAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQAFAAICAgApJO2XP/X8BVyAwAAZwgAABQAAAAA"
    "AAAAAAAAAAAA+LgAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAhQAFAAICAgApJO2XLlyP/cOBwAAZHEAAA0AAAAAAAAAAAAA"
    "AAAArLwAAHhsL3N0eWxlcy54bWxQSwECFAAUAAgICACkk7ZcNGgDnIcAAAChAAAAFQAAAAAAAAAAAAAAAAD1wwAAeGwvcGVy"
    "c29ucy9wZXJzb24ueG1sUEsBAhQAFAAICAgApJO2XHX4G/1YAQAAcwMAAA8AAAAAAAAAAAAAAAAAv8QAAHhsL3dvcmtib29r"
    "LnhtbFBLAQIUABQACAgIAKSTtlw/D4UaFAEAAMMDAAAaAAAAAAAAAAAAAAAAAFTGAAB4bC9fcmVscy93b3JrYm9vay54bWwu"
    "cmVsc1BLAQIUABQACAgIAKSTtlykb6EgsgAAACgBAAALAAAAAAAAAAAAAAAAALDHAABfcmVscy8ucmVsc1BLAQIUABQACAgI"
    "AKSTtlw0aNeWZDsAAEg9AAATAAAAAAAAAAAAAAAAAJvIAAB4bC9tZWRpYS9pbWFnZTEucG5nUEsBAhQAFAAICAgApJO2XFjk"
    "JsJyAQAAEQYAABMAAAAAAAAAAAAAAAAAQAQBAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAABAAEABMBAAA8wUBAAAA"
)



# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _nominal_text(text: str) -> str:
    """Strip count prefix and tolerance suffixes; return only the nominal value."""
    t = text.strip()
    # Remove leading count multiplier: "3X ", "2x ", "4× ", etc.
    t = re.sub(r"^\d+\s*[xX×]\s*", "", t)
    # Remove trailing tolerance suffixes
    t = re.sub(r"\s*[±]\s*[\d.]+.*$", "", t)
    t = re.sub(r"\s*\+[\d.]+\s*/?\s*-[\d.]+.*$", "", t)
    t = re.sub(r"\s*\+[\d.]+\s+-[\d.]+.*$", "", t)
    return t.strip() or text


def _parse_tolerances(
    text: str,
) -> tuple[Optional[float], Optional[float], Optional[float]]:
    """Return (nominal, plus_tol, minus_tol) from a dimension text string.

    All returned values are positive floats, or None when not parseable.
    """
    t = text.strip()
    # Strip count prefix: "2X", "4x", "3×", "2X " etc.
    t = re.sub(r"^\d+\s*[xX×]\s*", "", t)
    t = re.sub(r"^[Øøâ£⌀∅]\s*", "", t)
    t = re.sub(r"^[Rr]\s*", "", t)

    # "25.4 ±0.2"
    m = re.match(r"^([+-]?\d+\.?\d*)\s*[±]\s*([0-9.]+)", t)
    if m:
        return float(m.group(1)), float(m.group(2)), float(m.group(2))

    # "25.4 +0.3/-0.1" or "25.4 +0.3 -0.1"
    m = re.match(r"^([+-]?\d+\.?\d*)\s*\+([0-9.]+)\s*/?\s*-([0-9.]+)", t)
    if m:
        return float(m.group(1)), float(m.group(2)), float(m.group(3))

    # Plain number
    m = re.match(r"^([+-]?\d+\.?\d*)\s*[°%]?\s*$", t)
    if m:
        return float(m.group(1)), None, None

    # Fractional inch "1/4" → 0.25
    m = re.match(r"^(\d+)/(\d+)$", t)
    if m and int(m.group(2)) != 0:
        return int(m.group(1)) / int(m.group(2)), None, None

    return None, None, None


def _count_decimal_places(text: str) -> int:
    """Return the number of decimal digits in a dimension text's nominal value.

    Strips common prefixes (count multiplier, Ø, R, ±, …) before inspecting
    the number.
    Examples: "25" → 0,  "25.4" → 1,  "2X 16.00" → 2,  "4x Ø5.5" → 1.
    """
    t = text.strip()
    t = re.sub(r'^\d+\s*[xX×]\s*', '', t)   # strip "2X ", "4x " etc.
    t = re.sub(r'^[ØøRr⌀∅±+\-\s]*', '', t)
    m = re.match(r'^[\d]+\.([\d]+)', t)
    return len(m.group(1)) if m else 0


def _parse_title_block_tolerances(text: str) -> dict:
    """Parse general-tolerance declarations from a title-block text blob.

    Handles both signed and unsigned formats, e.g.:

        X                    NO DECIMALS 1         ← bare number (no ±)
        .X                   DECIMAL  .3
        .XX                  DECIMAL  .13
        ANGULAR OR BEND      1/2                   ← fraction = 0.5°

        X NO DECIMALS ±1                           ← signed format
        .X DECIMAL +/- 0.3
        ANGULAR OR BEND +/- 1.2 DEGREE

    Strategy: process line-by-line (the natural format for title blocks).
    For each line that matches a format marker (X, .X, .XX, ANGULAR/BEND),
    extract the tolerance value with _tol_value(), which first looks for an
    explicit ± / +/- sign and falls back to the last numeric token on the
    line (bare number or N/M fraction).

    Returns a dict keyed by decimal-place count (int) or "angular" (str),
    e.g. {0: 1.0, 1: 0.3, 2: 0.13, "angular": 0.5}.
    Returns an empty dict when no tolerance declarations are found.
    """
    result: dict = {}
    # Normalize horizontal whitespace only; preserve newlines so line-by-line
    # parsing works on multi-line title blocks.
    upper = re.sub(r'[ \t]+', ' ', text.upper())

    def _tol_value(segment: str) -> Optional[float]:
        """Extract a tolerance magnitude from a text segment.

        Tries explicit-sign formats first (±N, +/-N, +N/-N), then falls back
        to the last numeric token on the line — either a plain decimal number
        or an N/M fraction (e.g. "1/2" → 0.5).
        """
        # --- Explicit sign formats ---
        for pat in (
            r'[±]\s*(\d+\.?\d*)',                         # ±0.13
            r'[+]\s*/\s*[-]\s*(\d+\.?\d*)',               # +/- 0.13
            r'[+](\d+\.?\d*)\s*/?\s*-\d+\.?\d*',         # +0.13/-0.13
        ):
            m = re.search(pat, segment)
            if m:
                try:
                    return float(m.group(1))
                except ValueError:
                    pass

        # --- Bare number / fraction fallback ---
        # Find every numeric token; take the last one as the tolerance value.
        # Fractions like "1/2" are resolved to 0.5.
        tokens = list(re.finditer(r'(\d+)\s*/\s*(\d+)|(\d*\.?\d+)', segment))
        if tokens:
            last = tokens[-1]
            if last.group(1) is not None:       # N/M fraction
                den = int(last.group(2))
                return int(last.group(1)) / den if den else None
            try:
                return float(last.group(3))
            except (ValueError, TypeError):
                pass
        return None

    for raw_line in upper.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # Angular / bend (check before X so "BEND" lines are not caught by X)
        if re.search(r'\bANGULAR\b|\bBEND\b', line):
            t = _tol_value(line)
            if t is not None:
                result["angular"] = t
            continue

        # Whole-number: line starts with standalone "X" (not part of ".X…")
        if re.match(r'^X\b', line):
            t = _tol_value(line)
            if t is not None:
                result[0] = t
            continue

        # .X, .XX, .XXX … — line starts with ".X"
        m = re.match(r'^(\.X+)\b', line)
        if m:
            decimals = len(m.group(1)) - 1   # .X→1, .XX→2, .XXX→3
            t = _tol_value(line)
            if t is not None:
                result[decimals] = t

    # ------------------------------------------------------------------ #
    # Single-line fallback: if line-by-line found nothing, try splitting  #
    # on commas/semicolons (some title blocks pack everything on one line) #
    # ------------------------------------------------------------------ #
    if not result:
        for seg in re.split(r'[,;]', upper):
            seg = seg.strip()
            if not seg:
                continue
            if re.search(r'\bANGULAR\b|\bBEND\b', seg):
                t = _tol_value(seg)
                if t is not None:
                    result["angular"] = t
            elif re.match(r'^X\b', seg):
                t = _tol_value(seg)
                if t is not None:
                    result[0] = t
            else:
                dm = re.match(r'^(\.X+)\b', seg)
                if dm:
                    decimals = len(dm.group(1)) - 1
                    t = _tol_value(seg)
                    if t is not None:
                        result[decimals] = t

    return result


def _smart_callout_pos(
    block: dict,
    placed: list[tuple[float, float]],
    r_pt: float,
    prefer_left: bool = False,
) -> tuple[float, float]:
    """Pick the callout centre with fewest overlaps from 8 candidate positions.

    When prefer_left is True (tolerance dimensions whose ± or +/− text extends
    to the right of the nominal), left is tried first so the circle doesn't
    land on top of the tolerance annotation.  Otherwise right is tried first.
    The first candidate with zero overlaps wins; otherwise the fewest-overlap
    position is used.
    """
    bx = block["x"]
    by = block["y"]
    bw = block["width"]
    bh = block["height"]
    bcx = bx + bw / 2.0
    bcy = by + bh / 2.0
    gap = r_pt + 4.0

    if prefer_left:
        candidates = [
            (bx - gap,       bcy),           # left (preferred — tolerance text is to the right)
            (bx + bw + gap,  bcy),           # right
            (bcx,            by - gap),      # above
            (bcx,            by + bh + gap), # below
            (bx - gap,       by - gap),      # upper-left
            (bx + bw + gap,  by - gap),      # upper-right
            (bx - gap,       by + bh + gap), # lower-left
            (bx + bw + gap,  by + bh + gap), # lower-right
        ]
    else:
        candidates = [
            (bx + bw + gap,  bcy),           # right (preferred)
            (bx - gap,       bcy),           # left
            (bcx,            by - gap),      # above
            (bcx,            by + bh + gap), # below
            (bx + bw + gap,  by - gap),      # upper-right
            (bx - gap,       by - gap),      # upper-left
            (bx + bw + gap,  by + bh + gap), # lower-right
            (bx - gap,       by + bh + gap), # lower-left
        ]

    best = candidates[0]
    min_overlaps: Optional[int] = None
    diameter = r_pt * 2.0

    for cx, cy in candidates:
        overlaps = sum(
            1 for px, py in placed if math.hypot(cx - px, cy - py) < diameter
        )
        if min_overlaps is None or overlaps < min_overlaps:
            min_overlaps = overlaps
            best = (cx, cy)
            if overlaps == 0:
                break

    return best


def _find_column(headers: list[dict], keywords: list[str]) -> Optional[int]:
    """Return the column index of the first header whose text contains any keyword."""
    for h in headers:
        val = str(h.get("value", "")).lower().strip()
        for kw in keywords:
            if kw in val:
                return int(h["column_index"])
    return None


# ---------------------------------------------------------------------------
# Config — reads / writes config.json next to the executable
# ---------------------------------------------------------------------------

class Config:
    def __init__(self) -> None:
        self._path = os.path.join(_APP_DIR, CONFIG_FILENAME)
        self._data: dict[str, Any] = {
            "callout_color": DEFAULT_CALLOUT_COLOR,
            "callout_radius": DEFAULT_CALLOUT_RADIUS,
        }
        self._load()

    def _load(self) -> None:
        if os.path.isfile(self._path):
            try:
                with open(self._path, "r", encoding="utf-8") as fh:
                    saved = json.load(fh)
                self._data.update(saved)
            except Exception:
                pass

    def save(self) -> None:
        try:
            with open(self._path, "w", encoding="utf-8") as fh:
                json.dump(self._data, fh, indent=2)
        except Exception:
            pass

    @property
    def callout_color(self) -> str:
        return self._data.get("callout_color", DEFAULT_CALLOUT_COLOR)

    @callout_color.setter
    def callout_color(self, v: str) -> None:
        self._data["callout_color"] = v

    @property
    def callout_radius(self) -> int:
        return int(self._data.get("callout_radius", DEFAULT_CALLOUT_RADIUS))

    @callout_radius.setter
    def callout_radius(self, v: int) -> None:
        self._data["callout_radius"] = int(v)


# ---------------------------------------------------------------------------
# Background worker thread — runs detect_dimensions() without blocking the UI
# ---------------------------------------------------------------------------

class DetectionWorker(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, pdf_path: str, page_num: int) -> None:
        super().__init__()
        self.pdf_path = pdf_path
        self.page_num = page_num

    def run(self) -> None:
        try:
            result = detect_dimensions(self.pdf_path, self.page_num)
            self.finished.emit(result)
        except RuntimeError as exc:
            self.error.emit(str(exc))
        except Exception as exc:
            self.error.emit(f"{type(exc).__name__}: {exc}")


# ---------------------------------------------------------------------------
# Spinner widget — animated loading indicator drawn with QPainter
# ---------------------------------------------------------------------------

class SpinnerWidget(QWidget):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._angle = 0
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)

    def start(self) -> None:
        self._timer.start(40)

    def stop(self) -> None:
        self._timer.stop()

    def _tick(self) -> None:
        self._angle = (self._angle + 12) % 360
        self.update()

    def paintEvent(self, _event) -> None:  # noqa: N802
        w, h = self.width(), self.height()
        r = min(w, h) // 2 - 4
        cx, cy = w // 2, h // 2
        n = 10

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        for i in range(n):
            angle_deg = (self._angle + i * (360 // n)) % 360
            alpha = int(255 * (i + 1) / n)
            color = QColor(231, 76, 60, alpha)
            pen = QPen(color, 3, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap)
            painter.setPen(pen)

            rad = math.radians(angle_deg)
            x1 = cx + (r - 9) * math.cos(rad)
            y1 = cy - (r - 9) * math.sin(rad)
            x2 = cx + r * math.cos(rad)
            y2 = cy - r * math.sin(rad)
            painter.drawLine(int(x1), int(y1), int(x2), int(y2))

        painter.end()


# ---------------------------------------------------------------------------
# CalloutItem — draggable numbered circle overlay on the PDF scene
# ---------------------------------------------------------------------------

class CalloutItem(QGraphicsItem):
    def __init__(
        self,
        number: int,
        radius: int,
        color: str,
        on_moved: Optional[Callable] = None,
        on_drag_complete: Optional[Callable] = None,
        on_context_menu: Optional[Callable] = None,
        on_click: Optional[Callable] = None,
    ) -> None:
        super().__init__()
        self.number = number
        self.radius = radius
        self.color = QColor(color)
        self._on_moved = on_moved
        self._on_drag_complete = on_drag_complete
        self._on_context_menu = on_context_menu
        self._on_click = on_click
        self._drag_start_pos: Optional[QPointF] = None
        self._highlighted: bool = False

        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, True)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges, True)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)
        self.setZValue(10)
        self.setCursor(Qt.CursorShape.SizeAllCursor)

    def set_number(self, n: int) -> None:
        self.number = n
        self.update()

    def set_color(self, color_hex: str) -> None:
        self.color = QColor(color_hex)
        self.update()

    def set_radius(self, r: int) -> None:
        self.prepareGeometryChange()
        self.radius = r
        self.update()

    def set_highlighted(self, active: bool) -> None:
        self._highlighted = active
        self.update()

    def boundingRect(self) -> QRectF:
        # Extra space accommodates the selection ring (r+5 + 1.5 px pen)
        r = self.radius + 9.0
        return QRectF(-r, -r, 2 * r, 2 * r)

    def shape(self) -> QPainterPath:
        # Hit-test area stays tight so only the circle itself is clickable
        path = QPainterPath()
        r = float(self.radius)
        path.addEllipse(QRectF(-r, -r, 2 * r, 2 * r))
        return path

    def paint(self, painter: QPainter, _option, _widget) -> None:
        r = float(self.radius)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Selection ring — blue, drawn outermost when this callout is active
        if self._highlighted:
            painter.setPen(QPen(QColor(52, 152, 219), 3.0))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            rr = r + 6.5
            painter.drawEllipse(QRectF(-rr, -rr, 2 * rr, 2 * rr))

        # White fill + colored border
        painter.setPen(QPen(self.color, 2.0))
        painter.setBrush(QBrush(QColor(255, 255, 255, 230)))
        painter.drawEllipse(QRectF(-r, -r, 2 * r, 2 * r))

        # Number label — font shrinks automatically so it always fits the circle
        num_str = str(self.number)
        base_pt = max(6.0, r * 0.72)
        font = QFont("Arial", 1, QFont.Weight.Bold)
        font.setPointSizeF(base_pt)
        # Available horizontal space = diameter minus 2 px margin each side
        available_w = max(1.0, 2.0 * r - 4.0)
        text_w = QFontMetricsF(font).horizontalAdvance(num_str)
        if text_w > available_w:
            font.setPointSizeF(max(5.0, base_pt * available_w / text_w))
        painter.setFont(font)
        painter.setPen(QPen(self.color))
        painter.drawText(
            QRectF(-r, -r, 2 * r, 2 * r),
            Qt.AlignmentFlag.AlignCenter,
            num_str,
        )

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_start_pos = self.pos()
            if self._on_click is not None:
                self._on_click(self)
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event) -> None:
        super().mouseReleaseEvent(event)
        if (
            event.button() == Qt.MouseButton.LeftButton
            and self._drag_start_pos is not None
            and self._on_drag_complete is not None
        ):
            end_pos = self.pos()
            # Only record if actually moved (not just a click-to-select)
            if (abs(end_pos.x() - self._drag_start_pos.x()) > 0.5 or
                    abs(end_pos.y() - self._drag_start_pos.y()) > 0.5):
                self._on_drag_complete(self, self._drag_start_pos, end_pos)
        self._drag_start_pos = None

    def itemChange(self, change, value):
        if (
            change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged
            and self._on_moved is not None
        ):
            self._on_moved(self, value)
        return super().itemChange(change, value)

    def contextMenuEvent(self, event) -> None:
        """Right-click on a callout circle — fire the context-menu callback."""
        if self._on_context_menu is not None:
            self._on_context_menu(self, event.screenPos())
        event.accept()


# ---------------------------------------------------------------------------
# PDFViewer — zoomable QGraphicsView with add-callout mode
# ---------------------------------------------------------------------------

class PDFViewer(QGraphicsView):
    callout_place_requested  = pyqtSignal(float, float)  # scene x, y
    callout_delete_requested = pyqtSignal(object)        # selected CalloutItem
    zoom_changed             = pyqtSignal(int)           # zoom level in percent
    prev_page_requested      = pyqtSignal()              # ← arrow key pressed
    next_page_requested      = pyqtSignal()              # → arrow key pressed

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._scene = QGraphicsScene()
        self.setScene(self._scene)
        self._add_mode = False

        self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.setBackgroundBrush(QBrush(QColor("#1A1A2E")))
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setMinimumSize(400, 300)
        # Allow the view to receive key events after clicking a callout
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

    # --- Public API --------------------------------------------------------

    def current_zoom_percent(self) -> int:
        return round(self.transform().m11() * 100)

    def set_add_mode(self, enabled: bool) -> None:
        self._add_mode = enabled
        if enabled:
            self.setDragMode(QGraphicsView.DragMode.NoDrag)
            self.viewport().setCursor(Qt.CursorShape.CrossCursor)
        else:
            self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
            self.viewport().setCursor(Qt.CursorShape.ArrowCursor)

    def display_pixmap(self, pixmap: QPixmap) -> None:
        self._scene.clear()
        self._scene.addPixmap(pixmap)
        self._scene.setSceneRect(QRectF(pixmap.rect()))

    def fit_page(self) -> None:
        if self._scene.sceneRect().isValid():
            self.fitInView(self._scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatio)

    def add_callout_item(self, item: CalloutItem) -> None:
        self._scene.addItem(item)

    def remove_callout_item(self, item: CalloutItem) -> None:
        if item.scene() == self._scene:
            self._scene.removeItem(item)

    def clear_callout_items(self) -> None:
        for item in list(self._scene.items()):
            if isinstance(item, CalloutItem):
                self._scene.removeItem(item)

    def zoom_in(self) -> None:
        self.scale(1.25, 1.25)
        self.zoom_changed.emit(self.current_zoom_percent())

    def zoom_out(self) -> None:
        self.scale(0.8, 0.8)
        self.zoom_changed.emit(self.current_zoom_percent())

    def zoom_reset(self) -> None:
        self.resetTransform()
        self.fit_page()
        self.zoom_changed.emit(self.current_zoom_percent())

    # --- Qt overrides -------------------------------------------------------

    def wheelEvent(self, event) -> None:
        factor = 1.15 if event.angleDelta().y() > 0 else 1.0 / 1.15
        self.scale(factor, factor)
        self.zoom_changed.emit(self.current_zoom_percent())

    def mousePressEvent(self, event) -> None:
        if self._add_mode and event.button() == Qt.MouseButton.LeftButton:
            pos = self.mapToScene(event.pos())
            self.callout_place_requested.emit(pos.x(), pos.y())
            return
        super().mousePressEvent(event)
        # Keep keyboard focus on the view after any click so Delete works
        self.setFocus()

    def keyPressEvent(self, event) -> None:
        key = event.key()
        if key in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            for item in self._scene.selectedItems():
                if isinstance(item, CalloutItem):
                    self.callout_delete_requested.emit(item)
        elif key == Qt.Key.Key_Left:
            self.prev_page_requested.emit()
        elif key == Qt.Key.Key_Right:
            self.next_page_requested.emit()
        else:
            super().keyPressEvent(event)


# ---------------------------------------------------------------------------
# DimensionRow — one item in the right-hand checklist
# ---------------------------------------------------------------------------

class DimensionRow(QWidget):
    check_changed   = pyqtSignal(int, bool)  # (dim_index, is_checked)
    edit_requested  = pyqtSignal(int)         # (dim_index)
    focus_requested = pyqtSignal(int)         # (dim_index) — row focused, scroll drawing to callout

    def __init__(
        self,
        dim_index: int,
        text: str,
        category: str,           # kept for API compat — no longer displayed
        confidence: float,       # kept for API compat — no longer displayed
        callout_color: str = DEFAULT_CALLOUT_COLOR,
        applied_plus_tol: Optional[float] = None,
        applied_minus_tol: Optional[float] = None,
        tol_inferred: bool = False,
        initial_enabled: bool = True,
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)
        self.dim_index = dim_index
        self._base_bg    = "#F8F8F8" if dim_index % 2 == 0 else "#FFFFFF"
        self._tol_inferred = tol_inferred
        self._focused = False
        self._highlighted = False

        self.setMinimumHeight(36)
        self.setFocusPolicy(Qt.FocusPolicy.ClickFocus)
        self.setMaximumHeight(44)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 3, 6, 3)
        layout.setSpacing(6)

        # ── Toggle button (replaces bare QCheckBox) ──────────────────────────
        self.toggle_btn = QPushButton()
        self.toggle_btn.setCheckable(True)
        self.toggle_btn.setChecked(initial_enabled)
        self.toggle_btn.setFixedSize(58, 26)
        self.toggle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_btn.toggled.connect(self._on_toggled)
        layout.addWidget(self.toggle_btn)

        # ── Callout number badge ──────────────────────────────────────────────
        self.num_label = QLabel("1")
        self.num_label.setFixedSize(22, 22)
        self.num_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._badge_color = callout_color
        self._apply_badge_style(callout_color, active=initial_enabled)
        layout.addWidget(self.num_label)

        # ── Nominal text label ────────────────────────────────────────────────
        display = _nominal_text(text)
        if len(display) > 28:
            display = display[:26] + "…"
        self.text_lbl = QLabel(display)
        self.text_lbl.setToolTip(text)
        self.text_lbl.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred
        )
        layout.addWidget(self.text_lbl)

        # ── Tolerance label ───────────────────────────────────────────────────
        if applied_plus_tol is not None:
            if applied_minus_tol is not None and applied_minus_tol != applied_plus_tol:
                tol_str = f"+{applied_plus_tol:g}/−{applied_minus_tol:g}"
            else:
                tol_str = f"±{applied_plus_tol:g}"
            if tol_inferred:
                tol_str += "*"          # asterisk = inferred from title block
            self.tol_lbl = QLabel(tol_str)
            self.tol_lbl.setFixedWidth(68)
            self.tol_lbl.setAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            )
            tip = (
                "Inferred from title-block general tolerances"
                if tol_inferred else
                "Explicit tolerance from dimension text"
            )
            self.tol_lbl.setToolTip(tip)
        else:
            self.tol_lbl = QLabel("")
            self.tol_lbl.setFixedWidth(68)
        layout.addWidget(self.tol_lbl)

        # ── Edit button — wider, blue ─────────────────────────────────────────
        self.edit_btn = QPushButton("✎  Edit")
        self.edit_btn.setFixedHeight(24)
        self.edit_btn.setMinimumWidth(64)
        self.edit_btn.setToolTip("Edit this dimension")
        self.edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.edit_btn.setStyleSheet(
            "QPushButton{background:#3498DB;color:white;border:none;"
            "border-radius:4px;font-size:11px;font-weight:bold;padding:2px 7px;}"
            "QPushButton:hover{background:#2980B9;}"
            "QPushButton:pressed{background:#1F618D;}"
        )
        self.edit_btn.clicked.connect(lambda: self.edit_requested.emit(self.dim_index))
        layout.addWidget(self.edit_btn)

        # Clicking any part of the row focuses it (Delete/Backspace + center-on-callout)
        for _w in (self.num_label, self.text_lbl, self.tol_lbl,
                   self.toggle_btn, self.edit_btn):
            _w.installEventFilter(self)

        # Apply initial visual state (background, text colour, toggle label)
        self._refresh_row_style(initial_enabled)

    # --- Public API --------------------------------------------------------

    def set_callout_number(self, n: int) -> None:
        self.num_label.setText(str(n))
        self._apply_badge_style(self._badge_color, active=True)
        # Sync toggle without re-emitting signal
        self.toggle_btn.blockSignals(True)
        self.toggle_btn.setChecked(True)
        self.toggle_btn.blockSignals(False)
        self._refresh_row_style(True)

    def set_color(self, color_hex: str) -> None:
        self._badge_color = color_hex
        self._apply_badge_style(color_hex, active=self.toggle_btn.isChecked())

    def mark_excluded(self) -> None:
        self.num_label.setText("—")
        self._apply_badge_style("#AAAAAA", active=False)
        # Sync toggle without re-emitting signal
        self.toggle_btn.blockSignals(True)
        self.toggle_btn.setChecked(False)
        self.toggle_btn.blockSignals(False)
        self._refresh_row_style(False)

    # --- Private -----------------------------------------------------------

    def _apply_badge_style(self, color: str, active: bool) -> None:
        if active:
            self.num_label.setStyleSheet(
                f"background:{color};color:white;border-radius:11px;"
                f"font-weight:bold;font-size:11px;"
            )
        else:
            self.num_label.setStyleSheet(
                "background:#AAAAAA;color:white;border-radius:11px;"
                "font-weight:bold;font-size:11px;"
            )

    def _refresh_toggle_label(self, enabled: bool) -> None:
        """Update toggle button text and colour to reflect enabled/disabled state."""
        if enabled:
            self.toggle_btn.setText("✓  On")
            self.toggle_btn.setStyleSheet(
                "QPushButton{background:#27AE60;color:white;border:none;"
                "border-radius:5px;font-weight:bold;font-size:11px;}"
                "QPushButton:hover{background:#219A52;}"
                "QPushButton:checked{background:#27AE60;}"
            )
        else:
            self.toggle_btn.setText("✗  Off")
            self.toggle_btn.setStyleSheet(
                "QPushButton{background:#95A5A6;color:white;border:none;"
                "border-radius:5px;font-weight:bold;font-size:11px;}"
                "QPushButton:hover{background:#7F8C8D;}"
                "QPushButton:checked{background:#95A5A6;}"
            )

    def set_highlighted(self, active: bool) -> None:
        self._highlighted = active
        self._refresh_row_style(self.toggle_btn.isChecked())

    def paintEvent(self, event) -> None:
        """Paint row background — done here so the color is always reliable."""
        enabled = self.toggle_btn.isChecked()
        if enabled:
            if self._highlighted:
                bg = QColor("#AED6F1")
            elif self._focused:
                bg = QColor("#D6EAF8")
            else:
                bg = QColor(self._base_bg)
        else:
            bg = QColor("#C5D8E0") if self._focused else QColor("#E8E8E8")
        painter = QPainter(self)
        painter.fillRect(self.rect(), bg)
        painter.end()

    def _refresh_row_style(self, enabled: bool) -> None:
        """Update child-widget styles and trigger a background repaint."""
        if enabled:
            self.text_lbl.setStyleSheet(
                "background:transparent;font-size:12px;color:#2C3E50;"
            )
            if self._tol_inferred:
                self.tol_lbl.setStyleSheet(
                    "background:transparent;font-size:10px;"
                    "color:#888;font-style:italic;"
                )
            else:
                self.tol_lbl.setStyleSheet(
                    "background:transparent;font-size:10px;"
                    "color:#2C3E50;font-weight:bold;"
                )
        else:
            self.text_lbl.setStyleSheet(
                "background:transparent;font-size:12px;color:#AAAAAA;"
            )
            self.tol_lbl.setStyleSheet(
                "background:transparent;font-size:10px;color:#BBBBBB;"
            )
        self._refresh_toggle_label(enabled)
        self.update()  # triggers paintEvent to repaint background

    def _on_toggled(self, checked: bool) -> None:
        self._refresh_row_style(checked)
        self._apply_badge_style(
            self._badge_color if checked else "#AAAAAA", active=checked
        )
        self.check_changed.emit(self.dim_index, checked)

    # --- Keyboard delete support -------------------------------------------

    def mousePressEvent(self, event) -> None:
        """Clicking the row background focuses it for keyboard handling."""
        self.setFocus()
        super().mousePressEvent(event)

    def eventFilter(self, obj, event) -> bool:
        """Forward mouse clicks on child labels so the row gets focus."""
        if event.type() == QEvent.Type.MouseButtonPress:
            self.setFocus()
        return super().eventFilter(obj, event)

    def keyPressEvent(self, event) -> None:
        """Delete or Backspace unchecks (turns off) the focused dimension."""
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            if self.toggle_btn.isChecked():
                self.toggle_btn.setChecked(False)
            return
        super().keyPressEvent(event)

    def focusInEvent(self, event) -> None:
        self._focused = True
        self._refresh_row_style(self.toggle_btn.isChecked())
        self.focus_requested.emit(self.dim_index)
        super().focusInEvent(event)

    def focusOutEvent(self, event) -> None:
        self._focused = False
        self._refresh_row_style(self.toggle_btn.isChecked())
        super().focusOutEvent(event)


# ---------------------------------------------------------------------------
# ChecklistPanel — scrollable container for DimensionRow widgets
# ---------------------------------------------------------------------------

class ChecklistPanel(QWidget):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        header_bar = QWidget()
        header_bar.setFixedHeight(36)
        header_bar.setStyleSheet("background:#2C3E50;")
        hbl = QHBoxLayout(header_bar)
        hbl.setContentsMargins(10, 0, 10, 0)
        self.header_lbl = QLabel("Detected Dimensions")
        self.header_lbl.setStyleSheet("color:white;font-weight:bold;font-size:13px;")
        hbl.addWidget(self.header_lbl)
        hbl.addStretch()
        outer.addWidget(header_bar)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll.setStyleSheet("QScrollArea{border:none;}")

        self._container = QWidget()
        self._rows_layout = QVBoxLayout(self._container)
        self._rows_layout.setContentsMargins(0, 0, 0, 0)
        self._rows_layout.setSpacing(1)
        self._rows_layout.addStretch()

        self.scroll.setWidget(self._container)
        outer.addWidget(self.scroll)

    def set_count(self, n: int, n_active: Optional[int] = None) -> None:
        if n_active is not None and n_active != n:
            self.header_lbl.setText(f"Detected Dimensions  ({n_active} / {n})")
        else:
            self.header_lbl.setText(f"Detected Dimensions  ({n})")

    def clear_rows(self) -> None:
        while self._rows_layout.count() > 1:  # keep the trailing stretch
            item = self._rows_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.header_lbl.setText("Detected Dimensions")

    def add_row(self, row: DimensionRow) -> None:
        self._rows_layout.insertWidget(self._rows_layout.count() - 1, row)

    def reorder_rows(self, ordered_rows: list) -> None:
        """Rearrange existing row widgets to match ordered_rows (no rebuild).

        Widgets not present in ordered_rows are silently ignored so excluded
        rows (hidden by mark_excluded) end up at the bottom naturally.
        """
        # Detach every row from the layout without deleting it
        while self._rows_layout.count() > 1:          # keep trailing stretch
            self._rows_layout.takeAt(0)
        # Re-insert in the requested order
        for i, row in enumerate(ordered_rows):
            self._rows_layout.insertWidget(i, row)


# ---------------------------------------------------------------------------
# SettingsDialog
# ---------------------------------------------------------------------------

class SettingsDialog(QDialog):
    def __init__(self, config: Config, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Settings")
        self.setMinimumWidth(420)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        grp = QGroupBox("Callout Appearance")
        grp_layout = QFormLayout(grp)

        self._chosen_color = config.callout_color
        self.color_btn = QPushButton()
        self.color_btn.setFixedSize(80, 26)
        self._refresh_color_btn()
        self.color_btn.clicked.connect(self._pick_color)
        grp_layout.addRow("Callout Color:", self.color_btn)

        radius_row = QHBoxLayout()
        self.radius_slider = QSlider(Qt.Orientation.Horizontal)
        self.radius_slider.setRange(8, 26)
        self.radius_slider.setValue(config.callout_radius)
        self.radius_val_lbl = QLabel(str(config.callout_radius))
        self.radius_val_lbl.setFixedWidth(28)
        self.radius_slider.valueChanged.connect(
            lambda v: self.radius_val_lbl.setText(str(v))
        )
        radius_row.addWidget(self.radius_slider)
        radius_row.addWidget(self.radius_val_lbl)
        grp_layout.addRow("Callout Size:", radius_row)
        layout.addWidget(grp)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _refresh_color_btn(self) -> None:
        self.color_btn.setStyleSheet(
            f"background-color:{self._chosen_color};"
            f"border:1px solid #666;border-radius:3px;"
        )

    def _pick_color(self) -> None:
        color = QColorDialog.getColor(QColor(self._chosen_color), self, "Callout Color")
        if color.isValid():
            self._chosen_color = color.name()
            self._refresh_color_btn()

    def _accept(self) -> None:
        self.config.callout_color = self._chosen_color
        self.config.callout_radius = self.radius_slider.value()
        self.config.save()
        self.accept()


# ---------------------------------------------------------------------------
# LabelDialog — shown when user places a manual callout
# ---------------------------------------------------------------------------

class LabelDialog(QDialog):
    """Three-field callout entry: nominal, + tolerance, − tolerance.

    Tolerances are optional.  If only one is given it is used for both sides
    (symmetric ±).  If both are given and equal they are also written as ±.
    A live preview shows the formatted string that will be stored.
    """

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        *,
        title: str = "Add Callout",
        initial_nominal: str = "",
        initial_plus_tol: str = "",
        initial_minus_tol: str = "",
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(360)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── Input fields ──────────────────────────────────────────────────
        form = QFormLayout()
        form.setSpacing(8)

        def _field(placeholder: str, initial: str = "") -> QLineEdit:
            e = QLineEdit()
            e.setPlaceholderText(placeholder)
            e.setStyleSheet("color: black; background: white;")
            if initial:
                e.setText(initial)
            return e

        self.nominal_edit = _field("e.g.  25.4",                         initial_nominal)
        self.plus_edit    = _field("e.g.  0.13  (leave blank if none)",   initial_plus_tol)
        self.minus_edit   = _field("e.g.  0.10  (leave blank to match +)", initial_minus_tol)

        form.addRow("Nominal:", self.nominal_edit)
        form.addRow("+ Tolerance:", self.plus_edit)
        form.addRow("− Tolerance:", self.minus_edit)
        layout.addLayout(form)

        # ── Live preview ──────────────────────────────────────────────────
        preview_row = QHBoxLayout()
        preview_row.addWidget(QLabel("Preview:"))
        self.preview_lbl = QLabel("")
        self.preview_lbl.setStyleSheet(
            "font-weight: bold; color: #2C3E50; font-size: 13px;"
        )
        preview_row.addWidget(self.preview_lbl)
        preview_row.addStretch()
        layout.addLayout(preview_row)

        # ── Buttons ───────────────────────────────────────────────────────
        self.btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
        )
        self.btns.accepted.connect(self._accept)
        self.btns.rejected.connect(self.reject)
        layout.addWidget(self.btns)

        # Wire up live preview and Tab-to-accept on the last field
        for field in (self.nominal_edit, self.plus_edit, self.minus_edit):
            field.textChanged.connect(self._update_preview)
        self.minus_edit.returnPressed.connect(self.btns.accepted)
        self.nominal_edit.returnPressed.connect(self.plus_edit.setFocus)
        self.plus_edit.returnPressed.connect(self.minus_edit.setFocus)

        self._update_preview()

    # ── Helpers ───────────────────────────────────────────────────────────

    def _build_label(self) -> str:
        """Combine the three fields into a single dimension string."""
        nom   = self.nominal_edit.text().strip()
        plus  = self.plus_edit.text().strip()
        minus = self.minus_edit.text().strip()

        if not nom:
            return ""

        # Symmetrical ± when both sides are the same (or only one is given)
        if plus and minus:
            if plus == minus:
                return f"{nom} ±{plus}"
            return f"{nom} +{plus}/-{minus}"
        if plus:
            return f"{nom} ±{plus}"
        if minus:
            return f"{nom} ±{minus}"
        return nom

    def _update_preview(self) -> None:
        label = self._build_label()
        self.preview_lbl.setText(label if label else "—")
        ok_btn = self.btns.button(QDialogButtonBox.StandardButton.Ok)
        if ok_btn:
            ok_btn.setEnabled(bool(label))

    # ── Public API ────────────────────────────────────────────────────────

    @property
    def label_text(self) -> str:
        return self._build_label()

    def _accept(self) -> None:
        if self.label_text:
            self.accept()
        else:
            self.nominal_edit.setPlaceholderText("Nominal cannot be empty!")


# ---------------------------------------------------------------------------
# MainWindow
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("FAI Dimension Numberer")
        self.setMinimumSize(1200, 740)

        self.config = Config()
        self.template_structure: Optional[dict] = None
        self.dimensions: list[dict] = []       # global list across all pages
        self._pages_done: set[int] = set()     # pages already detected
        self.pdf_doc = None                    # fitz.Document
        self.current_page: int = 0
        self.current_pdf_path: str = ""
        self._add_callout_mode: bool = False
        self._worker: Optional[DetectionWorker] = None
        # General tolerances parsed from the title block of page 0.
        # Keys: int (decimal places) → float tolerance, "angular" → float.
        # e.g. {0: 1.0, 1: 0.3, 2: 0.13, "angular": 1.2}
        self.general_tolerances: dict = {}
        # Positions of callout circles that were burned into the PDF when it
        # was last exported.  Used to redact them before re-exporting so that
        # old numbers don't appear behind the new ones.
        self._session_burned_positions: list[dict] = []
        # Undo / redo stacks — each entry is a dict describing a reversible action.
        self._undo_stack: list[dict] = []
        self._redo_stack: list[dict] = []
        self._highlighted_row: Optional[DimensionRow] = None
        self._highlighted_callout: Optional[CalloutItem] = None

        self._build_ui()
        self._load_template()

    # =======================================================================
    # UI construction
    # =======================================================================

    def _build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._make_toolbar())

        self.splitter = QSplitter(Qt.Orientation.Horizontal)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)

        # Create PDFViewer BEFORE _make_viewer_controls so its methods can be bound
        self.pdf_viewer = PDFViewer()
        self.pdf_viewer.callout_place_requested.connect(self._on_callout_place_requested)
        self.pdf_viewer.callout_delete_requested.connect(self._on_callout_delete_requested)
        self.pdf_viewer.prev_page_requested.connect(self._go_prev_page)
        self.pdf_viewer.next_page_requested.connect(self._go_next_page)

        left_layout.addWidget(self._make_viewer_controls())
        left_layout.addWidget(self.pdf_viewer)
        self.pdf_viewer.zoom_changed.connect(lambda pct: self.zoom_lbl.setText(f"{pct}%"))

        self.right_stack = QStackedWidget()
        self.right_stack.setMinimumWidth(320)
        self.right_stack.setMaximumWidth(500)

        # Stack page 0: checklist
        self.checklist = ChecklistPanel()
        self.right_stack.addWidget(self.checklist)

        # Stack page 1: loading state
        loading_page = QWidget()
        loading_page.setStyleSheet("background:#F0F0F0;")
        ll = QVBoxLayout(loading_page)
        ll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ll.setSpacing(12)
        self.spinner = SpinnerWidget()
        self.spinner.setFixedSize(64, 64)
        ll.addWidget(self.spinner, alignment=Qt.AlignmentFlag.AlignHCenter)
        loading_lbl = QLabel("Analyzing drawing…")
        loading_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        loading_lbl.setStyleSheet("font-size:13px;color:#444;")
        ll.addWidget(loading_lbl)
        loading_sub = QLabel("Detecting dimensions via pattern matching.")
        loading_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        loading_sub.setStyleSheet("font-size:11px;color:#888;")
        ll.addWidget(loading_sub)
        self.right_stack.addWidget(loading_page)

        self.splitter.addWidget(left_widget)
        self.splitter.addWidget(self.right_stack)
        self.splitter.setSizes([820, 380])
        root.addWidget(self.splitter)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready — open a drawing to begin.")

    def _make_toolbar(self) -> QWidget:
        bar = QWidget()
        bar.setFixedHeight(52)
        bar.setStyleSheet(
            "background:#1E2132;"
            "border-bottom:1px solid #3A3A5C;"
        )
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(10, 6, 10, 6)
        layout.setSpacing(6)

        def btn(text: str, tip: str = "", checkable: bool = False) -> QPushButton:
            b = QPushButton(text)
            b.setToolTip(tip)
            b.setCheckable(checkable)
            b.setFixedHeight(34)
            b.setStyleSheet(
                "QPushButton{"
                "background:#3A3D5C;color:white;border:none;"
                "border-radius:5px;padding:0 14px;font-size:12px;}"
                "QPushButton:hover{background:#5A5D8C;}"
                "QPushButton:pressed{background:#2A2D4C;}"
                "QPushButton:checked{background:#C0392B;}"
                "QPushButton:disabled{background:#2A2A3A;color:#555;}"
            )
            return b

        def sep() -> QFrame:
            f = QFrame()
            f.setFrameShape(QFrame.Shape.VLine)
            f.setStyleSheet("color:#444;")
            return f

        self.open_btn = btn("📂  Open Drawing", "Open a PDF engineering drawing  (Ctrl+O)")
        self.open_btn.clicked.connect(self.open_drawing)

        self.add_callout_btn = btn(
            "✚  Add Callout",
            "Click on drawing to place a manual callout  (A to toggle, Esc to cancel)",
            checkable=True,
        )
        self.add_callout_btn.toggled.connect(self._toggle_add_callout_mode)

        self.export_pdf_btn = btn("📄  Export PDF", "Save annotated PDF  (Ctrl+P)")
        self.export_pdf_btn.clicked.connect(self.export_pdf)
        self.export_pdf_btn.setEnabled(False)

        self.export_record_btn = btn("📊  Export Record", "Save filled measurement record  (Ctrl+M)")
        self.export_record_btn.clicked.connect(self.export_measurement_record)
        self.export_record_btn.setEnabled(False)

        self.export_both_btn = btn("📦  Export Both", "Export annotated PDF + measurement record  (Ctrl+B)")
        self.export_both_btn.clicked.connect(self.export_both)
        self.export_both_btn.setEnabled(False)

        settings_btn = btn("⚙", "Settings")
        settings_btn.setFixedWidth(40)
        settings_btn.clicked.connect(self._open_settings)

        layout.addWidget(self.open_btn)
        layout.addWidget(sep())
        layout.addWidget(self.add_callout_btn)
        layout.addWidget(sep())
        layout.addWidget(self.export_pdf_btn)
        layout.addWidget(self.export_record_btn)
        layout.addWidget(self.export_both_btn)
        layout.addStretch()
        layout.addWidget(sep())
        layout.addWidget(settings_btn)

        return bar

    def _make_viewer_controls(self) -> QWidget:
        bar = QWidget()
        bar.setFixedHeight(36)
        bar.setStyleSheet("background:#F4F4F4;border-bottom:1px solid #DDD;")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(6, 3, 6, 3)
        layout.setSpacing(4)

        def small_btn(text: str, tip: str = "") -> QPushButton:
            b = QPushButton(text)
            b.setToolTip(tip)
            b.setFixedSize(28, 26)
            b.setStyleSheet(
                "QPushButton{background:#DDD;border:1px solid #BBB;"
                "border-radius:3px;font-size:13px;}"
                "QPushButton:hover{background:#CCC;}"
            )
            return b

        zi = small_btn("+", "Zoom in")
        zi.clicked.connect(self.pdf_viewer.zoom_in)
        zo = small_btn("−", "Zoom out")
        zo.clicked.connect(self.pdf_viewer.zoom_out)
        fit = QPushButton("Fit")
        fit.setFixedHeight(26)
        fit.setToolTip("Fit page to window")
        fit.setStyleSheet(
            "QPushButton{background:#DDD;border:1px solid #BBB;"
            "border-radius:3px;font-size:11px;padding:0 8px;}"
            "QPushButton:hover{background:#CCC;}"
        )
        fit.clicked.connect(self.pdf_viewer.zoom_reset)

        layout.addWidget(zi)
        layout.addWidget(zo)
        layout.addWidget(fit)

        self.zoom_lbl = QLabel("100%")
        self.zoom_lbl.setFixedWidth(44)
        self.zoom_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.zoom_lbl.setStyleSheet("color:#666; font-size:11px;")
        layout.addWidget(self.zoom_lbl)

        layout.addStretch()

        # Page navigation — Prev / label / Next
        self.prev_btn = small_btn("◀", "Previous page")
        self.prev_btn.clicked.connect(self._go_prev_page)
        self.prev_btn.setEnabled(False)

        self.page_lbl = QLabel("Page 1 / 1")
        self.page_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_lbl.setStyleSheet("color:#444; font-size:12px;")
        self.page_lbl.setFixedWidth(90)

        self.next_btn = small_btn("▶", "Next page")
        self.next_btn.clicked.connect(self._go_next_page)
        self.next_btn.setEnabled(False)

        layout.addWidget(self.prev_btn)
        layout.addWidget(self.page_lbl)
        layout.addWidget(self.next_btn)

        return bar

    # =======================================================================
    # Template loading
    # =======================================================================

    def _load_template(self) -> None:
        if parse_spreadsheet_template is None:
            return

        # Template is embedded as _TEMPLATE_B64. parse_spreadsheet_template
        # expects a file path, so decode to a temp file, parse, then delete.
        tmp_path: Optional[str] = None
        try:
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            with open(tmp_path, "wb") as fh:
                fh.write(base64.b64decode(_TEMPLATE_B64))
            self.template_structure = parse_spreadsheet_template(tmp_path)
            n_sheets = len(self.template_structure.get("sheet_names", []))
            self.status_bar.showMessage(
                f"Template loaded: {TEMPLATE_FILENAME}  ({n_sheets} sheet{'s' if n_sheets != 1 else ''})"
            )
        except Exception as exc:
            QMessageBox.warning(
                self,
                "Template Warning",
                f"Could not parse the embedded measurement template:<br>{exc}<br><br>"
                f"Measurement record export will be limited.",
            )
        finally:
            if tmp_path and os.path.isfile(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

    # =======================================================================
    # Opening a drawing
    # =======================================================================

    def open_drawing(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Engineering Drawing", "", "PDF Files (*.pdf)"
        )
        if not path:
            return

        try:
            if self.pdf_doc is not None:
                self.pdf_doc.close()
            self.pdf_doc = fitz.open(path)
        except Exception as exc:
            QMessageBox.critical(self, "Cannot Open PDF", f"Failed to open the drawing:\n\n{exc}")
            return

        self.current_pdf_path = path
        self.dimensions = []
        self._pages_done = set()
        self._session_burned_positions = []
        self._undo_stack = []

        # ── Session restore ──────────────────────────────────────────────────
        # Check whether a saved .autodim sidecar exists for this PDF.
        # If so, offer to restore the previous session instead of re-detecting.
        session = self._load_session_sidecar(path)
        if session:
            reply = QMessageBox.question(
                self,
                "Previous Session Found",
                "A saved session file was found for this drawing.\n\n"
                "Restore the previous callouts?\n\n"
                "• Yes — load the saved callout positions and skip re-detection.\n"
                "  When you re-export, old burned-in numbers will be removed\n"
                "  automatically before the updated ones are drawn.\n\n"
                "• No — run dimension detection from scratch (session ignored).",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._restore_from_session(session)
                return

        # ── Normal flow — detect from scratch ───────────────────────────────
        self.general_tolerances = self._detect_general_tolerances()
        if self.general_tolerances:
            parts: list[str] = []
            for k in sorted(k for k in self.general_tolerances if isinstance(k, int)):
                x_str = "X" if k == 0 else "." + "X" * k
                parts.append(f"{x_str} ±{self.general_tolerances[k]}")
            if "angular" in self.general_tolerances:
                parts.append(f"∠ ±{self.general_tolerances['angular']}°")
            self.status_bar.showMessage(
                "General tolerances detected: " + ",  ".join(parts)
            )

        n_pages = len(self.pdf_doc)
        self._update_page_nav(0, n_pages)
        self._go_to_page(0)

        self.export_pdf_btn.setEnabled(False)
        self.export_record_btn.setEnabled(False)
        self.export_both_btn.setEnabled(False)

    def _update_page_nav(self, current: int, total: int) -> None:
        self.page_lbl.setText(f"Page {current + 1} / {total}")
        self.prev_btn.setEnabled(current > 0)
        self.next_btn.setEnabled(current < total - 1)

    def _go_prev_page(self) -> None:
        if self.pdf_doc and self.current_page > 0:
            self._go_to_page(self.current_page - 1)

    def _go_next_page(self) -> None:
        if self.pdf_doc and self.current_page < len(self.pdf_doc) - 1:
            self._go_to_page(self.current_page + 1)

    def _go_to_page(self, new_page: int) -> None:
        if self.pdf_doc is None:
            return
        self._clear_highlight()
        self.current_page = new_page
        self._update_page_nav(new_page, len(self.pdf_doc))
        self._render_page()

        if new_page in self._pages_done:
            self._show_current_page_callouts()
            self._rebuild_checklist_for_page(new_page)
            self.right_stack.setCurrentIndex(0)
        else:
            self._run_detection()

    # =======================================================================
    # General-tolerance detection (title block, page 0)
    # =======================================================================

    def _detect_general_tolerances(self) -> dict:
        """Scan page 0 for a general-tolerance block.

        Tries the bottom-right corner first (where title blocks live on
        virtually every drawing), then falls back to the full page so the
        detection works even for non-standard layouts.
        """
        if self.pdf_doc is None or len(self.pdf_doc) == 0:
            return {}

        page = self.pdf_doc[0]
        pr = page.rect

        # Progressively larger crop windows anchored to the bottom-right,
        # then a full-page fallback.
        regions: list[Optional[fitz.Rect]] = [
            fitz.Rect(pr.width * 0.55, pr.height * 0.70, pr.width, pr.height),
            fitz.Rect(pr.width * 0.45, pr.height * 0.60, pr.width, pr.height),
            fitz.Rect(pr.width * 0.30, pr.height * 0.50, pr.width, pr.height),
            None,   # None → full page
        ]

        for clip in regions:
            text = page.get_text("text", clip=clip) if clip else page.get_text("text")
            tols = _parse_title_block_tolerances(text)
            if tols:
                return tols

        return {}

    def _apply_general_tolerances(self, dims: list) -> None:
        """For dims that carry no explicit tolerance, apply the matching
        general tolerance from the title block by decimal-place count.

        Stores results in ``dim["applied_plus_tol"]`` and
        ``dim["applied_minus_tol"]`` (both positive floats).  Dims that
        already have explicit tolerances parsed from their text are still
        populated so that the export path has a single field to consult.
        """
        if not self.general_tolerances:
            return

        for dim in dims:
            _nom, plus_t, minus_t = _parse_tolerances(dim["text"])

            if plus_t is not None:
                dim["applied_plus_tol"]  = plus_t
                dim["applied_minus_tol"] = minus_t if minus_t is not None else plus_t
                continue

            cat = dim.get("category", "")
            if cat == "angular":
                tol = self.general_tolerances.get("angular")
            else:
                decimals = _count_decimal_places(dim["text"])
                tol = self.general_tolerances.get(decimals)
                if tol is None:
                    for d in range(decimals - 1, -1, -1):
                        if d in self.general_tolerances:
                            tol = self.general_tolerances[d]
                            break

            if tol is not None:
                dim["applied_plus_tol"]  = tol
                dim["applied_minus_tol"] = tol

    # =======================================================================
    # Session sidecar — save / restore previously annotated drawings
    # =======================================================================

    @staticmethod
    def _sidecar_path(pdf_path: str) -> str:
        """Return the .autodim sidecar path for a given PDF path."""
        return os.path.splitext(pdf_path)[0] + SIDECAR_EXT

    def _load_session_sidecar(self, pdf_path: str) -> Optional[dict]:
        """Return parsed sidecar JSON if one exists next to pdf_path, else None."""
        sp = self._sidecar_path(pdf_path)
        if not os.path.isfile(sp):
            return None
        try:
            with open(sp, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            return None

    def _save_session_sidecar(self, pdf_path: str) -> None:
        """Persist all dimension data to a .autodim sidecar next to pdf_path."""
        # JSON requires string keys; convert int tolerance-level keys → str.
        tols_serialisable = {str(k): v for k, v in self.general_tolerances.items()}

        session: dict[str, Any] = {
            "version": "1.0",
            "callout_color": self.config.callout_color,
            "callout_radius": self.config.callout_radius,
            "general_tolerances": tols_serialisable,
            "dimensions": [
                {
                    "text":             d["text"],
                    "page":             d.get("page", 0),
                    "x":                d.get("x", 0.0),
                    "y":                d.get("y", 0.0),
                    "width":            d.get("width", 0.0),
                    "height":           d.get("height", 0.0),
                    "callout_x":        d["callout_x"],
                    "callout_y":        d["callout_y"],
                    "callout_num":      d.get("callout_num", 0),
                    "included":         d.get("included", True),
                    "category":         d.get("category", "other"),
                    "confidence":       d.get("confidence", 1.0),
                    "applied_plus_tol": d.get("applied_plus_tol"),
                    "applied_minus_tol":d.get("applied_minus_tol"),
                }
                for d in self.dimensions
            ],
        }
        sp = self._sidecar_path(pdf_path)
        try:
            with open(sp, "w", encoding="utf-8") as fh:
                json.dump(session, fh, indent=2)
        except Exception as exc:
            print(f"[session] could not save sidecar: {exc}")

    def _restore_from_session(self, session: dict) -> None:
        """Populate self.dimensions from a loaded sidecar and skip detection."""
        # Restore appearance config
        if "callout_color" in session:
            self.config.callout_color = session["callout_color"]
        if "callout_radius" in session:
            self.config.callout_radius = int(session["callout_radius"])

        # Restore general tolerances — JSON turns int keys into strings
        raw = session.get("general_tolerances", {})
        self.general_tolerances = {
            (int(k) if k.lstrip("-").isdigit() else k): v
            for k, v in raw.items()
        }

        # Rebuild dimensions list (no callout_item / row_widget yet)
        self.dimensions = []
        for d in session.get("dimensions", []):
            entry = {k: v for k, v in d.items()}
            entry["callout_item"] = None
            entry["row_widget"]   = None
            self.dimensions.append(entry)

        # Record which circles were burned into this PDF so _write_annotated_pdf
        # can redact them before drawing the updated ones.
        radius_pt = self.config.callout_radius / RENDER_ZOOM
        self._session_burned_positions = [
            {
                "page":   d.get("page", 0),
                "x":      d["callout_x"],
                "y":      d["callout_y"],
                "r_pt":   radius_pt,
            }
            for d in session.get("dimensions", [])
            if d.get("included", True) and d.get("callout_num", 0) > 0
        ]

        # ── Erase burned-in circles from the in-memory document ──────────────
        # Apply the same redactions to self.pdf_doc RIGHT NOW so the viewer
        # renders clean pages immediately.  The on-disk file is NOT modified;
        # _write_annotated_pdf opens it fresh and redacts independently.
        if self._session_burned_positions:
            pages_touched: set[int] = set()
            for bp in self._session_burned_positions:
                pg = bp["page"]
                if pg >= len(self.pdf_doc):
                    continue
                r = bp["r_pt"]
                cx, cy = bp["x"], bp["y"]
                try:
                    self.pdf_doc[pg].add_redact_annot(
                        fitz.Rect(cx - r, cy - r, cx + r, cy + r),
                        fill=(1, 1, 1),
                    )
                    pages_touched.add(pg)
                except Exception:
                    pass
            for pg in pages_touched:
                try:
                    self.pdf_doc[pg].apply_redactions()
                except Exception:
                    pass

        # Mark every page as already processed — no detection will run
        n_pages = len(self.pdf_doc)
        self._pages_done = set(range(n_pages))

        self._update_page_nav(0, n_pages)
        self._go_to_page(0)          # renders the now-clean page + shows callouts

        self.export_pdf_btn.setEnabled(True)
        self.export_record_btn.setEnabled(self.template_structure is not None)
        self.export_both_btn.setEnabled(self.template_structure is not None)

        n_active = sum(1 for d in self.dimensions if d.get("included", True))
        n_total  = len(self.dimensions)
        self.status_bar.showMessage(
            f"Session restored — {n_active} active callout{'s' if n_active != 1 else ''} "
            f"({n_total} total).  Move or toggle callouts, then re-export to apply changes."
        )

    def _render_page(self) -> None:
        if self.pdf_doc is None:
            return
        # Remove callout items from scene BEFORE scene.clear() inside display_pixmap.
        # This prevents Qt from deleting C++ objects we still hold references to.
        self.pdf_viewer.clear_callout_items()
        for dim in self.dimensions:
            dim["callout_item"] = None

        page = self.pdf_doc[self.current_page]

        # Render at native physical-pixel resolution so the image is sharp on
        # HiDPI / scaled displays (e.g. Windows 125 %, 150 %, 200 % DPI).
        # dpr == 1.0 on a standard display → no change in behaviour.
        dpr: float = self.devicePixelRatio()
        mat = fitz.Matrix(RENDER_ZOOM * dpr, RENDER_ZOOM * dpr)
        pix = page.get_pixmap(matrix=mat)
        img = QImage(
            pix.samples, pix.width, pix.height,
            pix.stride, QImage.Format.Format_RGB888,
        )
        pixmap = QPixmap.fromImage(img)
        # Tell Qt the logical size is still RENDER_ZOOM × PDF-pts; coordinate
        # conversions that rely on RENDER_ZOOM are unaffected.
        pixmap.setDevicePixelRatio(dpr)
        self.pdf_viewer.display_pixmap(pixmap)
        self.pdf_viewer.zoom_reset()

    # =======================================================================
    # Dimension detection
    # =======================================================================

    def _run_detection(self) -> None:
        self.right_stack.setCurrentIndex(1)
        self.spinner.start()
        self.status_bar.showMessage(
            f"Analyzing page {self.current_page + 1} of '{os.path.basename(self.current_pdf_path)}'…"
        )

        self._worker = DetectionWorker(self.current_pdf_path, self.current_page)
        self._worker.finished.connect(self._on_detection_done)
        self._worker.error.connect(self._on_detection_error)
        self._worker.start()

    @pyqtSlot(list)
    def _on_detection_done(self, dims: list) -> None:
        self.spinner.stop()

        page_num = self.current_page
        r_pt = self.config.callout_radius / RENDER_ZOOM

        # Build list of already-placed callout centres for this page
        placed: list[tuple[float, float]] = [
            (d["callout_x"], d["callout_y"])
            for d in self.dimensions
            if d.get("page") == page_num and d.get("included", True)
        ]

        for raw in dims:
            _, plus_t, _ = _parse_tolerances(raw.get("text", ""))
            cx, cy = _smart_callout_pos(raw, placed, r_pt, prefer_left=(plus_t is not None))
            placed.append((cx, cy))

            entry: dict[str, Any] = dict(raw)
            entry.update({
                "page":       page_num,
                "callout_x":  cx,
                "callout_y":  cy,
                "included":   True,
                "callout_num": 0,
                "callout_item": None,
                "row_widget": None,
            })
            self.dimensions.append(entry)

        self._pages_done.add(page_num)
        # Fill in general tolerances for dims that have none explicitly
        self._apply_general_tolerances(
            [d for d in self.dimensions if d.get("page") == page_num]
        )
        self._update_callout_numbers()
        self._show_current_page_callouts()
        self._rebuild_checklist_for_page(page_num)

        self.right_stack.setCurrentIndex(0)
        self.export_pdf_btn.setEnabled(True)
        self.export_record_btn.setEnabled(self.template_structure is not None)
        self.export_both_btn.setEnabled(self.template_structure is not None)

        n_page = sum(1 for d in self.dimensions if d.get("page") == page_num)
        n_total = len(self.dimensions)
        self.status_bar.showMessage(
            f"Page {page_num + 1}: {n_page} dimension{'s' if n_page != 1 else ''} found  "
            f"({n_total} total across {len(self._pages_done)} page{'s' if len(self._pages_done) != 1 else ''})."
        )

    @pyqtSlot(str)
    def _on_detection_error(self, msg: str) -> None:
        self.spinner.stop()
        self.right_stack.setCurrentIndex(0)
        QMessageBox.critical(
            self,
            "Detection Error",
            f"Dimension detection failed:\n\n{msg}\n\n"
            "If the PDF is scanned, run OCR on it first (e.g. ocrmypdf).",
        )
        self.status_bar.showMessage("Detection failed — see error dialog.")

    # =======================================================================
    # Checklist & callout item management
    # =======================================================================

    def _rebuild_checklist_for_page(self, page_num: int) -> None:
        """Rebuild the right-hand checklist showing only dims for page_num."""
        # Clear highlights before deleting old row widgets to avoid stale refs
        self._clear_highlight()
        # Clear row_widget references for all dims (old rows are about to be deleted)
        for dim in self.dimensions:
            dim["row_widget"] = None

        self.checklist.clear_rows()

        page_dims = [d for d in self.dimensions if d.get("page") == page_num]
        for dim in page_dims:
            # Determine whether the tolerance is explicit (in the text) or
            # inferred from the title-block general tolerances.
            _, explicit_plus, _ = _parse_tolerances(dim["text"])
            tol_inferred = (
                explicit_plus is None
                and dim.get("applied_plus_tol") is not None
            )
            row = DimensionRow(
                dim_index=self.dimensions.index(dim),
                text=dim["text"],
                category=dim["category"],
                confidence=dim["confidence"],
                callout_color=self.config.callout_color,
                applied_plus_tol=dim.get("applied_plus_tol"),
                applied_minus_tol=dim.get("applied_minus_tol"),
                tol_inferred=tol_inferred,
                initial_enabled=dim.get("included", True),
            )
            row.check_changed.connect(self._on_check_changed)
            row.edit_requested.connect(self._on_edit_requested)
            row.focus_requested.connect(self._on_row_focused)
            dim["row_widget"] = row
            self.checklist.add_row(row)

        n_active = sum(1 for d in page_dims if d.get("included", True))
        self.checklist.set_count(len(page_dims), n_active)
        # Sync badge numbers on the newly created rows
        self._update_callout_numbers()

    def _show_current_page_callouts(self) -> None:
        """Recreate callout items in the scene for the current page's dims."""
        # Items were already cleared from scene in _render_page; callout_item = None
        for dim in self.dimensions:
            if dim.get("page") != self.current_page:
                continue
            item = self._make_callout_item(dim)
            dim["callout_item"] = item
            self.pdf_viewer.add_callout_item(item)
        self._update_callout_numbers()

    def _make_callout_item(self, dim: dict) -> CalloutItem:
        """Create and position a CalloutItem for a dimension dict."""
        def on_moved(item: CalloutItem, pos: QPointF, d: dict = dim) -> None:
            d["callout_x"] = pos.x() / RENDER_ZOOM
            d["callout_y"] = pos.y() / RENDER_ZOOM

        def on_drag_complete(
            item: CalloutItem,
            old_scene: QPointF,
            new_scene: QPointF,
            d: dict = dim,
        ) -> None:
            self._push_undo({
                "type":      "move",
                "dim_index": self.dimensions.index(d),
                "old_x":     old_scene.x() / RENDER_ZOOM,
                "old_y":     old_scene.y() / RENDER_ZOOM,
                "new_x":     new_scene.x() / RENDER_ZOOM,
                "new_y":     new_scene.y() / RENDER_ZOOM,
            })
            # Renumber in new reading order now that the position has changed
            self._update_callout_numbers()

        def on_click(item: CalloutItem, d: dict = dim) -> None:
            self._clear_highlight()
            self._highlighted_callout = item
            item.set_highlighted(True)
            row = d.get("row_widget")
            if row is not None:
                self._highlighted_row = row
                row.set_highlighted(True)
                self.checklist.scroll.ensureWidgetVisible(row)

        item = CalloutItem(
            number=1,
            radius=self.config.callout_radius,
            color=self.config.callout_color,
            on_moved=on_moved,
            on_drag_complete=on_drag_complete,
            on_context_menu=self._on_callout_context_menu,
            on_click=on_click,
        )
        item.setPos(dim["callout_x"] * RENDER_ZOOM, dim["callout_y"] * RENDER_ZOOM)
        return item

    def _update_callout_numbers(self) -> None:
        """Reassign sequential callout numbers in reading order.

        Included callouts are sorted page-first, then top-to-bottom in
        bands (callouts within one circle-diameter vertically are treated as
        the same 'row'), then left-to-right within each band.  This means
        inserting or moving a callout anywhere on the page automatically
        renumbers everything else to maintain correct reading order.
        """
        r_pt  = self.config.callout_radius / RENDER_ZOOM
        band  = max(r_pt * 2.0, 15.0)   # vertical tolerance for same-row grouping

        def _reading_key(d: dict) -> tuple:
            return (
                d.get("page", 0),
                round(d.get("callout_y", 0.0) / band),  # row band
                d.get("callout_x", 0.0),                # left→right within row
            )

        # Assign numbers to included dims in reading order
        included_sorted = sorted(
            (d for d in self.dimensions if d.get("included", True)),
            key=_reading_key,
        )
        for num, dim in enumerate(included_sorted, start=1):
            dim["callout_num"] = num
            row: Optional[DimensionRow] = dim.get("row_widget")
            item: Optional[CalloutItem] = dim.get("callout_item")
            if row is not None:
                row.set_callout_number(num)
                row.set_color(self.config.callout_color)
            if item is not None:
                item.set_number(num)
                item.set_color(self.config.callout_color)
                item.setVisible(True)

        # Zero-out and hide excluded dims
        for dim in self.dimensions:
            if dim.get("included", True):
                continue
            dim["callout_num"] = 0
            row = dim.get("row_widget")
            item = dim.get("callout_item")
            if row is not None:
                row.mark_excluded()
            if item is not None:
                item.setVisible(False)

        # Reorder checklist rows to match the current reading order.
        # All dims (enabled and disabled) are sorted by their position on the page
        # so enabled rows always appear 1→N and disabled rows stay interleaved at
        # their spatial position rather than collapsing to the bottom.
        page_dims = [d for d in self.dimensions if d.get("page") == self.current_page]
        ordered_rows = [
            d["row_widget"]
            for d in sorted(page_dims, key=_reading_key)
            if d.get("row_widget") is not None
        ]
        if ordered_rows:
            self.checklist.reorder_rows(ordered_rows)

        # Keep the checklist header count in sync after any numbering change
        n_total  = len(page_dims)
        n_active = sum(1 for d in page_dims if d.get("included", True))
        self.checklist.set_count(n_total, n_active)

    @pyqtSlot(int, bool)
    def _on_check_changed(self, dim_index: int, checked: bool) -> None:
        if 0 <= dim_index < len(self.dimensions):
            old_state = self.dimensions[dim_index].get("included", True)
            self.dimensions[dim_index]["included"] = checked
            self._push_undo({
                "type":      "toggle",
                "dim_index": dim_index,
                "old_state": old_state,
                "new_state": checked,
            })
            self._update_callout_numbers()

    @pyqtSlot(int)
    def _on_edit_requested(self, dim_index: int) -> None:
        """Open a pre-filled LabelDialog so the user can correct a dimension."""
        if not (0 <= dim_index < len(self.dimensions)):
            return

        dim = self.dimensions[dim_index]

        # Pre-populate from the current text
        nom_str = _nominal_text(dim["text"])
        _, plus_t, minus_t = _parse_tolerances(dim["text"])
        plus_str  = f"{plus_t:g}"  if plus_t  is not None else ""
        minus_str = f"{minus_t:g}" if minus_t is not None else ""

        dlg = LabelDialog(
            self,
            title="Edit Dimension",
            initial_nominal=nom_str,
            initial_plus_tol=plus_str,
            initial_minus_tol=minus_str,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        # Update the dimension text and re-derive applied tolerances
        dim["text"] = dlg.label_text
        self._apply_general_tolerances([dim])

        # Rebuild the checklist so the row reflects the new values
        self._rebuild_checklist_for_page(self.current_page)
        self.status_bar.showMessage(
            f"Dimension #{dim_index + 1} updated to: {dim['text']}"
        )

    # =======================================================================
    # Keyboard shortcuts
    # =======================================================================

    def keyPressEvent(self, event) -> None:
        key   = event.key()
        mods  = event.modifiers()
        ctrl  = Qt.KeyboardModifier.ControlModifier
        shift = Qt.KeyboardModifier.ShiftModifier

        if key == Qt.Key.Key_Escape and self._add_callout_mode:
            self.add_callout_btn.setChecked(False)
        elif key == Qt.Key.Key_Z and (mods & ctrl) and (mods & shift):
            self._redo_last()
        elif key == Qt.Key.Key_Z and (mods & ctrl):
            self._undo_last()
        elif key == Qt.Key.Key_Y and (mods & ctrl):
            self._redo_last()
        elif key == Qt.Key.Key_O and (mods & ctrl):
            self.open_drawing()
        elif key == Qt.Key.Key_P and (mods & ctrl):
            self.export_pdf()
        elif key == Qt.Key.Key_M and (mods & ctrl):
            self.export_measurement_record()
        elif key == Qt.Key.Key_B and (mods & ctrl):
            self.export_both()
        elif key == Qt.Key.Key_A and not mods:
            if self.pdf_doc is not None:
                self.add_callout_btn.setChecked(not self.add_callout_btn.isChecked())
        elif key == Qt.Key.Key_Left:
            self._go_prev_page()
        elif key == Qt.Key.Key_Right:
            self._go_next_page()
        else:
            super().keyPressEvent(event)

    # =======================================================================
    # Undo
    # =======================================================================

    _MAX_UNDO = 50

    def _push_undo(self, action: dict) -> None:
        self._undo_stack.append(action)
        if len(self._undo_stack) > self._MAX_UNDO:
            self._undo_stack.pop(0)
        self._redo_stack.clear()  # new action invalidates the redo history

    def _undo_last(self) -> None:
        if not self._undo_stack:
            self.status_bar.showMessage("Nothing to undo.")
            return
        action = self._undo_stack.pop()
        if action["type"] == "move":
            self._redo_stack.append({
                "type": "move", "dim_index": action["dim_index"],
                "old_x": action["new_x"], "old_y": action["new_y"],
                "new_x": action["old_x"], "new_y": action["old_y"],
            })
            self._undo_move(action)
        elif action["type"] == "toggle":
            self._redo_stack.append({
                "type": "toggle", "dim_index": action["dim_index"],
                "old_state": action["new_state"], "new_state": action["old_state"],
            })
            self._undo_toggle(action)

    def _undo_move(self, action: dict) -> None:
        idx = action["dim_index"]
        if not (0 <= idx < len(self.dimensions)):
            return
        dim = self.dimensions[idx]
        old_x, old_y = action["old_x"], action["old_y"]
        item: Optional[CalloutItem] = dim.get("callout_item")
        if item is not None:
            # setPos triggers on_moved which updates dim["callout_x/y"] for us
            item.setPos(old_x * RENDER_ZOOM, old_y * RENDER_ZOOM)
        else:
            dim["callout_x"] = old_x
            dim["callout_y"] = old_y
        # Renumber in reading order after the position is restored
        self._update_callout_numbers()
        self.status_bar.showMessage(
            f"Undone: callout move  (now #{dim.get('callout_num', '?')})."
        )

    def _undo_toggle(self, action: dict) -> None:
        idx = action["dim_index"]
        if not (0 <= idx < len(self.dimensions)):
            return
        old_state: bool = action["old_state"]
        dim = self.dimensions[idx]
        row: Optional[DimensionRow] = dim.get("row_widget")
        if row is not None:
            # blockSignals prevents _on_check_changed from pushing another undo entry
            row.toggle_btn.blockSignals(True)
            row.toggle_btn.setChecked(old_state)
            row.toggle_btn.blockSignals(False)
            row._refresh_row_style(old_state)
            row._apply_badge_style(
                row._badge_color if old_state else "#AAAAAA", active=old_state
            )
        dim["included"] = old_state
        self._update_callout_numbers()
        self.status_bar.showMessage(
            f"Undone: toggle for dimension #{idx + 1} "
            f"→ {'On' if old_state else 'Off'}."
        )

    def _redo_last(self) -> None:
        if not self._redo_stack:
            self.status_bar.showMessage("Nothing to redo.")
            return
        action = self._redo_stack.pop()
        if action["type"] == "move":
            # Push the reverse back onto undo so Ctrl+Z can undo this redo
            self._undo_stack.append({
                "type": "move", "dim_index": action["dim_index"],
                "old_x": action["new_x"], "old_y": action["new_y"],
                "new_x": action["old_x"], "new_y": action["old_y"],
            })
            self._undo_move(action)
            dim = self.dimensions[action["dim_index"]] if 0 <= action["dim_index"] < len(self.dimensions) else {}
            self.status_bar.showMessage(f"Redone: callout move  (now #{dim.get('callout_num', '?')}).")
        elif action["type"] == "toggle":
            self._undo_stack.append({
                "type": "toggle", "dim_index": action["dim_index"],
                "old_state": action["new_state"], "new_state": action["old_state"],
            })
            self._undo_toggle(action)
            idx = action["dim_index"]
            self.status_bar.showMessage(
                f"Redone: toggle for dimension #{idx + 1} → {'On' if action['old_state'] else 'Off'}."
            )

    # =======================================================================
    # Add-callout mode
    # =======================================================================

    def _toggle_add_callout_mode(self, checked: bool) -> None:
        self._add_callout_mode = checked
        self.pdf_viewer.set_add_mode(checked)
        if checked:
            self.status_bar.showMessage(
                "Add Callout mode active — click on the drawing to place a callout.  Press Esc to cancel."
            )
        else:
            self.status_bar.showMessage("Normal mode.")

    @pyqtSlot(float, float)
    def _on_callout_place_requested(self, scene_x: float, scene_y: float) -> None:
        # Exit add mode immediately so the user isn't stuck in crosshair mode
        self.add_callout_btn.setChecked(False)

        dlg = LabelDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        label = dlg.label_text
        if not label:
            return

        pdf_x = scene_x / RENDER_ZOOM
        pdf_y = scene_y / RENDER_ZOOM
        idx = len(self.dimensions)

        dim: dict[str, Any] = {
            "index":       idx,
            "text":        label,
            "x":           pdf_x,
            "y":           pdf_y,
            "width":       0.0,
            "height":      0.0,
            "confidence":  1.0,
            "category":    "other",
            "page":        self.current_page,
            "included":    True,
            "callout_x":   pdf_x,
            "callout_y":   pdf_y,
            "callout_num": 0,
            "callout_item": None,
            "row_widget":   None,
        }
        self.dimensions.append(dim)
        self._apply_general_tolerances([dim])   # fill tolerances from title block

        _, explicit_plus, _ = _parse_tolerances(label)
        tol_inferred = (
            explicit_plus is None and dim.get("applied_plus_tol") is not None
        )
        row = DimensionRow(
            dim_index=idx,
            text=label,
            category="other",
            confidence=1.0,
            callout_color=self.config.callout_color,
            applied_plus_tol=dim.get("applied_plus_tol"),
            applied_minus_tol=dim.get("applied_minus_tol"),
            tol_inferred=tol_inferred,
            initial_enabled=True,
        )
        row.check_changed.connect(self._on_check_changed)
        row.edit_requested.connect(self._on_edit_requested)
        dim["row_widget"] = row
        self.checklist.add_row(row)
        self.checklist.set_count(
            sum(1 for d in self.dimensions if d.get("page") == self.current_page)
        )

        item = self._make_callout_item(dim)
        dim["callout_item"] = item
        self.pdf_viewer.add_callout_item(item)

        self._update_callout_numbers()
        self.export_pdf_btn.setEnabled(True)
        if self.template_structure is not None:
            self.export_record_btn.setEnabled(True)
            self.export_both_btn.setEnabled(True)

    @pyqtSlot(object)
    def _on_callout_delete_requested(self, item: "CalloutItem") -> None:
        """Delete/Backspace pressed while a callout circle is selected — uncheck it."""
        for dim in self.dimensions:
            if dim.get("callout_item") is item:
                row: Optional[DimensionRow] = dim.get("row_widget")
                if row is not None:
                    # Let the toggle drive everything (visual update + signal chain)
                    if row.toggle_btn.isChecked():
                        row.toggle_btn.setChecked(False)
                else:
                    # Dimension is on a different page — update state directly
                    dim["included"] = False
                    self._update_callout_numbers()
                return

    def _on_callout_context_menu(self, item: "CalloutItem", screen_pos) -> None:
        """Show a right-click context menu on a callout circle."""
        # Find which dimension this item belongs to
        dim_index: Optional[int] = None
        for i, dim in enumerate(self.dimensions):
            if dim.get("callout_item") is item:
                dim_index = i
                break
        if dim_index is None:
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #2B2D42;
                border: 1px solid #555;
                border-radius: 6px;
                padding: 4px 0px;
                color: #EEEEEE;
                font-size: 13px;
            }
            QMenu::item {
                padding: 7px 20px 7px 12px;
                border-radius: 4px;
                margin: 2px 4px;
            }
            QMenu::item:selected {
                background-color: #4A90D9;
                color: white;
            }
        """)
        edit_action   = menu.addAction("✏   Edit Dimension")
        menu.addSeparator()
        delete_action = menu.addAction("🗑   Delete Dimension")

        chosen = menu.exec(screen_pos)
        if chosen is edit_action:
            self._on_edit_requested(dim_index)
        elif chosen is delete_action:
            self._on_callout_delete_requested(item)

    # =======================================================================
    # Two-way callout ↔ checklist row linking
    # =======================================================================

    def _on_row_focused(self, dim_index: int) -> None:
        """Focusing a checklist row highlights the callout and centers the view."""
        self._clear_highlight()
        if not (0 <= dim_index < len(self.dimensions)):
            return
        dim = self.dimensions[dim_index]
        item = dim.get("callout_item")
        if item is None or item.scene() is None:
            return
        self._highlighted_callout = item
        item.set_highlighted(True)
        # Defer centerOn one event-loop tick so Qt finishes processing the
        # focus change before we scroll the viewer (avoids off-center snap).
        QTimer.singleShot(0, lambda: self.pdf_viewer.centerOn(item))

    def _clear_highlight(self) -> None:
        if self._highlighted_row is not None:
            self._highlighted_row.set_highlighted(False)
            self._highlighted_row = None
        if self._highlighted_callout is not None:
            self._highlighted_callout.set_highlighted(False)
            self._highlighted_callout = None

    # =======================================================================
    # Settings
    # =======================================================================

    def _open_settings(self) -> None:
        dlg = SettingsDialog(self.config, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            for dim in self.dimensions:
                if (item := dim.get("callout_item")) is not None:
                    item.set_color(self.config.callout_color)
                    item.set_radius(self.config.callout_radius)
                if (row := dim.get("row_widget")) is not None and dim.get("included"):
                    row.set_color(self.config.callout_color)

    # =======================================================================
    # Detect remaining pages (called before export)
    # =======================================================================

    def _detect_all_remaining_pages(self) -> None:
        """Synchronously detect dimensions on any pages not yet visited."""
        if self.pdf_doc is None or detect_dimensions is None:
            return

        unvisited = [p for p in range(len(self.pdf_doc)) if p not in self._pages_done]
        if not unvisited:
            return

        progress = QProgressDialog(
            "Detecting dimensions on remaining pages…",
            None,  # no cancel button
            0, len(unvisited),
            self,
        )
        progress.setWindowTitle("Analyzing Drawing")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(400)
        progress.setValue(0)

        r_pt = self.config.callout_radius / RENDER_ZOOM

        for step, page_num in enumerate(unvisited):
            progress.setValue(step)
            progress.setLabelText(
                f"Analyzing page {page_num + 1} of {len(self.pdf_doc)}…"
            )
            QApplication.processEvents()

            try:
                raw_dims = detect_dimensions(self.current_pdf_path, page_num)
            except Exception:
                self._pages_done.add(page_num)
                continue

            placed: list[tuple[float, float]] = [
                (d["callout_x"], d["callout_y"])
                for d in self.dimensions
                if d.get("page") == page_num and d.get("included", True)
            ]

            for raw in raw_dims:
                _, plus_t, _ = _parse_tolerances(raw.get("text", ""))
                cx, cy = _smart_callout_pos(raw, placed, r_pt, prefer_left=(plus_t is not None))
                placed.append((cx, cy))

                entry = dict(raw)
                entry.update({
                    "page":        page_num,
                    "callout_x":   cx,
                    "callout_y":   cy,
                    "included":    True,
                    "callout_num": 0,
                    "callout_item": None,
                    "row_widget":   None,
                })
                self.dimensions.append(entry)

            self._pages_done.add(page_num)
            self._apply_general_tolerances(
                [d for d in self.dimensions if d.get("page") == page_num]
            )

        progress.setValue(len(unvisited))
        self._update_callout_numbers()

    # =======================================================================
    # Export — annotated PDF
    # =======================================================================

    def export_pdf(self) -> None:
        if not self.current_pdf_path or self.pdf_doc is None:
            return

        self._detect_all_remaining_pages()

        base = os.path.splitext(os.path.basename(self.current_pdf_path))[0]
        default_out = os.path.join(
            os.path.dirname(self.current_pdf_path), f"{base}_annotated.pdf"
        )
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Export Annotated PDF", default_out, "PDF Files (*.pdf)"
        )
        if not out_path:
            return

        active = [d for d in self.dimensions if d.get("included")]
        if not active:
            QMessageBox.information(self, "Nothing to Export", "No callouts are currently included.")
            return

        try:
            self._write_annotated_pdf(out_path, active)
            self.status_bar.showMessage(f"PDF exported → {out_path}")
            reply = QMessageBox.information(
                self, "Export Complete",
                f"Annotated PDF saved:\n{out_path}\n\nReveal in Explorer?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._reveal_in_explorer(out_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", f"Could not write PDF:\n\n{exc}")

    # =======================================================================
    # Shared annotation helper
    # =======================================================================

    def _annotate_doc(
        self, doc: "fitz.Document", active_dims: list
    ) -> "dict[int, list]":
        """Draw callout circles on *doc* in-place.

        Returns a ``dims_by_page`` dict mapping page index → list of dims.
        Mutates the fitz.Document so caller can either save it or render
        pages to images.
        """
        r_pt = self.config.callout_radius / RENDER_ZOOM

        qt_color = QColor(self.config.callout_color)
        stroke_rgb = (qt_color.redF(), qt_color.greenF(), qt_color.blueF())

        # Base font size — will be scaled per number to always fit the circle.
        # diameter = 2*r_pt; target ≈ 60 % of diameter → r_pt * 1.2
        base_font_sz = max(8.0, r_pt * 1.2)
        # Available width inside the circle border (2 pt margin each side)
        _available_w = max(1.0, 2.0 * r_pt - 4.0)

        dims_by_page: dict[int, list] = {}
        for dim in active_dims:
            pg = dim.get("page", self.current_page)
            dims_by_page.setdefault(pg, []).append(dim)

        for page_num, page_dims in dims_by_page.items():
            if page_num >= len(doc):
                continue
            page = doc[page_num]
            for dim in page_dims:
                cx  = dim["callout_x"]
                cy  = dim["callout_y"]
                num_str = str(dim["callout_num"])

                circle_rect = fitz.Rect(cx - r_pt, cy - r_pt, cx + r_pt, cy + r_pt)

                # ── Font size: shrink to fit wide numbers inside the circle ──
                # Measure at the default size, then scale down proportionally
                # so the text always fits within the available diameter.
                try:
                    tw_default = fitz.get_text_length(
                        num_str, fontname="hebo", fontsize=base_font_sz
                    )
                except AttributeError:
                    tw_default = base_font_sz * 0.60 * len(num_str)
                if tw_default > _available_w and tw_default > 0:
                    font_sz = max(6.0, base_font_sz * _available_w / tw_default)
                else:
                    font_sz = base_font_sz

                # ── Circle: white fill + coloured border ─────────────────
                page.draw_oval(
                    circle_rect,
                    color=stroke_rgb, fill=(1.0, 1.0, 1.0),
                    width=2.0,
                )

                # ── Vertically centred number ─────────────────────────────
                # Use insert_text with manually computed centre so the
                # number always renders regardless of circle size or
                # PyMuPDF version (insert_textbox can silently swallow
                # text when the rect is just a little too tight).
                if dim["callout_num"] > 0:
                    # Horizontal: measure exact glyph width at final font_sz.
                    try:
                        tw = fitz.get_text_length(
                            num_str, fontname="hebo", fontsize=font_sz
                        )
                    except AttributeError:
                        tw = font_sz * 0.60 * len(num_str)
                    tx = cx - tw / 2.0
                    # Vertical: insert_text places the baseline at the given
                    # y-coordinate.  For digits the ascender ≈ 0.72×font_sz,
                    # so the visual centre sits 0.36×font_sz above baseline.
                    # Solve for baseline: baseline = cy + 0.36×font_sz.
                    ty = cy + font_sz * 0.36
                    page.insert_text(
                        fitz.Point(tx, ty),
                        num_str,
                        fontsize=font_sz,
                        fontname="hebo",
                        color=stroke_rgb,
                    )

        return dims_by_page

    # =======================================================================
    # Export — annotated PDF
    # =======================================================================

    def _write_annotated_pdf(self, out_path: str, active_dims: list) -> None:
        doc = fitz.open(self.current_pdf_path)

        # If we restored a session from a previously annotated PDF, the old
        # callout circles are burned into the content.  Redact (white-fill)
        # those exact positions before drawing the updated ones so that no
        # stale numbers show through underneath.
        if self._session_burned_positions:
            pages_touched: set[int] = set()
            for bp in self._session_burned_positions:
                pg = bp["page"]
                if pg >= len(doc):
                    continue
                r = bp["r_pt"]
                cx, cy = bp["x"], bp["y"]
                doc[pg].add_redact_annot(
                    fitz.Rect(cx - r, cy - r, cx + r, cy + r),
                    fill=(1, 1, 1),
                )
                pages_touched.add(pg)
            for pg in pages_touched:
                doc[pg].apply_redactions()

        self._annotate_doc(doc, active_dims)
        doc.save(out_path, garbage=4, deflate=True)
        doc.close()

        # Save / update the session sidecar next to the exported PDF so the
        # user can re-open this annotated file and resume editing later.
        self._save_session_sidecar(out_path)

        # Update burned positions to reflect what was just drawn, so a
        # subsequent re-export from this same session redacts correctly.
        r_pt = self.config.callout_radius / RENDER_ZOOM
        self._session_burned_positions = [
            {"page": d.get("page", 0), "x": d["callout_x"], "y": d["callout_y"], "r_pt": r_pt}
            for d in self.dimensions
            if d.get("included", True) and d.get("callout_num", 0) > 0
        ]

    # =======================================================================
    # Export — measurement record
    # =======================================================================

    def export_measurement_record(self) -> None:
        if load_workbook is None:
            QMessageBox.critical(self, "Missing Library", "openpyxl is not installed.")
            return

        self._detect_all_remaining_pages()

        # Sort by callout_num so spreadsheet rows match the numbered drawing
        active = sorted(
            (d for d in self.dimensions if d.get("included")),
            key=lambda d: d.get("callout_num", 0),
        )
        if not active:
            QMessageBox.information(self, "Nothing to Export", "No callouts are currently included.")
            return

        base = (
            os.path.splitext(os.path.basename(self.current_pdf_path))[0]
            if self.current_pdf_path else "drawing"
        )
        default_dir = os.path.dirname(self.current_pdf_path) if self.current_pdf_path else _APP_DIR
        default_out = os.path.join(default_dir, f"{base}_measurement_record.xlsx")

        out_path, _ = QFileDialog.getSaveFileName(
            self, "Export Measurement Record", default_out, "Excel Files (*.xlsx)"
        )
        if not out_path:
            return

        try:
            self._write_measurement_record(out_path, active)
            self.status_bar.showMessage(f"Measurement record exported → {out_path}")
            reply = QMessageBox.information(
                self, "Export Complete",
                f"Measurement record saved:\n{out_path}\n\n"
                "All original formulas, formatting, and conditional rules are preserved.\n\n"
                "Reveal in Explorer?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._reveal_in_explorer(out_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", f"Could not write measurement record:\n\n{exc}")

    def export_both(self) -> None:
        """Export annotated PDF and measurement record in one operation."""
        if not self.current_pdf_path or self.pdf_doc is None:
            return
        if load_workbook is None:
            QMessageBox.critical(self, "Missing Library", "openpyxl is not installed.")
            return

        self._detect_all_remaining_pages()

        active = [d for d in self.dimensions if d.get("included")]
        if not active:
            QMessageBox.information(self, "Nothing to Export", "No callouts are currently included.")
            return

        base = os.path.splitext(os.path.basename(self.current_pdf_path))[0]
        dir_ = os.path.dirname(self.current_pdf_path)

        pdf_path, _ = QFileDialog.getSaveFileName(
            self, "Export Annotated PDF",
            os.path.join(dir_, f"{base}_annotated.pdf"),
            "PDF Files (*.pdf)",
        )
        if not pdf_path:
            return

        xlsx_path, _ = QFileDialog.getSaveFileName(
            self, "Export Measurement Record",
            os.path.join(dir_, f"{base}_measurement_record.xlsx"),
            "Excel Files (*.xlsx)",
        )
        if not xlsx_path:
            return

        active_sorted = sorted(active, key=lambda d: d.get("callout_num", 0))
        errors: list[str] = []

        try:
            self._write_annotated_pdf(pdf_path, active)
        except Exception as exc:
            errors.append(f"PDF: {exc}")

        try:
            self._write_measurement_record(xlsx_path, active_sorted)
        except Exception as exc:
            errors.append(f"Measurement record: {exc}")

        if errors:
            QMessageBox.critical(self, "Export Errors", "\n\n".join(errors))
            return

        self.status_bar.showMessage(
            f"Exported: {os.path.basename(pdf_path)}, {os.path.basename(xlsx_path)}"
        )
        reply = QMessageBox.information(
            self, "Export Complete",
            f"Both files exported:\n• {pdf_path}\n• {xlsx_path}\n\nReveal in Explorer?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._reveal_in_explorer(pdf_path)

    def _reveal_in_explorer(self, path: str) -> None:
        """Open Windows Explorer with the given file selected."""
        try:
            subprocess.Popen(f'explorer /select,"{os.path.normpath(path)}"', shell=True)
        except Exception:
            pass

    def _write_measurement_record(self, out_path: str, active_dims: list) -> None:
        with open(out_path, "wb") as fh:
            fh.write(base64.b64decode(_TEMPLATE_B64))

        wb = load_workbook(out_path, data_only=False)

        # Use the "Dimensions" sheet (first sheet with data entry rows)
        ws = None
        for name in wb.sheetnames:
            sheet = wb[name]
            if sheet.max_row >= _DATA_START_ROW:
                ws = sheet
                break
        if ws is None:
            ws = wb.active

        for i, dim in enumerate(active_dims):
            row_idx = _DATA_START_ROW + i
            nominal, plus_tol, minus_tol = _parse_tolerances(dim["text"])
            # Fall back to title-block general tolerances when the dimension
            # text carries no explicit tolerance annotation.
            if plus_tol is None:
                plus_tol  = dim.get("applied_plus_tol")
            if minus_tol is None:
                minus_tol = dim.get("applied_minus_tol")

            def safe_write(col: int, value: Any) -> None:
                if value is None:
                    return
                cell = ws.cell(row=row_idx, column=col)
                existing = cell.value
                if isinstance(existing, str) and existing.startswith("="):
                    return
                cell.value = value

            safe_write(_COL_DIM_NUM,   dim["callout_num"])
            safe_write(_COL_NOMINAL,   nominal)
            safe_write(_COL_PLUS_TOL,  plus_tol)
            safe_write(_COL_MINUS_TOL, minus_tol)

        wb.save(out_path)
        wb.close()


# ---------------------------------------------------------------------------
# Application entry point
# ---------------------------------------------------------------------------

def main() -> None:
    missing: list[str] = []
    if fitz is None:
        missing.append("PyMuPDF  (pip install pymupdf)")
    if load_workbook is None:
        missing.append("openpyxl  (pip install openpyxl)")

    app = QApplication(sys.argv)
    app.setApplicationName("FAI Dimension Numberer")
    app.setApplicationVersion("1.1.0")
    app.setStyle("Fusion")

    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window,          QColor("#F5F5F5"))
    palette.setColor(QPalette.ColorRole.WindowText,      QColor("#1A1A1A"))
    palette.setColor(QPalette.ColorRole.Base,            QColor("#FFFFFF"))
    palette.setColor(QPalette.ColorRole.AlternateBase,   QColor("#F0F0F0"))
    palette.setColor(QPalette.ColorRole.Text,            QColor("#1A1A1A"))
    palette.setColor(QPalette.ColorRole.Highlight,       QColor("#2980B9"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#FFFFFF"))
    app.setPalette(palette)

    if missing:
        QMessageBox.critical(
            None,
            "Missing Dependencies",
            "The following required packages are not installed:\n\n"
            + "\n".join(f"  • {m}" for m in missing)
            + "\n\nInstall them and relaunch the application.",
        )
        sys.exit(1)

    if _DETECTOR_IMPORT_ERROR:
        QMessageBox.critical(
            None,
            "Missing Module: dimension_detector.py",
            f"Could not import dimension_detector.py:\n\n{_DETECTOR_IMPORT_ERROR}\n\n"
            f"Ensure dimension_detector.py is in the same folder as this executable:\n"
            f"{_APP_DIR}",
        )
        sys.exit(1)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
