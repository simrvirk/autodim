"""
dimension_detector.py
=====================
Engineering drawing dimension detector and measurement template parser.

Functions
---------
detect_dimensions(pdf_path, page_num) -> list[dict]
    Extract text blocks from a drawing PDF and classify which ones are
    engineering dimensions using the Claude API.

parse_spreadsheet_template(template_path) -> dict
    Fully parse an Excel measurement record template, preserving its
    structure, formulas, formatting, and conditional-formatting rules
    so the template can later be populated with real dimension data.
"""

from __future__ import annotations

import json
import os
from typing import Any, Optional

import re

import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Bilateral-tolerance stack detection
# ---------------------------------------------------------------------------

# Unsigned decimal value — used to recognise tolerance magnitude fragments
# e.g. "0.20", "0.00", "1.5"
_PLAIN_DECIMAL_RE = re.compile(r'^\d+\.?\d*$')

# Fallback patterns for the less-common stacked-line layout:
#   25.00        ← nominal
#   +0.13        ← explicit plus arm  (starts with '+')
#   -0.08        ← explicit minus arm (starts with '-')
_PLUS_ARM_RE  = re.compile(r'^\+\d+\.?\d*$')
_MINUS_ARM_RE = re.compile(r'^-\d+\.?\d*$')
# Plain nominal that may have an Ø/R prefix but no trailing sign
_NOMINAL_RE   = re.compile(r'^[ØøRr⌀∅]?\s*\d+\.?\d*$')


def _merge_bilateral_tolerances(blocks: list[dict]) -> list[dict]:
    """Collapse bilateral-tolerance stacks into the nominal block.

    CAD software renders bilateral tolerances in (at least) two ways:

    Pattern A — inline split (seen in this codebase's target drawings):
        The nominal text itself ends with '+', the plus-tolerance value is a
        separate block glued immediately to its right on the same line, and the
        minus-tolerance value is a block directly below the plus value:

            [2X 1.70+] [0.20]      ← same y; abutting x
                        [0.00]      ← same x as [0.20]; fractionally lower y

        After merging: "2X 1.70 +0.20/-0.00"

    Pattern B — stacked lines (less common):
        The three values appear as separate lines below one another:

            25.00        ← nominal
            +0.13        ← explicit plus arm  (text starts with '+')
            -0.08        ← explicit minus arm (text starts with '-')

        After merging: "25.00 +0.13/-0.08"

    In both cases the nominal block's text is updated to the combined form and
    the absorbed tolerance fragments are removed from the returned list.  The
    combined text is already understood by _parse_tolerances() and
    _classify_dimension() so no other code needs to change.
    """
    if not blocks:
        return blocks

    # Spatial order: top-to-bottom, left-to-right
    order = sorted(range(len(blocks)),
                   key=lambda i: (blocks[i]["y"], blocks[i]["x"]))

    absorbed: set[int] = set()   # original block indices to drop

    # ── Pattern A: nominal text ends with '+' ─────────────────────────────────
    for si, bi in enumerate(order):
        if bi in absorbed:
            continue
        nom_text_raw = blocks[bi]["text"].strip()
        if not nom_text_raw.endswith("+"):
            continue

        nom_block  = blocks[bi]
        nom_end_x  = nom_block["x"] + nom_block["width"]   # right edge of nominal

        # Find the plus-value block: same y row, x starts where the nominal ends
        plus_bi: Optional[int] = None
        for look in range(si + 1, min(si + 12, len(order))):
            cbi = order[look]
            if cbi in absorbed:
                continue
            c = blocks[cbi]
            # Must be on approximately the same baseline
            if abs(c["y"] - nom_block["y"]) > max(nom_block["height"], c["height"]):
                continue
            # x must start within a small gap of where the nominal text ends
            # (allow up to 8 pt gap, and up to 4 pt overlap for rounding)
            x_gap = c["x"] - nom_end_x
            if not (-4 <= x_gap <= 8):
                continue
            if _PLAIN_DECIMAL_RE.match(c["text"].strip()):
                plus_bi = cbi
                break

        if plus_bi is None:
            continue

        plus_block = blocks[plus_bi]

        # Find the minus-value block: same x range as plus_block, directly below
        minus_bi: Optional[int] = None
        for look in range(si + 1, min(si + 15, len(order))):
            cbi = order[look]
            if cbi in absorbed or cbi == plus_bi:
                continue
            c = blocks[cbi]
            # Must be strictly below the plus block's top edge
            if c["y"] <= plus_block["y"]:
                continue
            # Must not be too far below (within 4× the plus block's height,
            # which also handles the overlapping-fraction layout)
            if c["y"] - plus_block["y"] > plus_block["height"] * 4:
                break
            # Must overlap horizontally with the plus block
            c_x0, c_x1  = c["x"], c["x"] + c["width"]
            p_x0, p_x1  = plus_block["x"], plus_block["x"] + plus_block["width"]
            overlap = min(c_x1, p_x1) - max(c_x0, p_x0)
            if overlap < min(p_x1 - p_x0, c_x1 - c_x0) * 0.25:
                continue
            if _PLAIN_DECIMAL_RE.match(c["text"].strip()):
                minus_bi = cbi
                break

        # Build combined text: strip trailing '+' from nominal, append tols
        nom_base  = nom_text_raw[:-1].strip()          # "2X 1.70+" → "2X 1.70"
        plus_val  = plus_block["text"].strip()
        minus_val = blocks[minus_bi]["text"].strip() if minus_bi is not None else plus_val

        nom_block["text"] = f"{nom_base} +{plus_val}/-{minus_val}"
        absorbed.add(plus_bi)
        if minus_bi is not None:
            absorbed.add(minus_bi)

    # ── Pattern B: explicit "+X.XX" line below nominal ────────────────────────
    for si, bi in enumerate(order):
        if bi in absorbed:
            continue
        text = blocks[bi]["text"].strip()
        if not _PLUS_ARM_RE.match(text):
            continue

        plus_block = blocks[bi]

        # Search upward (in spatial order) for a plain nominal
        nom_bi: Optional[int] = None
        for look in range(si - 1, max(si - 10, -1), -1):
            cbi = order[look]
            if cbi in absorbed:
                continue
            c = blocks[cbi]
            if c["y"] >= plus_block["y"]:
                continue
            v_gap = plus_block["y"] - (c["y"] + c["height"])
            if v_gap > plus_block["height"] * 4:
                break
            c_x0, c_x1 = c["x"], c["x"] + c["width"]
            p_x0, p_x1 = plus_block["x"], plus_block["x"] + plus_block["width"]
            if min(c_x1, p_x1) - max(c_x0, p_x0) < min(p_x1-p_x0, c_x1-c_x0) * 0.25:
                continue
            if _NOMINAL_RE.match(c["text"].strip()):
                nom_bi = cbi
                break

        if nom_bi is None:
            continue

        # Search downward for the explicit minus arm
        minus_bi = None
        for look in range(si + 1, min(si + 10, len(order))):
            cbi = order[look]
            if cbi in absorbed:
                continue
            c = blocks[cbi]
            if c["y"] <= plus_block["y"]:
                continue
            if c["y"] - (plus_block["y"] + plus_block["height"]) > plus_block["height"] * 4:
                break
            c_x0, c_x1 = c["x"], c["x"] + c["width"]
            p_x0, p_x1 = plus_block["x"], plus_block["x"] + plus_block["width"]
            if min(c_x1, p_x1) - max(c_x0, p_x0) < min(p_x1-p_x0, c_x1-c_x0) * 0.25:
                continue
            if _MINUS_ARM_RE.match(c["text"].strip()):
                minus_bi = cbi
                break

        plus_val = text.lstrip("+")
        nom = blocks[nom_bi]
        if minus_bi is not None:
            minus_val = blocks[minus_bi]["text"].strip().lstrip("-")
            nom["text"] = f"{nom['text'].strip()} +{plus_val}/-{minus_val}"
            absorbed.add(minus_bi)
        else:
            nom["text"] = f"{nom['text'].strip()} +{plus_val}/-{plus_val}"
        absorbed.add(bi)

    return [b for i, b in enumerate(blocks) if i not in absorbed]


# ---------------------------------------------------------------------------
# Regex-based dimension classifier
# ---------------------------------------------------------------------------

def _classify_dimension(text: str) -> tuple[Optional[str], float]:
    """Return (category, confidence) if text looks like an engineering dimension.

    Returns (None, 0.0) for text that is not a dimension.
    Confidence is 0.0–1.0; only items with confidence > 0 are returned.
    """
    t = text.strip()

    # Reference dimensions in parentheses are for information only — skip them
    if re.match(r'^\(.*\)$', t):
        return None, 0.0

    # Reject obviously non-dimension text
    if not t or len(t) > 80:
        return None, 0.0
    word_count = len(t.split())
    if word_count > 8:
        return None, 0.0
    # Skip plain revision letters / single uppercase words (not datum context)
    if re.match(r'^[A-Z]{2,}$', t) and not re.search(r'\d', t):
        return None, 0.0

    # --- Standalone bilateral-tolerance arm (e.g. "+0.13" or "-0.08") -------
    # These are the upper/lower tolerance lines typeset below a nominal value.
    # _merge_bilateral_tolerances() folds them into the nominal before this
    # function is called; any that still reach here are not standalone dims.
    if re.match(r'^[+-]\d+\.?\d*$', t):
        return None, 0.0

    # --- Angular ---
    if re.search(r'\d+\.?\d*\s*°', t):
        return "angular", 0.92

    # --- Diameter ---
    if re.search(r'[Øø⌀∅]\s*\d+', t):
        return "diameter", 0.95
    if re.search(r'\bDIA\.?\s*\d+', t, re.IGNORECASE):
        return "diameter", 0.90

    # --- Radius ---
    if re.match(r'^[Rr]\s*\d+\.?\d*', t):
        return "radius", 0.92
    if re.search(r'\bCR\s*\d+\.?\d*', t, re.IGNORECASE):
        return "radius", 0.88

    # --- Thread / screw-hole callouts — intentionally ignored ---
    if re.match(r'^[Mm]\d+[xX×]\d+', t):
        return None, 0.0
    if re.search(r'\d+-\d+\s*(UNC|UNF|UNEF|NPT|NPS)\b', t, re.IGNORECASE):
        return None, 0.0
    if re.match(r'^#\d+-\d+', t):
        return None, 0.0

    # --- Surface finish ---
    if re.search(r'\b[Rr][aAzZqQwW]\s*\d+', t):
        return "finish", 0.92
    if re.search(r'\d+\.?\d*\s*RMS\b', t, re.IGNORECASE):
        return "finish", 0.90
    if re.match(r'^N\d{1,2}$', t):
        return "finish", 0.78

    # --- GD&T symbols and keywords — intentionally ignored ---
    if re.search(r'[⊕⊙⌖⊞⊟◎⌯⌦⌧]', t):
        return None, 0.0
    if re.search(
        r'\b(TRUE\s*POS|T\.P\.|FLATNESS|PERPENDICULARITY|CONCENTRICITY'
        r'|RUNOUT|CIRCULARITY|CYLINDRICITY|PARALLELISM|STRAIGHTNESS'
        r'|PROFILE|SYMMETRY)\b', t, re.IGNORECASE
    ):
        return None, 0.0

    # --- Hole / feature callouts — intentionally ignored ---
    if re.search(r'\b(C\'?BORE|CBORE|CSINK|SPOTFACE|THRU)\b', t, re.IGNORECASE):
        return None, 0.0

    # --- Chamfer / taper ---
    if re.search(r'\d+\s*[Xx×]\s*\d+\.?\d*\s*°', t):
        return "angular", 0.88
    if re.search(r'\bTAPER\b', t, re.IGNORECASE):
        return "linear", 0.80

    # --- Tolerance-only blocks (e.g. "±0.005") ---
    if re.match(r'^[±]\s*\d+\.?\d*$', t):
        return "linear", 0.88

    # --- Linear with explicit tolerance ---
    if re.search(r'\d+\.?\d*\s*[±]\s*\d+\.?\d*', t):
        return "linear", 0.92
    if re.search(r'\d+\.?\d*\s*\+\d+\.?\d*\s*/?\s*-\d+\.?\d*', t):
        return "linear", 0.92

    # --- Plain decimal number (likely a dimension if it has a decimal point) ---
    m = re.match(r'^([+-]?\d{1,5}\.\d{1,5})\s*$', t)
    if m:
        val = abs(float(m.group(1)))
        if 0.0001 <= val <= 99999:
            return "linear", 0.72

    return None, 0.0


# ---------------------------------------------------------------------------
# GD&T feature-control-frame detection (vector-path based)
# ---------------------------------------------------------------------------

def _collect_gdt_cells(page) -> list[tuple[float, float, float, float]]:
    """Return (x0, y0, x1, y1) for every small rectangular vector path that
    is likely a GD&T feature-control-frame symbol cell.

    In a CAD-exported PDF the GD&T symbol (flatness, cylindricity, etc.) is
    drawn as vector graphics inside a small rectangular cell.  The tolerance
    value (e.g. "0.2") is a separate text block sitting immediately to the
    right of that cell.

    The distinguishing geometry of a GD&T symbol cell vs a dimension-line
    arrowhead (the other common small rectangle in engineering drawings):

      • GD&T cell   — wide and flat: width ≈ 12–35 pt, height ≈ 5–15 pt,
                      aspect ratio (w/h) typically > 1.5
      • Arrowhead   — roughly square or taller: width ≈ height ≈ 7–10 pt,
                      aspect ratio ≈ 0.8–1.1

    We therefore require:
      • unfilled (fill is None)
      • width  10 – 40 pt
      • height  4 – 20 pt
      • width  > height × 1.3   (explicitly wider than tall → not an arrowhead)
    """
    cells: list[tuple[float, float, float, float]] = []
    try:
        for path in page.get_drawings():
            r = path.get("rect")
            if r is None:
                continue
            w = r.x1 - r.x0
            h = r.y1 - r.y0
            if (path.get("fill") is None
                    and path.get("color") is not None
                    and 10 <= w < 40          # wide enough for a symbol glyph
                    and 4 < h < 20            # one text-line tall
                    and w > h * 1.3):         # wider than tall → not an arrowhead
                cells.append((r.x0, r.y0, r.x1, r.y1))
    except Exception:
        pass   # get_drawings() unavailable on this PDF variant
    return cells


def _is_gdt_value(block: dict,
                  cells: list[tuple[float, float, float, float]]) -> bool:
    """Return True if *block* sits immediately to the right of a GD&T cell.

    Criteria:
      • The block's left edge is 0–15 pt to the right of the cell's right edge.
      • The block and the cell share at least 30 % of the shorter one's height.
    """
    bx0 = block["x"]
    by0 = block["y"]
    by1 = by0 + block["height"]
    for cx0, cy0, cx1, cy1 in cells:
        x_gap = bx0 - cx1
        if not (0 <= x_gap <= 15):
            continue
        overlap = min(by1, cy1) - max(by0, cy0)
        if overlap >= min(by1 - by0, cy1 - cy0) * 0.30:
            return True
    return False


# ---------------------------------------------------------------------------
# Public API — PDF dimension detection
# ---------------------------------------------------------------------------

def detect_dimensions(pdf_path: str, page_num: int = 0) -> list[dict]:
    """Detect engineering dimensions in a single page of a drawing PDF.

    Parameters
    ----------
    pdf_path : str
        Absolute or relative path to the PDF file.
    page_num : int
        Zero-based page index. Defaults to 0 (first page).

    Returns
    -------
    list[dict]
        One dict per detected dimension:
        {
            "index"     : int,    # position in the extracted block list
            "text"      : str,    # raw text string from the PDF
            "x"         : float,  # left edge of bounding box (points)
            "y"         : float,  # top edge of bounding box (points)
            "width"     : float,  # bounding box width (points)
            "height"    : float,  # bounding box height (points)
            "confidence": float,  # 0–1 from Claude
            "category"  : str,    # dimension category label
            "included"  : True    # always True for returned items
        }

    Raises
    ------
    FileNotFoundError
        If pdf_path does not point to an existing file.
    ValueError
        If page_num is out of range for the document.
    RuntimeError
        If the page contains no extractable text (scanned PDF) or if the
        Claude API returns an unexpected payload.
    """
    # ------------------------------------------------------------------
    # 1. Validate inputs
    # ------------------------------------------------------------------
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path!r}")

    doc = fitz.open(pdf_path)
    total_pages = len(doc)

    if page_num < 0 or page_num >= total_pages:
        doc.close()
        raise ValueError(
            f"page_num {page_num} is out of range for a {total_pages}-page document."
        )

    # ------------------------------------------------------------------
    # 2. Extract individual text lines with bounding boxes via PyMuPDF
    # ------------------------------------------------------------------
    page = doc[page_num]

    # Use get_text("dict") instead of get_text("blocks") so that each *line*
    # gets its own bounding box.  The "blocks" API merges adjacent lines into
    # one block, causing dimension strings like "11.00 +/- 0.2", "2X", "6.00"
    # to be concatenated into a single entry and mis-classified.
    raw_dict = page.get_text("dict")

    # Collect GD&T symbol cells *before* closing the doc — get_drawings()
    # requires the page object to still be alive.
    gdt_cells = _collect_gdt_cells(page)

    doc.close()

    text_blocks: list[dict] = []
    for block in raw_dict.get("blocks", []):
        if block.get("type") != 0:
            continue  # skip image blocks
        for line in block.get("lines", []):
            # Concatenate all spans on the line into one string
            line_text = "".join(
                span["text"] for span in line.get("spans", [])
            ).strip()
            if not line_text:
                continue
            bbox = line["bbox"]   # (x0, y0, x1, y1)
            text_blocks.append({
                "index" : len(text_blocks),
                "text"  : line_text,
                "x"     : round(bbox[0], 3),
                "y"     : round(bbox[1], 3),
                "width" : round(bbox[2] - bbox[0], 3),
                "height": round(bbox[3] - bbox[1], 3),
            })

    if not text_blocks:
        raise RuntimeError(
            "No text could be extracted from page "
            f"{page_num} of '{pdf_path}'. "
            "This PDF may be a scanned image. Run OCR (e.g. ocrmypdf) on it "
            "first, or use a vector-exported drawing PDF."
        )

    # ------------------------------------------------------------------
    # 3. Merge bilateral-tolerance stacks before classification.
    #
    # Drawing CAD software often typsets a nominal and its bilateral tolerance
    # as three separate text lines stacked vertically:
    #
    #     25.00          ← nominal
    #     +0.13          ← upper tolerance arm
    #     -0.08          ← lower tolerance arm
    #
    # Without this step the "+0.13" line is mis-classified as a dimension.
    # The merge folds the arms into the nominal ("25.00 +0.13/-0.08") and
    # removes the standalone tolerance lines from the block list.
    # ------------------------------------------------------------------
    text_blocks = _merge_bilateral_tolerances(text_blocks)

    # ------------------------------------------------------------------
    # 4. Classify each text block with regex patterns
    # ------------------------------------------------------------------
    results: list[dict] = []
    for block in text_blocks:
        category, confidence = _classify_dimension(block["text"])
        if category is None:
            continue
        # Skip values that live inside a GD&T feature control frame.
        # The frame's symbol cell is a small rectangular vector path
        # immediately to the left of the tolerance value text.
        if gdt_cells and _is_gdt_value(block, gdt_cells):
            continue
        results.append({
            "index"     : block["index"],
            "text"      : block["text"],
            "x"         : block["x"],
            "y"         : block["y"],
            "width"     : block["width"],
            "height"    : block["height"],
            "confidence": confidence,
            "category"  : category,
            "included"  : True,
        })

    # Return in document order (ascending y then x — top-to-bottom, left-to-right)
    results.sort(key=lambda d: (d["y"], d["x"]))
    return results


# ---------------------------------------------------------------------------
# Public API — Excel template parser
# ---------------------------------------------------------------------------

def parse_spreadsheet_template(template_path: str) -> dict:
    """Fully parse an Excel measurement record template (.xlsx).

    Reads the workbook without evaluating formulas (so formula strings are
    preserved) and extracts every structural and formatting detail needed to
    later populate a copy of the template with real dimension data.

    Parameters
    ----------
    template_path : str
        Path to the .xlsx template file.

    Returns
    -------
    dict
        Top-level keys:
        ├─ "file_path"        : str
        ├─ "sheet_names"      : list[str]
        └─ "sheets"           : dict[sheet_name -> sheet_info]

        Each sheet_info dict contains:
        ├─ "dimensions"          : str   (e.g. "A1:Z100")
        ├─ "max_row"             : int
        ├─ "max_col"             : int
        ├─ "freeze_panes"        : str | None
        ├─ "column_headers"      : list[{"column_letter", "column_index",
        │                                "row", "value"}]
        ├─ "formulas"            : list[{"cell", "formula"}]
        ├─ "merged_cells"        : list[{"range", "min_row", "max_row",
        │                                "min_col", "max_col"}]
        ├─ "conditional_formats" : list[{"range", "rules"}]
        ├─ "column_widths"       : dict[column_letter -> width]
        ├─ "row_heights"         : dict[row_number -> height]
        ├─ "cell_styles"         : dict[cell_address -> style_dict]
        └─ "data_entry_rows"     : list[int]   (rows that appear to be for data)

    Raises
    ------
    FileNotFoundError
        If template_path does not exist.
    ValueError
        If the file is not a recognisable .xlsx workbook.
    """
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Template not found: {template_path!r}")
    if not template_path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise ValueError(
            f"Expected an Excel workbook (.xlsx/.xlsm), got: {template_path!r}"
        )

    # Load twice:
    #   • formula_wb  — formulas as strings (data_only=False, default)
    #   • value_wb    — computed cell values (data_only=True) for style inspection
    #   We use formula_wb as the primary source; value_wb only for reading
    #   cached computed values where useful (not saved over, so no data loss).
    try:
        formula_wb = load_workbook(template_path, data_only=False)
    except Exception as exc:
        raise ValueError(f"Could not open workbook: {exc}") from exc

    result: dict[str, Any] = {
        "file_path"  : os.path.abspath(template_path),
        "sheet_names": formula_wb.sheetnames,
        "sheets"     : {},
    }

    for sheet_name in formula_wb.sheetnames:
        ws = formula_wb[sheet_name]
        sheet_info: dict[str, Any] = {}

        # ------------------------------------------------------------------
        # Basic dimensions
        # ------------------------------------------------------------------
        sheet_info["dimensions"]   = ws.dimensions
        sheet_info["max_row"]      = ws.max_row
        sheet_info["max_col"]      = ws.max_column
        sheet_info["freeze_panes"] = (
            str(ws.freeze_panes) if ws.freeze_panes else None
        )

        # ------------------------------------------------------------------
        # Column headers
        # Heuristic: scan the first 5 rows for cells whose row is a header.
        # We identify a header row as one where most occupied cells are
        # non-numeric strings.  We collect every cell that has a string value
        # in that row.
        # ------------------------------------------------------------------
        headers: list[dict] = []
        header_rows_found: set[int] = set()

        for row_idx in range(1, min(6, (ws.max_row or 0) + 1)):
            row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx,
                                          values_only=False))[0]
            str_count   = sum(1 for c in row_cells
                              if c.value is not None
                              and isinstance(c.value, str)
                              and not str(c.value).startswith("="))
            total_count = sum(1 for c in row_cells if c.value is not None)
            if total_count > 0 and str_count / total_count >= 0.5:
                header_rows_found.add(row_idx)
                for cell in row_cells:
                    if cell.value is not None:
                        headers.append({
                            "column_letter": get_column_letter(cell.column),
                            "column_index" : cell.column,
                            "row"          : cell.row,
                            "value"        : cell.value,
                        })

        sheet_info["column_headers"] = headers

        # ------------------------------------------------------------------
        # Formulas — every cell whose value starts with "="
        # ------------------------------------------------------------------
        formulas: list[dict] = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    formulas.append({
                        "cell"   : cell.coordinate,
                        "formula": val,
                    })
        sheet_info["formulas"] = formulas

        # ------------------------------------------------------------------
        # Merged cells
        # ------------------------------------------------------------------
        merged: list[dict] = []
        for rng in ws.merged_cells.ranges:
            merged.append({
                "range"  : str(rng),
                "min_row": rng.min_row,
                "max_row": rng.max_row,
                "min_col": rng.min_col,
                "max_col": rng.max_col,
            })
        sheet_info["merged_cells"] = merged

        # ------------------------------------------------------------------
        # Conditional formatting rules
        # openpyxl stores them as ConditionalFormattingList.
        # Each entry: (sqref, [Rule, ...])
        # ------------------------------------------------------------------
        cf_list: list[dict] = []
        for sqref, rules in ws.conditional_formatting._cf_rules.items():
            serialised_rules = []
            for rule in rules:
                rule_dict: dict[str, Any] = {
                    "type"    : rule.type,
                    "priority": rule.priority,
                    "operator": getattr(rule, "operator", None),
                    "formula" : (list(rule.formula)
                                 if getattr(rule, "formula", None)
                                 else None),
                }
                # Differential style (fill/font/border applied when rule fires)
                dxf = getattr(rule, "dxf", None)
                if dxf is not None:
                    dxf_info: dict[str, Any] = {}
                    if dxf.fill and dxf.fill.fgColor:
                        dxf_info["fill_fg_color"] = dxf.fill.fgColor.rgb
                    if dxf.font:
                        dxf_info["font_bold"]  = dxf.font.bold
                        dxf_info["font_color"] = (
                            dxf.font.color.rgb if dxf.font.color else None
                        )
                    if dxf.border:
                        dxf_info["has_border"] = True
                    rule_dict["dxf"] = dxf_info

                # Color scale / data bar / icon set specifics
                for attr in ("colorScale", "dataBar", "iconSet"):
                    if getattr(rule, attr, None) is not None:
                        rule_dict[attr] = str(getattr(rule, attr))

                serialised_rules.append(rule_dict)

            cf_list.append({
                "range": str(sqref),
                "rules": serialised_rules,
            })
        sheet_info["conditional_formats"] = cf_list

        # ------------------------------------------------------------------
        # Column widths
        # ------------------------------------------------------------------
        col_widths: dict[str, float | None] = {}
        for col_letter, col_dim in ws.column_dimensions.items():
            col_widths[col_letter] = col_dim.width
        sheet_info["column_widths"] = col_widths

        # ------------------------------------------------------------------
        # Row heights
        # ------------------------------------------------------------------
        row_heights: dict[int, float | None] = {}
        for row_idx, row_dim in ws.row_dimensions.items():
            row_heights[row_idx] = row_dim.height
        sheet_info["row_heights"] = row_heights

        # ------------------------------------------------------------------
        # Cell styles — serialise only non-default cells to keep the dict lean
        # ------------------------------------------------------------------
        cell_styles: dict[str, dict] = {}
        for row in ws.iter_rows():
            for cell in row:
                style = _serialise_cell_style(cell)
                if style:
                    cell_styles[cell.coordinate] = style
        sheet_info["cell_styles"] = cell_styles

        # ------------------------------------------------------------------
        # Data-entry rows heuristic
        # A row is a "data entry" row if:
        #   • it is below all detected header rows
        #   • it has at least one empty cell in a column that has a header
        #   • it is not entirely empty (it has some structure/formula)
        # ------------------------------------------------------------------
        header_row_max = max(header_rows_found) if header_rows_found else 1
        header_col_indices = {h["column_index"] for h in headers}

        data_entry_rows: list[int] = []
        for row_idx in range(header_row_max + 1, (ws.max_row or 0) + 1):
            row_cells = {
                c.column: c
                for c in ws[row_idx]
                if c.column in header_col_indices
            }
            has_formula = any(
                isinstance(c.value, str) and c.value.startswith("=")
                for c in row_cells.values()
            )
            has_empty = any(
                c.value is None for c in row_cells.values()
            )
            # Rows with a mix of formulas and empty cells are data-entry rows
            if has_empty and (has_formula or _row_has_style(ws, row_idx)):
                data_entry_rows.append(row_idx)

        sheet_info["data_entry_rows"] = data_entry_rows

        result["sheets"][sheet_name] = sheet_info

    formula_wb.close()
    return result


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _serialise_cell_style(cell) -> dict:
    """Return a dict of non-default style attributes for a cell.

    Returns an empty dict if the cell has no meaningful styling, so callers
    can skip storing it.
    """
    style: dict[str, Any] = {}

    # Font
    f = cell.font
    if f:
        font_info: dict[str, Any] = {}
        if f.bold:
            font_info["bold"] = True
        if f.italic:
            font_info["italic"] = True
        if f.underline:
            font_info["underline"] = f.underline
        if f.size and f.size != 11:
            font_info["size"] = f.size
        if f.name and f.name not in ("Calibri", "Arial", ""):
            font_info["name"] = f.name
        if f.color and f.color.type == "rgb" and f.color.rgb not in (
            "00000000", "FF000000"
        ):
            font_info["color"] = f.color.rgb
        if font_info:
            style["font"] = font_info

    # Fill
    fl = cell.fill
    if fl and fl.fill_type and fl.fill_type != "none":
        fill_info: dict[str, Any] = {"fill_type": fl.fill_type}
        try:
            if fl.fgColor and fl.fgColor.type == "rgb":
                fill_info["fg_color"] = fl.fgColor.rgb
            if fl.bgColor and fl.bgColor.type == "rgb":
                fill_info["bg_color"] = fl.bgColor.rgb
        except Exception:
            pass
        style["fill"] = fill_info

    # Alignment
    al = cell.alignment
    if al:
        align_info: dict[str, Any] = {}
        if al.horizontal and al.horizontal != "general":
            align_info["horizontal"] = al.horizontal
        if al.vertical and al.vertical != "bottom":
            align_info["vertical"] = al.vertical
        if al.wrap_text:
            align_info["wrap_text"] = True
        if align_info:
            style["alignment"] = align_info

    # Number format
    nf = cell.number_format
    if nf and nf not in ("General", "@", ""):
        style["number_format"] = nf

    # Border — just flag its presence; full border serialisation is verbose
    bd = cell.border
    if bd:
        sides = [bd.left, bd.right, bd.top, bd.bottom]
        if any(s and s.border_style for s in sides):
            style["has_border"] = True

    return style


def _row_has_style(ws, row_idx: int) -> bool:
    """Return True if any cell in the row has non-default styling."""
    for cell in ws[row_idx]:
        if _serialise_cell_style(cell):
            return True
    return False
